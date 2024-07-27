
import sys
import os
from dotenv import load_dotenv
from datetime import datetime, timedelta
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import numbers
import json

try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side




# Add the path to the parent directory of "washify" to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'washify')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'sitewash')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'hamilton')))

# print(sys.path)

from washify_weekly import get_week_dates as washify_week_dates
from washify_weekly import generate_weekly_report  as washify_week_report

from sitewatch_weekly import get_week_dates as sitewatch_week_dates
from sitewatch_weekly import generate_weekly_report as sitewatch_week_report

from hamilton_weekly  import get_week_dates as hamilton_week_dates
from hamilton  import generate_report as hamilton_week_report

from custom_mailer import send_email,get_excel_files

current_file_path = os.path.dirname(os.path.abspath(__file__))
# print(current_file_path)

data_path = os.path.join(current_file_path,"data")

def do_sum(xl_map,start_index,range):
    "This will do row based some "
    total=0
    for i in range:
        val = xl_map[start_index][i+1]
        if isinstance(val,float) or isinstance(val,int):
            total+=val
    
    return total

def do_sum_location(xl_map,location:list):
    "This will take array of row ,col"
    total =0
    for row,col in location:
        val = xl_map[row][col]
        if isinstance(val,float) or isinstance(val,int):
            total+=val
    return total



def do_avg(val):
    "return  the average of the value"
    result = 0
    if val:
        result = val/4
    return result
    
def Average_retail_visit__GA_SC_fucntion(retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC,
                                         retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC):    
    
    result = 0
    try:
        Average_retail_visit__GA_SC = sum([retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC])/sum(
        [retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC ]
    )
        result = Average_retail_visit__GA_SC
    except Exception as e:
        print(f"Exception in Average_retail_visit__GA_SC_fucntion() {e}")
        

    return result 

def Average_memeber_visit_GA_SC_function(Total_revenue_GA_SC,
        retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC,
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC,
        total_cars_GA_SC
        ):
    result = 0
    
    try:
        
        Average_memeber_visit_GA_SC = (Total_revenue_GA_SC - sum([retail_revenue_monday_to_friday_GA_SC , retail_revenue_saturday_to_sunday_GA_SC]))/(total_cars_GA_SC - sum([
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC]))
        
        result = Average_memeber_visit_GA_SC
        
    except Exception as e:
        print(f"Exception in Average_memeber_visit_GA_SC_function() {e}")
    
    return result

def cost_per_labour_hour_monday_to_friday_GA_SC_fucntion(car_count_monday_to_friday_GA_SC,staff_hours_monday_to_friday_GA_SC):
    result = 0
    try:
        cost_per_labour_hour_monday_to_friday_GA_SC = car_count_monday_to_friday_GA_SC/staff_hours_monday_to_friday_GA_SC
        result =cost_per_labour_hour_monday_to_friday_GA_SC
    except Exception as e:
        print(f"Exception in cost_per_labour_hour_monday_to_friday_GA_SC() {e}")

    return result


def cost_per_labour_hour_saturday_to_sunday_GA_SC_function(car_count_saturday_to_sunday_GA_SC,staff_hours_saturday_to_sunday_GA_SC):
    result = 0
    
    try:
        cost_per_labour_hour_saturday_to_sunday_GA_SC = car_count_saturday_to_sunday_GA_SC/staff_hours_saturday_to_sunday_GA_SC
        result = cost_per_labour_hour_saturday_to_sunday_GA_SC
    except Exception as e:
        print(f"Exception in staff_hours_saturday_to_sunday_GA_SC_fucntion() {e}")
        
    return result

def Total_cars_per_man_hour_GA_SC_function(total_cars_GA_SC, staff_hours_monday_to_friday_GA_SC,staff_hours_saturday_to_sunday_GA_SC):
    result =0 
    try:
            Total_cars_per_man_hour_GA_SC = total_cars_GA_SC/sum([
            staff_hours_monday_to_friday_GA_SC,staff_hours_saturday_to_sunday_GA_SC])
            result = Total_cars_per_man_hour_GA_SC
    
    except Exception as e:
        print(f"Exception in  Total_cars_per_man_hour_GA_SC_function() {e}")
    
    return result
 
def Conversion_rate_GA_SC_function(Total_club_plans_sold_GA_SC,retail_car_count_monday_to_friday_GA_SC,
                                   retail_car_count_saturday_to_sunday_GA_SC):
    result = 0 
    try:
        Conversion_rate_GA_SC = Total_club_plans_sold_GA_SC/sum([
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC])     
        
        result =  Conversion_rate_GA_SC
    except Exception as e:
        print(f"Exceptionin Conversion_rate_GA_SC_function() {e}")
    
    return result

def Conversion_rate_Total_function(Total_club_plans_sold_Total,
                                   retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total):
    result = 0
    
    try:
        
        Conversion_rate_Total = Total_club_plans_sold_Total/sum(
        [retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total])  
        
        result = Conversion_rate_Total
    except Exception as e:
        print(f"Exception in Conversion_rate_Total_function()  {e}") 
    
    return result

def Conversion_rate_ILL_function(Total_club_plans_sold_ILL,
                                 retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Conversion_rate_ILL = Total_club_plans_sold_ILL/sum([retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL])         
        result = Conversion_rate_ILL
    except Exception as e:
        print(f"Exception in Conversion_rate_ILL_function() {e}")
        
    return result

def Total_cars_per_man_hour_total_function(Total_cars_Total,
                                           staff_hours_monday_to_friday_Total,
                                           staff_hours_saturday_to_sunday_Total):
    result = 0
    
    try:
        Total_cars_per_man_hour_total = Total_cars_Total/sum(
        [staff_hours_monday_to_friday_Total,staff_hours_saturday_to_sunday_Total])
        
        result = Total_cars_per_man_hour_total
    
    except Exception as e:
        print(f"Exception in Total_cars_per_man_hour_total_function() {e}")
        
    return result


def Total_cars_per_man_hour_ILL_function(total_cars_in_ILL,
                                         staff_hours_monday_to_friday_ILL,
                                         staff_hours_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Total_cars_per_man_hour_ILL = total_cars_in_ILL/(sum([
        staff_hours_monday_to_friday_ILL+staff_hours_saturday_to_sunday_ILL]))

        result = Total_cars_per_man_hour_ILL
    
    except Exception as e:
        print(f"Exception Total_cars_per_man_hour_ILL_function() {e}")
    
    return result


def cost_per_labour_hour_saturday_to_sunday_ILL_function(car_count_saturday_to_sunday_ILL,staff_hours_saturday_to_sunday_ILL):
    result = 0
    
    try:
        cost_per_labour_hour_saturday_to_sunday_ILL = car_count_saturday_to_sunday_ILL/staff_hours_saturday_to_sunday_ILL
        
        result = cost_per_labour_hour_saturday_to_sunday_ILL
    
    except Exception as e:
        print(f"Exception cost_per_labour_hour_saturday_to_sunday_ILL_function() {e}")
        
    return result


def cost_per_labour_hour_monday_to_friday_Total_function(car_count_monday_to_friday_Total,staff_hours_monday_to_friday_Total):
    result = 0
    
    try:
        cost_per_labour_hour_monday_to_friday_Total = car_count_monday_to_friday_Total/staff_hours_monday_to_friday_Total
        result = cost_per_labour_hour_monday_to_friday_Total
    
    except Exception as e:
        print(f"Exception cost_per_labour_hour_monday_to_friday_Total_function() {e}")
    
    return result

def cost_per_labour_hour_monday_to_friday_ILL_function(car_count_monday_to_friday_ILL,staff_hours_monday_to_friday_ILL):
    result = 0
    
    try:
        cost_per_labour_hour_monday_to_friday_ILL = car_count_monday_to_friday_ILL/staff_hours_monday_to_friday_ILL
        result =  cost_per_labour_hour_monday_to_friday_ILL
    
    except Exception as e:
        print(f"Exception cost_per_labour_hour_monday_to_friday_ILL_function() {e}")
        
    return result

def Average_memeber_visit_Total_function(Total_revenue_Total,retail_revenue_monday_to_friday_Total
                        ,retail_revenue_saturday_to_sunday_Total,Total_cars_Total,retail_car_count_monday_to_friday_Total,
                        retail_car_count_saturday_to_sunday_Total):
    result = 0
    
    try:
        Average_memeber_visit_Total = (Total_revenue_Total -sum([retail_revenue_monday_to_friday_Total,retail_revenue_saturday_to_sunday_Total]))/(Total_cars_Total - sum([
        retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total]))
        result =  Average_memeber_visit_Total
    
    except Exception as e:
        print(f"Exception Average_memeber_visit_Total_function() {e}")  
    
    return result    

def Average_memeber_visit_ILL_function(Total_revenue_ILL,retail_revenue_monday_to_friday_ILL,
                                       retail_revenue_saturday_to_sunday_ILL,total_cars_in_ILL,
                                       retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Average_memeber_visit_ILL = (Total_revenue_ILL - sum([retail_revenue_monday_to_friday_ILL,retail_revenue_saturday_to_sunday_ILL]))/(total_cars_in_ILL - sum([
        retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL]))
        
        result = Average_memeber_visit_ILL
    
    except Exception as e:
        print(f"Exception Average_memeber_visit_ILL_function() {e}")
    
    return result

def Average_retail_visit_Total_function(retail_revenue_monday_to_friday_Total,
                                        retail_revenue_saturday_to_sunday_Total,
                                        retail_car_count_monday_to_friday_Total,
                                        retail_car_count_saturday_to_sunday_Total):
    result = 0
    
    try:
        Average_retail_visit_Total = sum([retail_revenue_monday_to_friday_Total,retail_revenue_saturday_to_sunday_Total])/sum([
        retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total])
        
        result = Average_retail_visit_Total
    
    except Exception as e:
        print(f"Exception Average_retail_visit_Total_function() {e}")
        
    return result
            

def Average_retail_visit_IL_function(retail_revenue_monday_to_friday_ILL,
                                     retail_revenue_saturday_to_sunday_ILL,
                                     retail_car_count_monday_to_friday_ILL,
                                     retail_car_count_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Average_retail_visit_IL = sum([retail_revenue_monday_to_friday_ILL,retail_revenue_saturday_to_sunday_ILL])/sum(
        [retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL])

        result = Average_retail_visit_IL
    
    except Exception as e:
        print(f"Exception Average_retail_visit_IL_function() {e}")
    
    return result

def add_commas(value):
    # Check if the value is an integer or can be converted to an integer
    try:
        num = float(value)
        if num >= 1000:
            return "{:,}".format(num)
    except ValueError:
        pass
    
    # Return the original value if it's not an integer or less than 1000
    return value                

#new xl maps functions 
def update_place_to_xlmap(xl_map,place_index,place_dictionary)->list:
    "Will return updates place dictionary"
    xl_map[2][place_index] = place_dictionary.get("car_count_monday_to_friday")
    xl_map[3][place_index]=place_dictionary.get("car_count_saturday_sunday")
    xl_map[4][place_index]=place_dictionary.get("retail_car_count_monday_to_friday")
    xl_map[5][place_index]=place_dictionary.get("retail_car_count_saturday_sunday")
    
    xl_map[7][place_index]=place_dictionary.get("retail_revenue_monday_to_friday")
    xl_map[8][place_index]=place_dictionary.get("retail_revenue_saturday_sunday")
    xl_map[9][place_index]=place_dictionary.get("total_revenue_monday_to_friday")
    xl_map[10][place_index]=place_dictionary.get("total_revenue_saturday_sunday")
    
    xl_map[14][place_index]=place_dictionary.get("labour_hours_monday_to_friday")
    xl_map[15][place_index]=place_dictionary.get("labour_hours_saturday_sunday")
    xl_map[16][place_index]=place_dictionary.get("cars_per_labour_hour_monday_to_friday")
    xl_map[17][place_index]=place_dictionary.get("cars_per_labour_hour_saturday_sunday")
    
    xl_map[19][place_index] = place_dictionary.get("arm_plans_sold_cnt")
    xl_map[20][place_index]= place_dictionary.get("conversion_rate")
    xl_map[21][place_index] = place_dictionary.get("total_arm_planmembers_cnt")
        
    return xl_map

def set_colour(val,row,col,worksheet,colours):
    darkgreen_format,light_green_format,darkred_format,lightred_format =colours
    cell=worksheet.cell(row,col)
    if val>=20:
       cell.fill = darkgreen_format
    
    elif val >=10:
        cell.fill = light_green_format
        
        
    elif val>=-10:
        cell.fill = darkred_format
    else : #val>=-20
       cell.fill = lightred_format

def set_colour_for_avg_retail(current_week,past_4_weeks, row, col, worksheet, colours):
    """This will do colur coding of for the xl sheet 10 5 0 -5 - 10

    Args:
        current_week (_type_): _description_
        past_4_weeks (_type_): _description_
    """

    if not all([current_week, past_4_weeks]):
        return
    
    darkgreen_format,light_green_format,darkred_format,lightred_format =colours
    cell=worksheet.cell(row,col)

    # percentage = ((current_week - past_4_weeks) / past_4_weeks)*100
    
    # if percentage >10:
    #     #Dark green
    #     cell.fill = darkgreen_format
        
    # elif percentage > 5 and percentage <=9:
    #     #light green
    #     cell.fill = light_green_format
        
    # elif percentage <= 5  and percentage <=-5:
    #     print("neutral")
    # elif percentage < -5 and percentage>=-9:
    #     cell.fill = lightred_format
    #     #light red
        
    # elif percentage < -10:
    #     cell.fill = darkred_format
    percentage = ((current_week - past_4_weeks) / past_4_weeks)*100
    
    print("percentage :",percentage)
    if percentage >=10:
        #Dark green
        print("Dark green ")
        cell.fill = darkgreen_format
    elif percentage >= 5:
        print("light green")
        cell.fill = light_green_format
    elif percentage > -5:
        print("neutral")
        
    elif percentage > -10:
        print("light red ")
        cell.fill = lightred_format
    elif percentage <=-10:
        print("Dark red")  
        cell.fill = darkred_format

def set_colour_new(current_week,past_4_weeks,row,col,worksheet,colours):
    """This will do colur coding of for the xl sheet 10 5 0 -5 - 10

    Args:
        current_week (_type_): _description_
        past_4_weeks (_type_): _description_
    """

    if not all([current_week, past_4_weeks]):
        return
    print(f"location on xl : {row},{col}")
    print(f"current : {current_week} , past week  { past_4_weeks}")
    darkgreen_format,light_green_format,darkred_format,lightred_format =colours
    cell=worksheet.cell(row,col)

    percentage = ((current_week - past_4_weeks) / past_4_weeks)*100
    

    print("percentage :",percentage)
    # if percentage >=10:
    #     #Dark green
    #     print("Dark green ")
    #     cell.fill = darkgreen_format
        
    # elif percentage >= 5:
    #     print("light green")
    #     cell.fill = light_green_format
        
    # elif percentage >= -5:
    #     print("neutral")
    
    # elif percentage >=-9:
    #     print("light red ")
    #     cell.fill = lightred_format
        
    # elif percentage <=-10:
    #     cell.fill = darkred_format
    # brlow is 20 percent 
    # if percentage >=20:
    #     #Dark green
    #     print("Dark green ")
    #     cell.fill = darkgreen_format
        
    # elif percentage >= 10:
    #     print("light green")
    #     cell.fill = light_green_format
    # elif percentage >= -9:
    #     print("neutral")
    
    # elif percentage >=-19:
    #     print("light red ")
    #     cell.fill = lightred_format
    # elif percentage <=-20:
    #     print("Dark red")
    #     cell.fill = darkred_format
    
    if percentage >=10:
        #Dark green
        print("Dark green ")
        cell.fill = darkgreen_format
    elif percentage >= 5:
        print("light green")
        cell.fill = light_green_format
    elif percentage > -5:
        print("neutral")
        
    elif percentage > -10:
        print("light red ")
        cell.fill = lightred_format
    elif percentage <=-10:
        print("Dark red")  
        cell.fill = darkred_format

def chnage_total_car_count_fun(curent_car_cnt,past_4_car_cnt):
    chnage = None
    past_4_car_cnt_avg = past_4_car_cnt/4
    chnage = ((curent_car_cnt- past_4_car_cnt_avg)/past_4_car_cnt_avg)*100

    return chnage

def chnage_total_revenue_fun(curent_revenue,past_4_revenue):
    chnage = None
    past_4_revenue_avg = past_4_revenue/4
    chnage = ((curent_revenue - past_4_revenue_avg)/past_4_revenue_avg)*100
    print(f"change:{chnage} = ({curent_revenue}-{past_4_revenue_avg})/{past_4_revenue_avg}")
    return chnage

def handle_zero_divison(a, b):
    if b == 0:
        return ''
    else:
        return a/b

def prepare_xlmap(data,comment="The comment section",filename="test.xlsx",sheet_name="sheet1"):
    # Load the existing workbook using openpyxl
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.create_sheet(sheet_name)
    except Exception as _:
        print(f"creating neww xl file !! {filename}")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
    
    
    xl_map = [
    [""],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ]
    
    cols = ["    ","Totals","ILL",
                "GA / SC","Sudz - Beverly",'Fuller-Calumet',
                "Fuller-Cicero","Fuller-Matteson","Fuller-Elgin",
                "Splash-Peoria","Getaway-Macomb","Getaway-Morton",
                "Getaway-Ottawa","Getaway-Peru","Sparkle-Belair",
                "Sparkle-Evans","Sparkle-Furrys Ferry","Sparkle-Greenwood",
                "Sparkle-Grovetown 1","Sparkle-Grovetown 2","Sparkle-North Augusta",
                "Sparkle-Peach Orchard","Sparkle-Windsor Spring"]

    index_names =["Car Count Mon - Fri","Car Count Sat - Sun","Retail Car Count Mon - Fri",
                "Retail Car Count Sat - Sun","Total Cars","Retail Revenue Mon - Fri",
                "Retail Revenue Sat - Sun","Total Revenue Mon - Fri","Total Revenue Sat - Sun",
                "Total Revenue","Avg. Retail Visit","Avg. Member Visit",
                "Staff Hours Mon - Fri","Staff Hours Sat - Sun","Cars Per Labor Hour Mon - Fri",
                "Cars Per Labor Hour Sat & Sun","Total Cars Per Man Hour","Total Club Plans Sold",
                "Conversion Rate","Total Club Plan Members"]

    xl_map[0][0] = comment #to maps
    
    
    
    for col in range(len(cols)): #column names to maps
        xl_map[1][col]=cols[col]
        
    for index  in range(len(index_names)):  #index names to map
        xl_map[index+2][0]=index_names[index]
    
    Fuller_Cicero = data.get("Fuller-Cicero")  

    if Fuller_Cicero:
        Fuller_Cicero_index=6
        update_place_to_xlmap(xl_map,Fuller_Cicero_index,Fuller_Cicero)
    Fuller_Matteson = data.get("Fuller-Matteson")
    
    if Fuller_Matteson:
        Fuller_Matteson_index=7
        update_place_to_xlmap(xl_map,Fuller_Matteson_index,Fuller_Matteson)
    
    Fuller_Elgin = data.get("Fuller-Elgin")
    
    if Fuller_Elgin:
        Fuller_Elgin_index = 8
        update_place_to_xlmap(xl_map,Fuller_Elgin_index,Fuller_Elgin)
        
      
    Sparkle_Belair = data.get("Sparkle-Belair")
    
    if Sparkle_Belair:
        Sparkle_Belair_index=14

        update_place_to_xlmap(xl_map,Sparkle_Belair_index,Sparkle_Belair)
        
    Sparkle_Evans = data.get("Sparkle-Evans")
    
    if Sparkle_Evans:
        Sparkle_Evans_index=15

        update_place_to_xlmap(xl_map,Sparkle_Evans_index,Sparkle_Evans)

    Sparkle_North_Augusta = data.get("Sparkle-North Augusta")
    
    if Sparkle_North_Augusta:
        Sparkle_North_Augusta_index=20

        update_place_to_xlmap(xl_map,Sparkle_North_Augusta_index,Sparkle_North_Augusta)
    
    Sparkle_Greenwood = data.get("Sparkle-Greenwood")
    
    if Sparkle_Greenwood:
        Sparkle_Greenwood_index=17

        update_place_to_xlmap(xl_map,Sparkle_Greenwood_index,Sparkle_Greenwood)
    
    
    Sparkle_Grovetown_1 = data.get("Sparkle-Grovetown 1")
    
    if Sparkle_Grovetown_1:
        Sparkle_Grovetown_1_index=18

        update_place_to_xlmap(xl_map,Sparkle_Grovetown_1_index,Sparkle_Grovetown_1)
        
    Sparkle_Windsor_Spring = data.get("Sparkle-Windsor Spring")    
    
    if Sparkle_Windsor_Spring:
        Sparkle_Windsor_Spring_index=22 
        
        update_place_to_xlmap(xl_map,Sparkle_Windsor_Spring_index,Sparkle_Windsor_Spring)
        
    Sparkle_Furrys_Ferry = data.get("Sparkle-Furrys Ferry")
    
    if Sparkle_Furrys_Ferry:
        Sparkle_Furrys_Ferry_index= 16
        
        update_place_to_xlmap(xl_map,Sparkle_Furrys_Ferry_index,Sparkle_Furrys_Ferry)      

    Sparkle_Peach_Orchard = data.get("Sparkle-Peach Orchard")
    
    if Sparkle_Peach_Orchard:
        Sparkle_Peach_Orchard_index=21
        
        update_place_to_xlmap(xl_map,Sparkle_Peach_Orchard_index,Sparkle_Peach_Orchard)
        
    Sparkle_Grovetown_2 =  data.get("Sparkle-Grovetown 2")   
      
    if Sparkle_Grovetown_2:
        Sparkle_Grovetown_2_index =19
        
        update_place_to_xlmap(xl_map,Sparkle_Grovetown_2_index,Sparkle_Grovetown_2)
        
    Fuller_Calumet = data.get("Fuller-Calumet")
    
    if Fuller_Calumet:
        Fuller_Calumet_index= 5
        
        update_place_to_xlmap(xl_map,Fuller_Calumet_index,Fuller_Calumet)
        
        
    Sudz_Beverly = data.get("Sudz - Beverly")
    
    if Sudz_Beverly:
        Sudz_Beverly_index = 4
        
        update_place_to_xlmap(xl_map,Sudz_Beverly_index,Sudz_Beverly)
        
    ##Hamilton sites 
    
    Getaway_Macomb = data.get("Getaway-Macomb")
    
    if Getaway_Macomb:
        Getaway_Macomb_index = 10
        update_place_to_xlmap(xl_map,Getaway_Macomb_index,Getaway_Macomb)
        
    Getaway_Morton = data.get("Getaway-Morton")
    
    if Getaway_Morton:
        Getaway_Morton_index=11
        update_place_to_xlmap(xl_map,Getaway_Morton_index,Getaway_Morton)
        
    Getaway_Ottawa = data.get("Getaway-Ottawa")
    
    if Getaway_Ottawa:
        Getaway_Ottawa_index = 12
        update_place_to_xlmap(xl_map,Getaway_Ottawa_index,Getaway_Ottawa)
        
    Getaway_Peru = data.get("Getaway-Peru")    
    
    if Getaway_Peru:
        Getaway_Peru_index = 13
        update_place_to_xlmap(xl_map,Getaway_Peru_index,Getaway_Peru)
        
    
    #Hamilton 
    
    Splash_Peoria = data.get("Splash-Peoria")
    
    if Splash_Peoria:
        Splash_Peoria_index = 9
        update_place_to_xlmap(xl_map,Splash_Peoria_index,Splash_Peoria)

    #computation for first three columns 
    # sum([ xl_map[2][i+1] for i in range(3,13)])
    
    #first index
    car_count_monday_to_friday_ILL = do_sum(xl_map,2,range(3,13))
    xl_map[2][2] = car_count_monday_to_friday_ILL
    
    car_count_monday_to_friday_GA_SC = do_sum(xl_map,2,range(13,22))
    xl_map[2][3] = car_count_monday_to_friday_GA_SC
    
    car_count_monday_to_friday_Total = sum([car_count_monday_to_friday_ILL,car_count_monday_to_friday_GA_SC])
    
    xl_map[2][1]=car_count_monday_to_friday_Total
    
    
    
    #secound index 
    car_count_saturday_to_sunday_ILL = do_sum(xl_map,3,range(3,13))
    xl_map[3][2] = car_count_saturday_to_sunday_ILL
    
    car_count_saturday_to_sunday_GA_SC = do_sum(xl_map,3,range(13,22))
    xl_map[3][3] = car_count_saturday_to_sunday_GA_SC
    
    car_count_saturday_to_sunday_Total = sum([car_count_saturday_to_sunday_ILL, car_count_saturday_to_sunday_GA_SC])
    
    xl_map[3][1] = car_count_saturday_to_sunday_Total
    
    #third row 
    
    retail_car_count_monday_to_friday_ILL = do_sum(xl_map,4,range(3,13))
    xl_map[4][2] = retail_car_count_monday_to_friday_ILL
    
    retail_car_count_monday_to_friday_GA_SC = do_sum(xl_map,4,range(13,22))
    xl_map[4][3] = retail_car_count_monday_to_friday_GA_SC
    
    retail_car_count_monday_to_friday_Total = sum([retail_car_count_monday_to_friday_ILL,retail_car_count_monday_to_friday_GA_SC])
    
    xl_map[4][1]=retail_car_count_monday_to_friday_Total
    
    # Reatil car cound satruday to sunday
    
    retail_car_count_saturday_to_sunday_ILL = do_sum(xl_map,5,range(3,13))
    xl_map[5][2] = retail_car_count_saturday_to_sunday_ILL
    
    retail_car_count_saturday_to_sunday_GA_SC = do_sum(xl_map,5,range(13,22))
    xl_map[5][3] = retail_car_count_saturday_to_sunday_GA_SC
    
    retail_car_count_saturday_to_sunday_Total = sum([retail_car_count_saturday_to_sunday_ILL,retail_car_count_saturday_to_sunday_GA_SC])
    
    xl_map[5][1]=retail_car_count_saturday_to_sunday_Total
    
    loc_1 = [[2,3],[3,3]]
    total_cars_GA_SC=do_sum_location(xl_map,loc_1)
    
    xl_map[6][3]=total_cars_GA_SC
    
    loc_2=[[2,2],[3,2]]
    total_cars_in_ILL = do_sum_location(xl_map,loc_2)
    
    xl_map[6][2] = total_cars_in_ILL
    
    loc_3=[[6,3],[6,2]]
    
    Total_cars_Total = do_sum_location(xl_map,loc_3)
    xl_map[6][1]=Total_cars_Total
    
    #Retail Revenue Monday to Friday
    retail_revenue_monday_to_friday_ILL = do_sum(xl_map,7,range(3,13))
    xl_map[7][2] = retail_revenue_monday_to_friday_ILL
    
    retail_revenue_monday_to_friday_GA_SC = do_sum(xl_map,7,range(13,22))
    xl_map[7][3] = retail_revenue_monday_to_friday_GA_SC
    
    retail_revenue_monday_to_friday_Total = sum([retail_revenue_monday_to_friday_ILL,retail_revenue_monday_to_friday_GA_SC])
    
    xl_map[7][1]=retail_revenue_monday_to_friday_Total
    
    
    #Reatil Revenue Saturda to Sunday
    retail_revenue_saturday_to_sunday_ILL = do_sum(xl_map,8,range(3,13))
    xl_map[8][2] = retail_revenue_saturday_to_sunday_ILL
    
    retail_revenue_saturday_to_sunday_GA_SC = do_sum(xl_map,8,range(13,22))
    xl_map[8][3] = retail_revenue_saturday_to_sunday_GA_SC
    
    retail_revenue_saturday_to_sunday_Total = sum([retail_revenue_saturday_to_sunday_ILL,retail_revenue_saturday_to_sunday_GA_SC])
    
    xl_map[8][1]=retail_revenue_saturday_to_sunday_Total
    
    #Total Revnue Monday to Friday
    Total_revenue_monday_to_friday_ILL = do_sum(xl_map,9,range(3,13))
    xl_map[9][2] = Total_revenue_monday_to_friday_ILL
    
    Total_revenue_monday_to_friday_GA_SC = do_sum(xl_map,9,range(13,22))
    xl_map[9][3] = Total_revenue_monday_to_friday_GA_SC
    
    Total_revenue_monday_to_friday = sum([Total_revenue_monday_to_friday_ILL,Total_revenue_monday_to_friday_GA_SC])
    
    xl_map[9][1]=Total_revenue_monday_to_friday
    
    # Total Revenue Saturday to Sunday
    Total_revenue_saturday_to_sunday_ILL = do_sum(xl_map,10,range(3,13))
    xl_map[10][2] = Total_revenue_saturday_to_sunday_ILL
    
    Total_revenue_saturday_to_sunday_GA_SC = do_sum(xl_map,10,range(13,22))
    xl_map[10][3] = Total_revenue_saturday_to_sunday_GA_SC
    
    Total_revenue_saturday_sunday = sum([Total_revenue_saturday_to_sunday_ILL,Total_revenue_saturday_to_sunday_GA_SC])
    
    xl_map[10][1]=Total_revenue_saturday_sunday 
    
    # Total Revenue 
    loc_4=[[9,2],[10,2]]
    Total_revenue_ILL = do_sum_location(xl_map,loc_4)
    xl_map[11][2] =Total_revenue_ILL
    
    loc_5=[[9,3],[10,3]]
    Total_revenue_GA_SC = do_sum_location(xl_map,loc_5)
    xl_map[11][3]= Total_revenue_GA_SC
    
    Total_revenue_Total = sum([Total_revenue_ILL,Total_revenue_GA_SC])
    xl_map[11][1] =Total_revenue_Total
    
    #Average Retail Visit

    
    
    Average_retail_visit_IL_val =Average_retail_visit_IL_function(retail_revenue_monday_to_friday_ILL,
                                     retail_revenue_saturday_to_sunday_ILL,
                                     retail_car_count_monday_to_friday_ILL,
                                     retail_car_count_saturday_to_sunday_ILL)
    xl_map[12][2]=round(Average_retail_visit_IL_val,2)
    
    

    Average_retail_visit__GA_SC_val = Average_retail_visit__GA_SC_fucntion(retail_revenue_monday_to_friday_GA_SC
                                                                            ,retail_revenue_saturday_to_sunday_GA_SC,
                                                                            retail_car_count_monday_to_friday_GA_SC,
                                                                            retail_car_count_saturday_to_sunday_GA_SC)
    xl_map[12][3] = round(Average_retail_visit__GA_SC_val,2)
    
    Average_retail_visit_Total_val =Average_retail_visit_Total_function(retail_revenue_monday_to_friday_Total,
                                        retail_revenue_saturday_to_sunday_Total,
                                        retail_car_count_monday_to_friday_Total,
                                        retail_car_count_saturday_to_sunday_Total)
    xl_map[12][1] = round(Average_retail_visit_Total_val,2)
    
    #Average Member visit 
    
    Average_memeber_visit_ILL_val =Average_memeber_visit_ILL_function(Total_revenue_ILL,retail_revenue_monday_to_friday_ILL,
                                       retail_revenue_saturday_to_sunday_ILL,total_cars_in_ILL,
                                       retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL)
    xl_map[13][2] = round(Average_memeber_visit_ILL_val,2)
    
    Average_memeber_visit_GA_SC_val = Average_memeber_visit_GA_SC_function(Total_revenue_GA_SC,
        retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC,
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC,total_cars_GA_SC)
    
    xl_map[13][3] = round(Average_memeber_visit_GA_SC_val,2)
    
    Average_memeber_visit_Total_val =Average_memeber_visit_Total_function(Total_revenue_Total,retail_revenue_monday_to_friday_Total
                        ,retail_revenue_saturday_to_sunday_Total,Total_cars_Total,retail_car_count_monday_to_friday_Total,
                        retail_car_count_saturday_to_sunday_Total)
    
    xl_map[13][1] = round(Average_memeber_visit_Total_val,2)
    
    #Staff Hours Monday to Friday
    
    staff_hours_monday_to_friday_ILL = do_sum(xl_map,14,range(3,13))
    xl_map[14][2] = staff_hours_monday_to_friday_ILL
    
    staff_hours_monday_to_friday_GA_SC = do_sum(xl_map,14,range(13,22))
    
    xl_map[14][3] = staff_hours_monday_to_friday_GA_SC
    
    staff_hours_monday_to_friday_Total = sum([staff_hours_monday_to_friday_ILL,staff_hours_monday_to_friday_GA_SC])
    
    xl_map[14][1] = staff_hours_monday_to_friday_Total
    
    
    #Staff Hours Saturday to Sunday

    staff_hours_saturday_to_sunday_ILL = do_sum(xl_map,15,range(3,13))
    xl_map[15][2] = staff_hours_saturday_to_sunday_ILL
    
    staff_hours_saturday_to_sunday_GA_SC = do_sum(xl_map,15,range(13,22))
    
    xl_map[15][3] = staff_hours_saturday_to_sunday_GA_SC
    
    staff_hours_saturday_to_sunday_Total = sum([staff_hours_saturday_to_sunday_ILL,staff_hours_saturday_to_sunday_GA_SC])
    
    xl_map[15][1] = staff_hours_saturday_to_sunday_Total  
    
    # Cost per labour hour  Monday to friday
    cost_per_labour_hour_monday_to_friday_ILL_val =cost_per_labour_hour_monday_to_friday_ILL_function(car_count_monday_to_friday_ILL,staff_hours_monday_to_friday_ILL)
    xl_map[16][2] = round(cost_per_labour_hour_monday_to_friday_ILL_val,2)
    
    
    cost_per_labour_hour_monday_to_friday_GA_SC_val = cost_per_labour_hour_monday_to_friday_GA_SC_fucntion(car_count_monday_to_friday_GA_SC,staff_hours_monday_to_friday_GA_SC)
    xl_map[16][3] = round(cost_per_labour_hour_monday_to_friday_GA_SC_val,2)
    
    
    cost_per_labour_hour_monday_to_friday_Total_val =cost_per_labour_hour_monday_to_friday_Total_function(car_count_monday_to_friday_Total,staff_hours_monday_to_friday_Total)
    
    xl_map[16][1] = round(cost_per_labour_hour_monday_to_friday_Total_val,2)
    
    #Cost per laobour hour Saturday and Sunday
    cost_per_labour_hour_saturday_to_sunday_ILL_val = cost_per_labour_hour_saturday_to_sunday_ILL_function(car_count_saturday_to_sunday_ILL,staff_hours_saturday_to_sunday_ILL)
    xl_map[17][2] = round(cost_per_labour_hour_saturday_to_sunday_ILL_val,2)
    
    
    
    cost_per_labour_hour_saturday_to_sunday_GA_SC_val = cost_per_labour_hour_saturday_to_sunday_GA_SC_function(car_count_saturday_to_sunday_GA_SC,staff_hours_saturday_to_sunday_GA_SC)
    xl_map[17][3] = round(cost_per_labour_hour_saturday_to_sunday_GA_SC_val,2)
    
    cost_per_labour_hour_saturday_to_sunday_Total= car_count_saturday_to_sunday_Total/staff_hours_saturday_to_sunday_Total if staff_hours_saturday_to_sunday_Total!=0 else ""
    
    xl_map[17][1] = round(cost_per_labour_hour_saturday_to_sunday_Total,2) if cost_per_labour_hour_saturday_to_sunday_Total else ""
    
    # Total cars per man hour
    

    Total_cars_per_man_hour_ILL_val =Total_cars_per_man_hour_ILL_function(total_cars_in_ILL,
                                         staff_hours_monday_to_friday_ILL,
                                         staff_hours_saturday_to_sunday_ILL)
    xl_map[18][2] = round(Total_cars_per_man_hour_ILL_val,2)
    
    Total_cars_per_man_hour_GA_SC_val = Total_cars_per_man_hour_GA_SC_function(total_cars_GA_SC, staff_hours_monday_to_friday_GA_SC,
                                                                               staff_hours_saturday_to_sunday_GA_SC)
    
    xl_map[18][3] = round(Total_cars_per_man_hour_GA_SC_val,2)
    
    
    Total_cars_per_man_hour_total_val =Total_cars_per_man_hour_total_function(Total_cars_Total,
                                           staff_hours_monday_to_friday_Total,
                                           staff_hours_saturday_to_sunday_Total)
    xl_map[18][1] = round(Total_cars_per_man_hour_total_val,2)
    
    #Total club plans sold 
    Total_club_plans_sold_ILL = do_sum(xl_map,19,range(3,13))
    
    xl_map[19][2] = Total_club_plans_sold_ILL
    
    Total_club_plans_sold_GA_SC = do_sum(xl_map,19,range(13,22))
    
    xl_map[19][3] = Total_club_plans_sold_GA_SC
    
    Total_club_plans_sold_Total = sum([Total_club_plans_sold_ILL,Total_club_plans_sold_GA_SC])
    
    xl_map[19][1] = Total_club_plans_sold_Total
    
    
    #Conversion Rate 
    Conversion_rate_ILL_val =Conversion_rate_ILL_function(Total_club_plans_sold_ILL,
                                                          retail_car_count_monday_to_friday_ILL,
                                                          retail_car_count_saturday_to_sunday_ILL)
    xl_map[20][2] = round((Conversion_rate_ILL_val * 100),2)
    
    Conversion_rate_GA_SC_val =Conversion_rate_GA_SC_function(Total_club_plans_sold_GA_SC,retail_car_count_monday_to_friday_GA_SC,
                                   retail_car_count_saturday_to_sunday_GA_SC)
    
    xl_map[20][3]= round((Conversion_rate_GA_SC_val * 100),2)
    
    

    Conversion_rate_Total_val = Conversion_rate_Total_function(Total_club_plans_sold_Total,
                                                               retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total)
    xl_map[20][1] = round((Conversion_rate_Total_val * 100),2)
    
    
    #Total club plan members 
    Total_club_planmembers_ILL = do_sum(xl_map,21,range(3,13))
    
    xl_map[21][2] = Total_club_planmembers_ILL
    
    Total_club_planmembers_GA_SC = do_sum(xl_map,21,range(13,22))
    
    xl_map[21][3] = Total_club_planmembers_GA_SC 
    
    Total_club_planmembers_Total = sum([Total_club_planmembers_ILL,Total_club_planmembers_GA_SC])
    
    xl_map[21][1] = Total_club_planmembers_Total
    
    #Toatl Cars for all locations 
    total_cars_row = 6 
    for i in range(3,22):
        xl_map[total_cars_row][i+1]= do_sum_location(xl_map,[[2,i+1],[3,i+1]])
    
    
    revenue_row = 11
    
    #Total Revenue calculatio n for all locations 
    for i in range(3,22):
        xl_map[revenue_row][i+1]= do_sum_location(xl_map,[[9,i+1],[10,i+1]])
    
    #Average_retail_visit
    average_retail_visit_row = 12
    for i in range(3,22):
        retail_revenue_mon_sun = do_sum_location(xl_map,location=[[7,i+1],[8,i+1]])
        retail_car_count_mon_sun = do_sum_location(xl_map,location=[[4,i+1],[5,i+1]])
        
        average_retail_visit_val = retail_revenue_mon_sun/retail_car_count_mon_sun if retail_car_count_mon_sun != 0 else ""
        xl_map[average_retail_visit_row][i+1]=   round(average_retail_visit_val,2) if average_retail_visit_val else ""
    
    
    #Average member visit
    average_member_visit_row = 13
    for i in range(3,22):
        total_revenue = xl_map[11][i+1]
        total_revenue = total_revenue if isinstance(total_revenue,int) or isinstance(total_revenue,float) else 0
        
        retail_revenue_mon_sun = do_sum_location(xl_map,location=[[7,i+1],[8,i+1]])
        
        total_cars = xl_map[6][i+1]
        total_cars = total_cars if isinstance(total_cars,int) or isinstance(total_cars,float) else 0
        
        retail_car_count_mon_sun  = do_sum_location(xl_map,location=[[4,i+1],[5,i+1]])
        
        average_member_visit_val = (total_revenue-retail_revenue_mon_sun)/(total_cars-retail_car_count_mon_sun) if total_cars-retail_car_count_mon_sun !=0 else ""
        
        xl_map[average_member_visit_row][i+1] = round(average_member_visit_val,2) if average_member_visit_val else ""
    
    #Total cars per man hour 
    total_cars_per_man_hour_row = 18 
    for i in range(3,22):
        total_cars = xl_map[6][i+1]
        total_cars = total_cars if isinstance(total_cars,int) or isinstance(total_cars,float) else 0
        staff_hours_monday_to_saturday = do_sum_location(xl_map,[[14,i+1],[15,i+1]])
        
        total_cars_per_man_hour_val = total_cars/staff_hours_monday_to_saturday if staff_hours_monday_to_saturday !=0 else ""
        
        xl_map[total_cars_per_man_hour_row][i+1] = round(total_cars_per_man_hour_val,2) if total_cars_per_man_hour_val else ""
 
        
      
    # Define cell styles (assuming they are the same as before)
    bg_color = PatternFill(start_color='0b3040', end_color='0b3040', fill_type='solid')
    font_color = Font(color='FFFFFF')

    bg_color_index = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    font_color_index = Font(color='000000')

    darkgreen_format = PatternFill(start_color='0ee85e', end_color='0ee85e', fill_type='solid')
    light_green_format = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    darkred_format = PatternFill(start_color='FC0303', end_color='FC0303', fill_type='solid')
    lightred_format = PatternFill(start_color='D98484', end_color='D98484', fill_type='solid')
    yellow=  PatternFill(start_color='D0D48A', end_color='D0D48A', fill_type='solid')
    
    #writing to  actual sheet
    #first row comment section
    # Write comments in the first row
    # worksheet.append([comment])
    
    for row in range(len(xl_map)):
        for col in range(len(xl_map[row])):
            val = xl_map[row][col]
            cell = worksheet.cell(row=row+1, column=col+1, value=val)  # offset by 0 rows for header and comment
            
            # val = add_commas(val)
            # val = float(val)
            # print("type:",type(val),val)
            if row == 1 and col != 0:
                cell.fill = bg_color
                cell.font = font_color

            elif col == 0 and 1 < row < 22:
                cell.fill = bg_color_index
                cell.font = font_color_index

            # elif val  and row == 12 and col > 0:
            #     if val >= 10:
            #         cell.fill = darkgreen_format
                    
            #     elif val>=5 and val <10: # [5,9] (inclusive intervals)
            #         cell.fill = light_green_format
                    
            #     elif val>=-5 and val <5:  # [-5,4]
            #         cell.fill = lightred_format
        
            #     elif  val <=-10 and val<-5:
            #         cell.fill = darkred_format

            # elif val  and row in [16, 17, 18, 20] and col > 0:
            #     if val >= 20:
            #         cell.fill = darkgreen_format
            #     elif val>=10 and val <20:
            #         cell.fill = light_green_format
            #     elif val>=-10 and val <10:
            #         cell.fill = lightred_format
            #     elif val <=-20 or val<-10:
            #         cell.fill = darkred_format

            # elif row == 5 and col > 2:
            #     cell.value = ""

            # elif row in [17] and col > 2:
            #     cell.value = ""

            # elif row == 21:
            #     cell.value = ""

            # elif row > 21 and col > 0:
            #     cell.value = ""

    # Add legend or additional information below the table
    legend_start_row = 24

    legend_styles = {
        "Very Concerning": PatternFill(start_color="fc0303", end_color="fc0303", fill_type="solid"),
        "Concerning": PatternFill(start_color="d98484", end_color="d98484", fill_type="solid"),
        "Neutral": PatternFill(start_color="d0d48a", end_color="d0d48a", fill_type="solid"),
        "Positive": PatternFill(start_color="8ad493", end_color="8ad493", fill_type="solid"),
        "Very Positive": PatternFill(start_color="0ee85e", end_color="0ee85e", fill_type="solid"),
    }

    worksheet.cell(row=legend_start_row, column=1, value='Legend')
    worksheet.cell(row=legend_start_row + 1, column=1, value='Very Concerning').fill = legend_styles["Very Concerning"]
    worksheet.cell(row=legend_start_row + 2, column=1, value='Concerning').fill = legend_styles["Concerning"]
    worksheet.cell(row=legend_start_row + 3, column=1, value='Neutral').fill = legend_styles["Neutral"]
    worksheet.cell(row=legend_start_row + 4, column=1, value='Positive').fill = legend_styles["Positive"]
    worksheet.cell(row=legend_start_row + 5, column=1, value='Very Positive').fill = legend_styles["Very Positive"]
    
    #setting cloumn width as deafult 
    column_width = 20  # You can change this value to whatever width you need
    first_col_width =25
    for col in range(1, 24):  # Columns A to W are 1 to 23
        column_letter = get_column_letter(col)
        if col==1:
            worksheet.column_dimensions[column_letter].width = first_col_width
        else:
            worksheet.column_dimensions[column_letter].width = column_width
       
    
    # Define the border style


    thick_border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'))
    
    thick_border_bottom = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    bottom=Side(style='thick'))




    # Apply the border to a range of cells (e.g., A1:C3)
    # for row in worksheet.iter_rows(min_row=3, max_row=21, min_col=2, max_col=4):
    #     for cell in row:
    #         cell.border = thin_border

    for row in range(3,23):
        cell0 = worksheet.cell(row=row,column=1)
        cell1 = worksheet.cell(row=row,column=2)
        cell2 = worksheet.cell(row=row,column=3)
        cell3 = worksheet.cell(row=row,column=4)
        
        if row in [7,12,13,19,21,22]:
            cell0.border=thick_border_bottom
            cell1.border=thick_border_bottom
            cell2.border=thick_border_bottom
            cell3.border=thick_border_bottom
        else:
            cell0.border=thick_border
            cell1.border=thick_border
            cell2.border=thick_border
            cell3.border=thick_border
    
    #number formates 3 dots       
    for row in worksheet.iter_rows():
        for cell in row:
            row_index = cell.row
            if isinstance(cell.value, (int, float)) and row_index in [17,18,19]:
                cell.number_format = '#,##0.0'
            elif isinstance(cell.value, (int, float)) and row_index in [13,14,21]:
                cell.number_format = '#,##0.00'
            elif isinstance(cell.value, (int, float)): #:
                cell.number_format = '#,##0'
    #Doller sysmbol     
    for row in range(8,13):
        cell1 = worksheet.cell(row=row,column=2)
        cell2 = worksheet.cell(row=row,column=3)
        cell3 = worksheet.cell(row=row,column=4)
        
        # cell1.border=thick_border
        # cell2.border=thick_border
        # cell3.border=thick_border
        
        cells=[cell1,cell2,cell3]
        for cell in cells:
            if isinstance(cell.value, (int, float)) and cell.value >= 1000:
                cell.number_format = '"$"#,##0'     
    
    #finding past week averages for Total car count
    all_locations = [Sudz_Beverly,Fuller_Calumet,
                    Fuller_Cicero,Fuller_Matteson,
                    Fuller_Elgin,Splash_Peoria,Getaway_Macomb,Getaway_Morton,
                    Getaway_Ottawa,Getaway_Peru,Sparkle_Belair,
                    Sparkle_Evans,Sparkle_Furrys_Ferry,Sparkle_Greenwood,
                    Sparkle_Grovetown_1,Sparkle_Grovetown_2,Sparkle_North_Augusta,
                    Sparkle_Peach_Orchard,Sparkle_Windsor_Spring]
    
    
    lst_main = ["Totals","ILL",
                "GA / SC"]
    ga_sc = all_locations[10:]
    ill = all_locations[0:10]

    ill_past_4_weeks_car_cnt = [loc_data.get("past_4_week_cnt") for loc_data in ill if loc_data]
    ill_sum_past = sum(ill_past_4_weeks_car_cnt)/4
    current_ill_week_cnt = xl_map[6][2]
    ill_average_percent = ((current_ill_week_cnt - ill_sum_past)/ ill_sum_past)*100
    
    ga_sc_past_4_weeks_car_cnt = [loc_data.get("past_4_week_cnt") for loc_data in ga_sc if loc_data]
    ga_sc_sum_past = sum(ga_sc_past_4_weeks_car_cnt)/4
    current_ga_sc_week_cnt = xl_map[6][3]
    ga_sc_average_percent= ((current_ga_sc_week_cnt-ga_sc_sum_past)/ga_sc_sum_past)*100
    
    totals_past = ill_sum_past +ga_sc_sum_past
    totals_current = xl_map[6][1]
    total_average_percent = ((totals_current-totals_past)/totals_past)*100
    
    ill_total_revenue_past_4_weeks = [loc_data.get("past_4_weeks_total_revenue") for loc_data in ill if loc_data]
    ill_total_revenue_avg = sum(ill_total_revenue_past_4_weeks)/4
    ill_curent_revenue = xl_map[11][2]
    ill_avg_revenue_change  = ((ill_curent_revenue-ill_total_revenue_avg)/ill_total_revenue_avg)*100
    
    ga_sc_total_revenue_past_4_weeks = [loc_data.get("past_4_weeks_total_revenue") for loc_data in ga_sc if loc_data]
    ga_sc_avg_revenue = sum(ga_sc_total_revenue_past_4_weeks)/4
    ga_sc_curent_revenue  = xl_map[11][3]
    
    ga_sc_avg_revenue_change = ((ga_sc_curent_revenue - ga_sc_avg_revenue)/ga_sc_avg_revenue)*100
    
    total_revenue_past_4_avg_total = ill_total_revenue_avg + ga_sc_avg_revenue
    total_reveneu_curent = ill_curent_revenue + ga_sc_curent_revenue
    
    total_revenue_total_change = ((total_reveneu_curent - total_revenue_past_4_avg_total)/total_revenue_past_4_avg_total)*100
    
    
    ill_past_4_weeks_arm_plan_sold = [loc_data.get("past_4_weeks_arm_plans_sold_cnt") for loc_data in ill if loc_data]
    ill_past_4_weeks_arm_plan_sold_sum = sum(ill_past_4_weeks_arm_plan_sold)
    
    ill_past_4_retail_car_count = [loc_data.get("past_4_weeks_retail_car_count") for loc_data in ill if loc_data]
    ill_past_4_retail_car_count_sum = sum(ill_past_4_retail_car_count)
    
    ill_past_4_conversation_rate = (ill_past_4_weeks_arm_plan_sold_sum/ill_past_4_retail_car_count_sum)
    
    ill_current_conversation_rate = xl_map[20][2]
    ill_conversation_rate_change = ill_current_conversation_rate - ill_past_4_conversation_rate
    
    ga_sc_past_4_weeks_arm_plans_sold = [loc_data.get("past_4_weeks_arm_plans_sold_cnt") for loc_data in ga_sc if loc_data]
    ga_sc_past_4_weeks_arm_plans_sold_sum = sum(ga_sc_past_4_weeks_arm_plans_sold)
    
    ga_sc_past_4_weeks_retail_car_count = [loc_data.get("past_4_weeks_retail_car_count") for loc_data in ga_sc if loc_data]
    ga_sc_past_4_weeks_retail_car_count_sum = sum(ga_sc_past_4_weeks_retail_car_count)
    
    ga_sc_past_4_conversation_rate = ( ga_sc_past_4_weeks_arm_plans_sold_sum/ga_sc_past_4_weeks_retail_car_count_sum)
    
    ga_sc_current_conversation_rate  = xl_map[20][3]
    ga_sc_conversation_change = ga_sc_current_conversation_rate - ga_sc_past_4_conversation_rate 
    
    past_total_arm_plans_sold = ill_past_4_weeks_arm_plan_sold_sum + ga_sc_past_4_weeks_arm_plans_sold_sum
    
    past_total_reatil_car_count = ill_past_4_retail_car_count_sum + ga_sc_past_4_weeks_retail_car_count_sum
    
    past_total_conversation_change = (past_total_arm_plans_sold/past_total_reatil_car_count)
    
    current_total_conversation_rate = xl_map[20][1]
    
    total_conversation_change = current_total_conversation_rate - past_total_conversation_change

    # Car Count Mon-fri
    ill_past_4_weeks_car_cnt_mon_fri = [loc_data.get("past_4_week_car_cnt_mon_fri", 0) for loc_data in ill if loc_data]
    ill_sum_past_mon_fri = sum(ill_past_4_weeks_car_cnt_mon_fri)/4
    current_ill_week_cnt_mon_fri = xl_map[2][2]

    ga_sc_past_4_weeks_car_cnt_mon_fri = [loc_data.get("past_4_week_car_cnt_mon_fri", 0) for loc_data in ga_sc if loc_data]
    ga_sc_sum_past_mon_fri = sum(ga_sc_past_4_weeks_car_cnt_mon_fri)/4
    current_ga_sc_week_cnt_mon_fri = xl_map[2][3]

    curr_total_car_cnt_mon_fri = current_ill_week_cnt_mon_fri+current_ga_sc_week_cnt_mon_fri
    past_4_week_total_car_cnt_mon_fri = ill_sum_past_mon_fri+ga_sc_sum_past_mon_fri

    # Car Count Sat-Sun
    ill_past_4_weeks_car_cnt_sat_sun = [loc_data.get("past_4_week_car_cnt_sat_sun", 0) for loc_data in ill if loc_data]
    ill_sum_past_sat_sun = sum(ill_past_4_weeks_car_cnt_sat_sun)/4
    current_ill_week_cnt_sat_sun = xl_map[3][2]

    ga_sc_past_4_weeks_car_cnt_sat_sun = [loc_data.get("past_4_week_car_cnt_sat_sun", 0) for loc_data in ga_sc if loc_data]
    ga_sc_sum_past_sat_sun = sum(ga_sc_past_4_weeks_car_cnt_sat_sun)/4
    current_ga_sc_week_cnt_sat_sun = xl_map[3][3]

    curr_total_car_cnt_sat_sun = current_ill_week_cnt_sat_sun+current_ga_sc_week_cnt_sat_sun
    past_4_week_total_car_cnt_sat_sun = ill_sum_past_sat_sun+ga_sc_sum_past_sat_sun

    # Retail Car Count mon-fri
    ill_past_4_weeks_retail_car_cnt_mon_fri = [loc_data.get("past_4_week_retail_car_count_mon_fri", 0) for loc_data in ill if loc_data]
    ill_sum_retail_car_cnt_mon_fri = sum(ill_past_4_weeks_retail_car_cnt_mon_fri)/4
    current_ill_week_retail_cnt_mon_fri = xl_map[4][2]

    ga_sc_past_4_weeks_retail_car_cnt_mon_fri = [loc_data.get("past_4_week_retail_car_count_mon_fri",0) for loc_data in ga_sc if loc_data]
    ga_sc_sum_retail_car_cnt_mon_fri = sum(ga_sc_past_4_weeks_retail_car_cnt_mon_fri)/4
    current_ga_sc_week_retail_cnt_mon_fri = xl_map[4][3]

    curr_total_retail_car_cnt_mon_fri = current_ill_week_retail_cnt_mon_fri+current_ga_sc_week_retail_cnt_mon_fri
    past_4_week_total_retail_car_cnt_mon_fri = ill_sum_retail_car_cnt_mon_fri+ga_sc_sum_retail_car_cnt_mon_fri

    # Retail Car Count Sat-Sun
    ill_past_4_weeks_retail_car_cnt_sat_sum = [loc_data.get("past_4_week_retail_car_count_sat_sun", 0) for loc_data in ill if loc_data]
    ill_sum_retail_car_cnt_sat_avg = sum(ill_past_4_weeks_retail_car_cnt_sat_sum)/4
    current_ill_week_retail_cnt_sat_sum = xl_map[5][2]

    ga_sc_past_4_weeks_retail_car_cnt_sat_sun = [loc_data.get("past_4_week_retail_car_count_sat_sun", 0) for loc_data in ga_sc if loc_data]
    ga_sc_sum_retail_car_cnt_sat_avg = sum(ga_sc_past_4_weeks_retail_car_cnt_sat_sun)/4
    current_ga_sc_week_retail_cnt_sat_sum = xl_map[5][3]

    curr_total_retail_car_cnt_sat_sun = current_ill_week_retail_cnt_sat_sum+current_ga_sc_week_retail_cnt_sat_sum
    past_4_week_total_retail_car_cnt_sat_sun = ill_sum_retail_car_cnt_sat_avg+ga_sc_sum_retail_car_cnt_sat_avg

    # Retail revenue mon-fri
    ill_past_4_week_retail_revenue_mon_fri = [loc_data.get("past_4_week_retail_revenue_mon_fri", 0) for loc_data in ill if loc_data]
    ill_past_4_week_retail_revenue_mon_fri_avg = sum(ill_past_4_week_retail_revenue_mon_fri)/4
    current_ill_retail_revenue_mon_fri = xl_map[7][2]

    ga_sc_past_4_week_retail_revenue_mon_fri = [loc_data.get("past_4_week_retail_revenue_mon_fri", 0) for loc_data in ga_sc if loc_data]
    ga_sc_past_4_week_retail_revenue_mon_fri_avg = sum(ga_sc_past_4_week_retail_revenue_mon_fri)/4
    current_ga_sc_retail_revenue_mon_fri = xl_map[7][3]

    curr_total_retail_revenue_mon_fri = current_ill_retail_revenue_mon_fri+current_ga_sc_retail_revenue_mon_fri
    past_4_total_retail_revenue_mon_fri = ill_past_4_week_retail_revenue_mon_fri_avg+ga_sc_past_4_week_retail_revenue_mon_fri_avg

    # Retail revenue sat-sun
    ill_past_4_week_retail_revenue_sat_sun = [loc_data.get("past_4_week_retail_revenue_sat_sun", 0) for loc_data in ill if loc_data]
    ill_past_4_week_retail_revenue_sat_sun_avg = sum(ill_past_4_week_retail_revenue_sat_sun)/4
    current_ill_retail_revenue_sat_sun = xl_map[8][2]

    ga_sc_past_4_week_retail_revenue_sat_sun = [loc_data.get("past_4_week_retail_revenue_sat_sun", 0) for loc_data in ga_sc if loc_data]
    ga_sc_past_4_week_retail_revenue_sat_sun_avg = sum(ga_sc_past_4_week_retail_revenue_sat_sun)/4
    current_ga_sc_retail_revenue_sat_sun = xl_map[8][3]

    curr_total_retail_revenue_sat_sun = current_ill_retail_revenue_sat_sun+current_ga_sc_retail_revenue_sat_sun
    past_4_total_retail_revenue_sat_sun = ill_past_4_week_retail_revenue_sat_sun_avg+ga_sc_past_4_week_retail_revenue_sat_sun_avg

    # Total Revenue mon-fri
    ill_past_4_week_total_revenue_mon_fri = [loc_data.get("past_4_week_total_revenue_mon_fri", 0) for loc_data in ill if loc_data]
    ill_past_4_week_total_revenue_mon_fri_avg = sum(ill_past_4_week_total_revenue_mon_fri)/4
    current_ill_total_revenue_mon_fri = xl_map[9][2]

    ga_sc_past_4_week_total_revenue_mon_fri = [loc_data.get("past_4_week_total_revenue_mon_fri", 0) for loc_data in ga_sc if loc_data]
    ga_sc_past_4_week_total_revenue_mon_fri_avg = sum(ga_sc_past_4_week_total_revenue_mon_fri)/4
    current_ga_sc_total_revenue_mon_fri = xl_map[9][3]

    curr_total_total_revenue_mon_fri = current_ill_total_revenue_mon_fri+current_ga_sc_total_revenue_mon_fri
    past_4_total_total_revenue_mon_fri = ill_past_4_week_total_revenue_mon_fri_avg+ga_sc_past_4_week_total_revenue_mon_fri_avg


    # Total revene sat-sun
    ill_past_4_week_total_revenue_sat_sun = [loc_data.get("past_4_week_total_revenue_sat_sun", 0) for loc_data in ill if loc_data]
    ill_past_4_week_total_revenue_sat_sun_avg = sum(ill_past_4_week_total_revenue_sat_sun)/4
    current_ill_total_revenue_sat_sun = xl_map[10][2]

    ga_sc_past_4_week_total_revenue_sat_sun = [loc_data.get("past_4_week_total_revenue_sat_sun", 0) for loc_data in ga_sc if loc_data]
    ga_sc_past_4_week_total_revenue_sat_sun_avg = sum(ga_sc_past_4_week_total_revenue_sat_sun)/4
    current_ga_sc_total_revenue_sat_sun = xl_map[10][3]

    curr_total_total_revenue_sat_sun = current_ill_total_revenue_sat_sun+current_ga_sc_total_revenue_sat_sun
    past_4_total_total_revenue_sat_sun = ill_past_4_week_total_revenue_sat_sun_avg+ga_sc_past_4_week_total_revenue_sat_sun_avg

    # Avg retail visit
    ill_past_4_weeks_retail_car_count_mon_fri = [loc_data.get("past_4_week_retail_car_count_mon_fri", 0) for loc_data in ill if loc_data]
    ill_past_4_weeks_retail_car_count_mon_fri_avg = sum(ill_past_4_weeks_retail_car_count_mon_fri)/4
    
    ill_past_4_weeks_retail_car_count_sat_sun = [loc_data.get("past_4_week_retail_car_count_sat_sun", 0) for loc_data in ill if loc_data]
    ill_past_4_weeks_retail_car_count_sat_sun_avg = sum(ill_past_4_weeks_retail_car_count_sat_sun)/4
    
    ill_past_4_retail_car_count_avg = ill_past_4_weeks_retail_car_count_mon_fri_avg+ill_past_4_weeks_retail_car_count_sat_sun_avg
    
    ill_past_4_retail_revenue_avg = ill_past_4_week_retail_revenue_mon_fri_avg+ill_past_4_week_retail_revenue_sat_sun_avg
    
    ill_past_4_week_avg_ratail_visit = (ill_past_4_retail_revenue_avg) /(ill_past_4_retail_car_count_avg)
    ill_current_avg_ratail_visit = xl_map[12][2]
    
    
    ga_sc_past_4_weeks_retail_car_count_mon_fri = [loc_data.get("past_4_week_retail_car_count_mon_fri", 0) for loc_data in ga_sc if loc_data]
    ga_sc_past_4_retail_car_count_mon_fri_avg = sum(ga_sc_past_4_weeks_retail_car_count_mon_fri)/4
    
    
    ga_sc_past_4_weeks_retail_car_count_sat_sun = [loc_data.get("past_4_week_retail_car_count_sat_sun", 0) for loc_data in ga_sc if loc_data]
    ga_sc_past_4_retail_car_count_sat_sun_avg = sum(ga_sc_past_4_weeks_retail_car_count_sat_sun)/4
    
    ga_sc_past_4_weeks_retail_car_count_avg = ga_sc_past_4_retail_car_count_mon_fri_avg + ga_sc_past_4_retail_car_count_sat_sun_avg
    
    ga_sc_past_4_retail_revenue_avg = ga_sc_past_4_week_retail_revenue_mon_fri_avg+ga_sc_past_4_week_retail_revenue_sat_sun_avg
    
    ga_sc_past_4_week_avg_ratail_visit = (ga_sc_past_4_retail_revenue_avg)/(ga_sc_past_4_weeks_retail_car_count_avg)
    ga_sc_curr_avg_ratail_visit = xl_map[12][3]
    
    past_4_weeks_retail_revenue_total_avg  = ill_past_4_retail_revenue_avg + ga_sc_past_4_retail_revenue_avg
    
    past_4_weeks_retail_car_count_avg      = ill_past_4_retail_car_count_avg  + ga_sc_past_4_weeks_retail_car_count_avg
    
    total_past_4_week_avg_ratail_visit = (past_4_weeks_retail_revenue_total_avg)/(past_4_weeks_retail_car_count_avg)
    total_curr_avg_ratail_visit = xl_map[12][1]

    # past_4_week_labour_hours_mon_fri
    ill_past_4_week_labour_hours_mon_fri = [loc_data.get("past_4_week_labour_hours_mon_fri", 0) for loc_data in ill if loc_data]
    ill_past_4_week_labour_hours_mon_fri_sum = sum(ill_past_4_week_labour_hours_mon_fri)
    ill_past_4_week_labour_hours_mon_fri_avg = ill_past_4_week_labour_hours_mon_fri_sum / 4
    
    ill_current_week_cars_per_labour_hour_mon_fri = xl_map[16][2]
    ' ill_sum_past_mon_fri   is the avg of  car count mon fri past week '
    ill_past_4_weeks_cars_per_labour_hour_mon_fri = ill_sum_past_mon_fri / ill_past_4_week_labour_hours_mon_fri_avg

    ga_sc_past_4_week_labour_hours_mon_fri = [loc_data.get("past_4_week_labour_hours_mon_fri", 0) for loc_data in ga_sc if loc_data]
    ga_sc_past_4_week_labour_hours_mon_fri_sum = sum(ga_sc_past_4_week_labour_hours_mon_fri)
    ga_sc_past_4_week_labour_hours_mon_fri_avg = ga_sc_past_4_week_labour_hours_mon_fri_sum / 4
    
    
    ga_sc_current_week_cars_per_labour_hour_mon_fri  = xl_map[16][3]
    ga_sc_past_4_weeks_cars_per_labour_hour_mon_fri  = ga_sc_sum_past_mon_fri / ga_sc_past_4_week_labour_hours_mon_fri_avg
    
    
    total_past_4_week_cars_mon_fri  = ill_sum_past_mon_fri + ga_sc_sum_past_mon_fri
    total_past_4_weeks_labour_hours_mon_fri = ill_past_4_week_labour_hours_mon_fri_avg + ga_sc_past_4_week_labour_hours_mon_fri_avg
    total_current_cars_per_labour_hour_mon_fri = xl_map[16][1]
    total_past_4_week_cars_per_labour_hour_mon_fri  = total_past_4_week_cars_mon_fri / total_past_4_weeks_labour_hours_mon_fri
    
    

    # past_4_week_labour_hours_sat_sun
    ill_past_4_week_labour_hours_sat_sun = [loc_data.get("past_4_week_labour_hours_sat_sun", 0) for loc_data in ill if loc_data]
    ill_past_4_week_labour_hours_sat_sun_sum = sum(ill_past_4_week_labour_hours_sat_sun)
    ill_past_4_week_labour_hours_sat_sun_avg = ill_past_4_week_labour_hours_sat_sun_sum / 4 

    ill_current_cars_per_labour_hour_sat_sun  = xl_map[17][2]
    'here ill_sum_past_sat_sun is the avg of car count of past 4 weeks '
    ill_past_4_week_cars_per_labour_hour_sat_sun = ill_sum_past_sat_sun / ill_past_4_week_labour_hours_sat_sun_avg

    ga_sc_past_4_week_labour_hours_sat_sun = [loc_data.get("past_4_week_labour_hours_sat_sun", 0) for loc_data in ga_sc if loc_data]
    ga_sc_past_4_week_labour_hours_sat_sun_sum = sum(ga_sc_past_4_week_labour_hours_sat_sun)
    ga_sc_past_4_week_labour_hours_sat_sun_avg = ga_sc_past_4_week_labour_hours_sat_sun_sum / 4 
    
    ga_sc_current_cars_per_labour_hour_sat_sun  =  xl_map[17][3]
    'ga_sc_sum_past_sat_sun is the avg of past 4 week car count'
    ga_sc_past_4_week_cars_per_labour_hour_sat_sun =  ga_sc_sum_past_sat_sun/ ga_sc_past_4_week_labour_hours_sat_sun_avg
    
    
    
    total_current_cars_per_labour_hour_sat_sun = xl_map[17][1]
    total_past_4_week_cars_sat_sun = ill_sum_past_sat_sun + ga_sc_sum_past_sat_sun
    total_past_4_labour_hours_sat_sun = ill_past_4_week_labour_hours_sat_sun_avg + ga_sc_past_4_week_labour_hours_sat_sun_avg
    total_cars_per_labour_hour_sat_sun = total_past_4_week_cars_sat_sun / total_past_4_labour_hours_sat_sun
    

    # Total Cars Per Man Hour
    ill_past_4_week_total_cars_per_man_hour = (ill_sum_past_mon_fri+ill_sum_past_sat_sun)/((ill_past_4_week_labour_hours_mon_fri_avg+ill_past_4_week_labour_hours_sat_sun_avg))
    ill_curr_total_cars_per_man_hour = xl_map[18][2]

    ga_sc_past_4_week_total_cars_per_man_hour = (ga_sc_sum_past_mon_fri+ga_sc_sum_past_sat_sun)/((ga_sc_past_4_week_labour_hours_mon_fri_avg+ga_sc_past_4_week_labour_hours_sat_sun_avg))
    ga_sc_curr_total_cars_per_man_hour = xl_map[18][3]

    total_past_4_week_total_cars_per_man_hour = (past_4_week_total_car_cnt_mon_fri + past_4_week_total_car_cnt_sat_sun) / (total_past_4_weeks_labour_hours_mon_fri +total_past_4_labour_hours_sat_sun) 
    total_curr_total_cars_per_man_hour = xl_map[18][1]
    
    
    colours = darkgreen_format,light_green_format,darkred_format,lightred_format
    print(f"ill avg : {ill_average_percent}")
    print("ill avg revenue:",ill_avg_revenue_change)
    print("ga sc avg revenue :",ga_sc_avg_revenue_change)
    print("total revenue total change:",total_revenue_total_change)
    print(f"ga_sc average :{ga_sc_average_percent}")
    print(f"total_average : {total_average_percent}")
    print("ill conversation change :",ill_conversation_rate_change)
    print("gasc conversation chane :",ga_sc_conversation_change)
    print(f"conversation rate in ga sc :{ga_sc_current_conversation_rate},- {ga_sc_past_4_conversation_rate }")
    print("total conversation change :",total_conversation_change)
    # set_colour(ill_average_percent,7,3,worksheet,colours) #for ill
    set_colour_new(current_ill_week_cnt,ill_sum_past,7,3,worksheet,colours)  #Total cars ill
    # set_colour(ga_sc_average_percent,7,4,worksheet,colours) #for gasc total
    set_colour_new(current_ga_sc_week_cnt,ga_sc_sum_past,7,4,worksheet,colours)
    # set_colour(total_average_percent,7,2,worksheet,colours) #for total total
    set_colour_new(totals_current,totals_past,7,2,worksheet,colours)  #total  cars total
    
    # set_colour(ill_avg_revenue_change,12,3,worksheet,colours) #ill 
    set_colour_new(ill_curent_revenue,ill_total_revenue_avg,12,3,worksheet,colours)
    # set_colour(ga_sc_avg_revenue_change,12,4,worksheet,colours) #ga sc
    set_colour_new(ga_sc_curent_revenue,ga_sc_avg_revenue,12,4,worksheet,colours)  #ga sc revenue 
    # set_colour(total_revenue_total_change,12,2,worksheet,colours)
    set_colour_new(total_reveneu_curent,total_revenue_past_4_avg_total,12,2,worksheet,colours) #totalk revenue 
    print("Conversion Rate")
    set_colour_new(ill_current_conversation_rate,ill_past_4_conversation_rate,21,3,worksheet,colours)  #conversation rate
    set_colour_new(ga_sc_current_conversation_rate,ga_sc_past_4_conversation_rate,21,4,worksheet,colours)
    set_colour_new(current_total_conversation_rate,past_total_conversation_change,21,2,worksheet,colours)

    print('Car count mon fri')
    set_colour_new(current_ill_week_cnt_mon_fri, ill_sum_past_mon_fri, 3, 3, worksheet, colours) #car count mon-fri ill
    set_colour_new(current_ga_sc_week_cnt_mon_fri, ga_sc_sum_past_mon_fri, 3, 4, worksheet, colours) #car count mon-fri ga-sc
    set_colour_new(curr_total_car_cnt_mon_fri, past_4_week_total_car_cnt_mon_fri, 3, 2, worksheet, colours) # total car count mon-fri

    print('Car count sat sun')
    set_colour_new(current_ill_week_cnt_sat_sun, ill_sum_past_sat_sun, 4, 3, worksheet, colours) #car count sat-sun ill
    set_colour_new(current_ga_sc_week_cnt_sat_sun, ga_sc_sum_past_sat_sun, 4, 4, worksheet, colours) #car count sat-sun ga-sc
    set_colour_new(curr_total_car_cnt_sat_sun, past_4_week_total_car_cnt_sat_sun, 4, 2, worksheet, colours) # total car count sat-sun

    print('Retail count mon fri')
    set_colour_new(current_ill_week_retail_cnt_mon_fri, ill_sum_retail_car_cnt_mon_fri, 5, 3, worksheet, colours) # retail count mon-fri ill
    set_colour_new(current_ga_sc_week_retail_cnt_mon_fri, ga_sc_sum_retail_car_cnt_mon_fri, 5, 4, worksheet, colours) # retail count mon-fri
    set_colour_new(curr_total_retail_car_cnt_mon_fri, past_4_week_total_retail_car_cnt_mon_fri, 5, 2, worksheet, colours) # retail count mon-fri
    
    print('Retail count sat sun')
    set_colour_new(current_ill_week_retail_cnt_sat_sum, ill_sum_retail_car_cnt_sat_avg, 6, 3, worksheet, colours) # retail count sat-sun ill
    set_colour_new(current_ga_sc_week_retail_cnt_mon_fri, ga_sc_sum_retail_car_cnt_mon_fri, 6, 4, worksheet, colours) # retail count sat-sun
    set_colour_new(curr_total_retail_car_cnt_sat_sun, past_4_week_total_retail_car_cnt_sat_sun, 6, 2, worksheet, colours) # retail count sat-sun

    print('Retail Revenue count mon fri')
    set_colour_new(current_ill_retail_revenue_mon_fri, ill_past_4_week_retail_revenue_mon_fri_avg, 8, 3, worksheet, colours)
    set_colour_new(current_ga_sc_retail_revenue_mon_fri, ga_sc_past_4_week_retail_revenue_mon_fri_avg, 8, 4, worksheet, colours)
    set_colour_new(curr_total_retail_revenue_mon_fri, past_4_total_retail_revenue_mon_fri, 8, 2, worksheet, colours)
    
    print('Retail Revenue count sat sun')
    set_colour_new(current_ill_retail_revenue_sat_sun, ill_past_4_week_retail_revenue_sat_sun_avg, 9, 3, worksheet, colours) 
    set_colour_new(current_ga_sc_retail_revenue_sat_sun, ga_sc_past_4_week_retail_revenue_sat_sun_avg, 9, 4, worksheet, colours) 
    set_colour_new(curr_total_retail_revenue_sat_sun, past_4_total_retail_revenue_sat_sun, 9, 2, worksheet, colours)

    print('Total Revenue count Mon Fri')
    set_colour_new(current_ill_total_revenue_mon_fri, ill_past_4_week_total_revenue_mon_fri_avg, 10, 3, worksheet, colours)
    set_colour_new(current_ga_sc_total_revenue_mon_fri, ga_sc_past_4_week_total_revenue_mon_fri_avg, 10, 4, worksheet, colours) 
    set_colour_new(curr_total_total_revenue_mon_fri, past_4_total_total_revenue_mon_fri, 10, 2, worksheet, colours)
    
    print('Total Revenue count Sat Sun')
    set_colour_new(current_ill_total_revenue_sat_sun, ill_past_4_week_total_revenue_sat_sun_avg, 11, 3, worksheet, colours) 
    set_colour_new(current_ga_sc_total_revenue_sat_sun, ga_sc_past_4_week_total_revenue_sat_sun_avg, 11, 4, worksheet, colours) 
    set_colour_new(curr_total_total_revenue_sat_sun, past_4_total_total_revenue_sat_sun, 11, 2, worksheet, colours) 

    print('Avg retail visit')
    set_colour_for_avg_retail(ill_current_avg_ratail_visit, ill_past_4_week_avg_ratail_visit, 13, 3, worksheet, colours)
    set_colour_for_avg_retail(ga_sc_curr_avg_ratail_visit, ga_sc_past_4_week_avg_ratail_visit, 13, 4, worksheet, colours)
    set_colour_for_avg_retail(total_curr_avg_ratail_visit, total_past_4_week_avg_ratail_visit, 13, 2, worksheet, colours)

    print('Cars Per Labor Hour Mon - Fri')
    set_colour_new(ill_current_week_cars_per_labour_hour_mon_fri, ill_past_4_weeks_cars_per_labour_hour_mon_fri, 17, 3, worksheet, colours) 
    set_colour_new(ga_sc_current_week_cars_per_labour_hour_mon_fri, ga_sc_past_4_weeks_cars_per_labour_hour_mon_fri, 17, 4, worksheet, colours) 
    set_colour_new(total_current_cars_per_labour_hour_mon_fri, total_past_4_week_cars_per_labour_hour_mon_fri, 17, 2, worksheet, colours)

    print('Cars Per Labor Hour Sat & Sun')
    set_colour_new(ill_current_cars_per_labour_hour_sat_sun, ill_past_4_week_cars_per_labour_hour_sat_sun, 18, 3, worksheet, colours) 
    set_colour_new(ga_sc_current_cars_per_labour_hour_sat_sun, ga_sc_past_4_week_cars_per_labour_hour_sat_sun, 18, 4, worksheet, colours) 
    set_colour_new(total_current_cars_per_labour_hour_sat_sun, total_cars_per_labour_hour_sat_sun, 18, 2, worksheet, colours)

    #Total Cars Per Man Hour
    set_colour_new(ill_curr_total_cars_per_man_hour, ill_past_4_week_total_cars_per_man_hour, 19, 3, worksheet, colours)
    set_colour_new(ga_sc_curr_total_cars_per_man_hour, ga_sc_past_4_week_total_cars_per_man_hour, 19, 4, worksheet, colours)
    set_colour_new(total_curr_total_cars_per_man_hour, total_past_4_week_total_cars_per_man_hour, 19, 2, worksheet, colours)

    print()
    
    
    loc_names = ["Sudz - Beverly",'Fuller-Calumet',
                "Fuller-Cicero","Fuller-Matteson","Fuller-Elgin",
                "Splash-Peoria","Getaway-Macomb","Getaway-Morton",
                "Getaway-Ottawa","Getaway-Peru","Sparkle-Belair",
                "Sparkle-Evans","Sparkle-Furrys Ferry","Sparkle-Greenwood",
                "Sparkle-Grovetown 1","Sparkle-Grovetown 2","Sparkle-North Augusta",
                "Sparkle-Peach Orchard","Sparkle-Windsor Spring"]
    
    for index,place_dictionary in enumerate(all_locations):
        #Total Cars
        current_week_total_cars = xl_map[6][index+4]
        past_4_week_total_cars = place_dictionary.get("past_4_week_cnt")
        change_in_total_car_count_percent  = chnage_total_car_count_fun(current_week_total_cars,past_4_week_total_cars)
        past_4_week_total_cars_avg  = do_avg(past_4_week_total_cars)
        #set_colour(change_in_total_car_count_percent,7,index+5,worksheet,colours) #for total cars 
        set_colour_new(current_week_total_cars,past_4_week_total_cars_avg,7,index+5,worksheet,colours)
        
        print(f"{loc_names[index]}=>chnage car count  {change_in_total_car_count_percent}")
        current_week_conversatio_rate = place_dictionary.get("conversion_rate")
        past_4_week_conversation_rate =  place_dictionary.get("past_4_week_conversion_rate")
        
        set_colour_new(current_week_conversatio_rate,past_4_week_conversation_rate,21,index+5,worksheet,colours) #conversation rate colours
        #Total Revenue
        #print(f"{loc_names[index]}=>chnage conversation rate   {change_in_conversationrate}")
        current_revenue_total = place_dictionary.get("total_revenue")
        past_4_week_revenue_total = place_dictionary.get("past_4_weeks_total_revenue")
        past_4_week_revenue_total_avg = do_avg(past_4_week_revenue_total)
        
        change_in_total_revenue = chnage_total_revenue_fun(current_revenue_total,past_4_week_revenue_total)
        print(f"{loc_names[index]}=>chnage total revenue    {change_in_total_revenue}")
        set_colour_new(current_revenue_total, past_4_week_revenue_total_avg,12, index+5,worksheet,colours)
        # set_colour(change_in_total_revenue,12,index+5,worksheet,colours)

        #Car Count Mon - Fri
        past_4_car_count_mon_fri_avg = place_dictionary.get('past_4_week_car_cnt_mon_fri', 0)/4
        curr_week_car_count_mon_fri = xl_map[2][index+4]
        set_colour_new(curr_week_car_count_mon_fri, past_4_car_count_mon_fri_avg, 3, index+5, worksheet, colours)

        past_4_car_count_sat_sun_avg = place_dictionary.get('past_4_week_car_cnt_sat_sun', 0)/4
        curr_week_car_count_sat_sun = xl_map[3][index+4]
        set_colour_new(curr_week_car_count_sat_sun, past_4_car_count_sat_sun_avg, 4, index+5, worksheet, colours)

        past_4_week_retail_car_count_mon_fri_avg = place_dictionary.get('past_4_week_retail_car_count_mon_fri', 0)/4
        curr_week_retail_car_count_sat_sun = xl_map[4][index+4]
        set_colour_new(curr_week_retail_car_count_sat_sun, past_4_week_retail_car_count_mon_fri_avg, 5, index+5, worksheet, colours)
        
        past_4_week_retail_car_count_sat_sun_avg = place_dictionary.get('past_4_week_retail_car_count_sat_sun', 0)/4
        curr_week_retail_car_count_sat_sun = xl_map[5][index+4]
        set_colour_new(curr_week_retail_car_count_sat_sun, past_4_week_retail_car_count_sat_sun_avg, 6, index+5, worksheet, colours)

        past_4_week_retail_revenue_mon_fri_avg = place_dictionary.get('past_4_week_retail_revenue_mon_fri', 0)/4
        curr_week_retail_revenue_sat_sun = xl_map[7][index+4]
        set_colour_new(curr_week_retail_revenue_sat_sun, past_4_week_retail_revenue_mon_fri_avg, 8, index+5, worksheet, colours)

        past_4_week_retail_revenue_sat_sun_avg = place_dictionary.get('past_4_week_retail_revenue_sat_sun', 0)/4
        curr_week_retail_revenue_sat_sun = xl_map[8][index+4]
        set_colour_new(curr_week_retail_revenue_sat_sun, past_4_week_retail_revenue_sat_sun_avg, 9, index+5, worksheet, colours)

        #Total Revenue Mon - Fri
        past_4_week_total_revenue_mon_fri_avg = place_dictionary.get('past_4_week_total_revenue_mon_fri', 0)/4
        curr_week_total_revenue_mon_fri = xl_map[9][index+4]
        set_colour_new(curr_week_total_revenue_mon_fri, past_4_week_total_revenue_mon_fri_avg, 10, index+5, worksheet, colours)

        #Total Revenue Sat - Sun
        print("Total Revenue Sat - Sun")
        past_4_week_total_revenue_sat_sun_avg = place_dictionary.get('past_4_week_total_revenue_sat_sun', 0)/4
        curr_week_total_revenue_sat_sun = xl_map[10][index+4]
        set_colour_new(curr_week_total_revenue_sat_sun, past_4_week_total_revenue_sat_sun_avg, 11, index+5, worksheet, colours)

        #Avg. Retail Visit
        past_retail_car_count_mon_fri_avg = place_dictionary.get('past_4_week_retail_car_count_mon_fri', 0)/4
        past_retail_car_count_sat_sun_avg = place_dictionary.get('past_4_week_retail_car_count_sat_sun', 0)/4
        
        past_retail_car_count_avg  = past_retail_car_count_mon_fri_avg  + past_retail_car_count_sat_sun_avg
        
        print("Avg. Retail Visit")
        past_4_week_avg_ratail_visit = (past_4_week_retail_revenue_mon_fri_avg+past_4_week_retail_revenue_sat_sun_avg)/(past_retail_car_count_avg)
        curr_avg_ratail_visit = xl_map[12][index+4]
        set_colour_new(curr_avg_ratail_visit, past_4_week_avg_ratail_visit, 13, index+5, worksheet, colours)

        #Cars Per Labor Hour Mon - Fri
        past_4_week_labour_hours_mon_fri_avg = handle_zero_divison(place_dictionary.get('past_4_week_car_cnt_mon_fri', 0),place_dictionary.get('past_4_week_labour_hours_mon_fri', 0))
        if xl_map[16][index+4] != '':
            curr_week_labour_hours_mon_fri_avg = xl_map[16][index+4]
            set_colour_new(curr_week_labour_hours_mon_fri_avg, past_4_week_labour_hours_mon_fri_avg, 17, index+5, worksheet, colours)

        past_4_week_labour_hours_sat_sun_avg = handle_zero_divison(place_dictionary.get('past_4_week_car_cnt_sat_sun', 0),place_dictionary.get('past_4_week_labour_hours_sat_sun', 0))
        if xl_map[17][index+4] != '':
            curr_week_labour_hours_sat_sun_avg = xl_map[17][index+4]
            set_colour_new(curr_week_labour_hours_sat_sun_avg, past_4_week_labour_hours_sat_sun_avg, 18, index+5, worksheet, colours)

        # Total Cars Per Man Hour
        if xl_map[18][index+4] != '':
            past_4_week_total_car_count = place_dictionary.get('past_4_week_car_cnt_mon_fri', 0)+place_dictionary.get('past_4_week_car_cnt_sat_sun', 0)
            past_4_week_total_man_hour = place_dictionary.get('past_4_week_labour_hours_mon_fri', 0)+place_dictionary.get('past_4_week_labour_hours_sat_sun', 0)
            past_4_week_total_cars_per_man_hour = past_4_week_total_car_count/past_4_week_total_man_hour
            curr_total_cars_per_man_hour = xl_map[18][index+4]
            set_colour_new(curr_total_cars_per_man_hour, past_4_week_total_cars_per_man_hour, 19, index+5, worksheet, colours)

        
        print("\n"*2)
        
    
    

    #applying bold font
    # Define a bold font style
    bold_font = Font(bold=True)
    
    for row in worksheet.iter_rows(min_row=7, max_row=7, min_col=1, max_col=4):
        for cell in row:
            cell.font = bold_font

    for row in worksheet.iter_rows(min_row=12, max_row=13, min_col=1, max_col=4):
        for cell in row:
            cell.font = bold_font
            
    for row in worksheet.iter_rows(min_row=19, max_row=19, min_col=1, max_col=4):
        for cell in row:
            cell.font = bold_font
            
    for row in worksheet.iter_rows(min_row=21, max_row=21, min_col=1, max_col=4):
        for cell in row:
            cell.font = bold_font
    # Save the modified workbook
    workbook.save(filename)




# print(data_path)

# Your existing code

def get_week_dates_for_storage():
    "will retun in '%Y' ==> 2024"
    # Get the current date
    today = datetime.today()
    
    # Find the current week's Monday date
    current_week_monday = today - timedelta(days=today.weekday())
    
    # Find the current week's Sunday date
    current_week_sunday = current_week_monday + timedelta(days=6)
    
    # Format the dates in dd/mm/yyyy format
    #monday_date_str = current_week_monday.strftime("%m/%d/%Y")
    sunday_date_str = current_week_sunday.strftime("%Y") #%m_%Y
    
    #old ret f"{monday_date_str}-{sunday_date_str}".replace('/','_')
    return sunday_date_str



def create_storage_directory(path):
    created_path=None
    try:
        path = os.path.join(data_path,path)
        os.makedirs(path, exist_ok=True)
        print(f"Directory '{path}' created successfully")
        created_path =path
    except OSError as error:
        print(f"Directory '{path}' cannot be created: {error}")
    
    return created_path



if __name__=="__main__":
    import time 
    start = time.time()

    load_dotenv()

    class emailConfig:
        env_vars    = os.environ
        FROM_EMAIL   = env_vars.get("FROM_EMAIL")
        FROM_NAME      = env_vars.get("FROM_NAME")
        SMTP_SERVER   = env_vars.get("SMTP_SERVER")
        SMTP_PORT = env_vars.get("SMTP_PORT")
        SMTP_USER=env_vars.get("SMTP_USER")
        SMTP_PASSWORD=env_vars.get("SMTP_PASSWORD")
        TO_EMAIL=env_vars.get("TO_EMAIL")
    
    # Configuration
    subject = 'Weekly reports'
    body = 'This is the body of the email.'
    to_email = emailConfig.TO_EMAIL
    from_email = emailConfig.FROM_EMAIL
    from_name = emailConfig.FROM_NAME
    smtp_server = emailConfig.SMTP_SERVER
    smtp_port = emailConfig.SMTP_PORT
    smtp_user = emailConfig.SMTP_USER
    smtp_password = emailConfig.SMTP_PASSWORD
    
    
    #cc_emails=["CR@SparkleCW.com","FZ@SparkleCW.com","Rick@SparkleStatus.com","Shane@SparkleStatus.com", "mgiamalis@firmament.com"]
    
    # # -----------------Actual script  ----------------------------#
    
    path = get_week_dates_for_storage()
    path="test_10_2" #"07-2024"#"06-2024"
    storage_path = create_storage_directory(path)
    monday_date_str, friday_date_str, saturday_date_str, sunday_date_str = sitewatch_week_dates()
    # print(monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    monday_date_str="2024-07-01"
    friday_date_str = "2024-07-05"
    saturday_date_str = "2024-07-06"
    sunday_date_str="2024-07-07"  #Y-M-D
    
    # sitewatch_report = sitewatch_week_report("",monday_date_str,friday_date_str,saturday_date_str, sunday_date_str)
    
    # monday_date_str, friday_date_str, saturday_date_str, sunday_date_str =  washify_week_dates()
    
    ##testing dates 
    monday_date_str =  "07/01/2024"
    friday_date_str =  "07/05/2024"
    saturday_date_str = "07/06/2024"
    sunday_date_str  =  "07/07/2024"  #M/D/Y
    
    print(monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    # washify_report = washify_week_report("", monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    
    # #for hamilton dates
    monday_date_str = "2024-07-01"
    friday_date_str = "2024-07-05"
    saturday_date_str = "2024-07-06"
    sunday_date_str  = "2024-07-07"
    
    # # monday_date_str, friday_date_str, saturday_date_str, sunday_date_str = hamilton_week_dates()
    # hamilton_report = hamilton_week_report(monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    
    # data = sitewatch_report
    
    # data.update(washify_report)
    
    # data.update(hamilton_report)

    # with open("all_data_t1.json", 'w') as f:
    #     json.dump(data, f, indent=4)

    with open("all_data_t1.json", 'r') as f:
        data = json.load(f)

    
    comment =f"Ending {sunday_date_str}"
    sheet_name= sunday_date_str.replace("/","-")
    filename=f"{path}.xlsx"
    file_name_with_fullpath = os.path.join(storage_path,filename)
    prepare_xlmap(data,comment,sheet_name=sheet_name,filename=file_name_with_fullpath)
    
    # Directory containing Excel files
    directory_path = storage_path
    attachments = get_excel_files(directory_path)
    
    #Sending email to email address
    body = f'weekly report Ending {sunday_date_str}'
    #send_email(subject, body, to_email, from_email, from_name, smtp_server, smtp_port, smtp_user, smtp_password, attachments)#
    
    #prepare_xlmap(data,comment,sheet_name=sheet_name)
    # # -----------------Actual script  ----------------------------#
    print(f"\nTotal Time took:{time.time()-start}")
    



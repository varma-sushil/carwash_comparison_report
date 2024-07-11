
import sys
import os
from dotenv import load_dotenv
from datetime import datetime, timedelta
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, Font
import json

try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side




# Add the path to the parent directory of "washify" to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'washify')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'sitewash')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'hamilton')))

# print(sys.path)

from washify_weekly import get_week_dates as washify_week_dates
from washify_weekly import generate_weekly_report  as washify_week_report

from sitewatch_weekly import get_week_dates as sitewatch_week_dates
from sitewatch_weekly import generate_weekly_report as sitewatch_week_report

from hamilton_weekly  import get_week_dates as hamilton_week_dates
from hamilton  import generate_report as hamilton_week_report

from custom_mailer import send_email,get_excel_files

current_file_path = os.path.dirname(os.path.abspath(__file__))
# print(current_file_path)

data_path = os.path.join(current_file_path,"data")

def do_sum(xl_map,start_index,range):
    "This will do row based some "
    total=0
    for i in range:
        val = xl_map[start_index][i+1]
        if isinstance(val,float) or isinstance(val,int):
            total+=val
    
    return total

def do_sum_location(xl_map,location:list):
    "This will take array of row ,col"
    total =0
    for row,col in location:
        val = xl_map[row][col]
        if isinstance(val,float) or isinstance(val,int):
            total+=val
    return total

def Average_retail_visit__GA_SC_fucntion(retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC,
                                         retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC):    
    
    result = 0
    try:
        Average_retail_visit__GA_SC = sum([retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC])/sum(
        [retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC ]
    )
        result = Average_retail_visit__GA_SC
    except Exception as e:
        print(f"Exception in Average_retail_visit__GA_SC_fucntion() {e}")
        

    return result 

def Average_memeber_visit_GA_SC_function(Total_revenue_GA_SC,
        retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC,
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC,
        total_cars_GA_SC
        ):
    result = 0
    
    try:
        
        Average_memeber_visit_GA_SC = (Total_revenue_GA_SC - sum([retail_revenue_monday_to_friday_GA_SC , retail_revenue_saturday_to_sunday_GA_SC]))/(total_cars_GA_SC - sum([
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC]))
        
        result = Average_memeber_visit_GA_SC
        
    except Exception as e:
        print(f"Exception in Average_memeber_visit_GA_SC_function() {e}")
    
    return result

def cost_per_labour_hour_monday_to_friday_GA_SC_fucntion(car_count_monday_to_friday_GA_SC,staff_hours_monday_to_friday_GA_SC):
    result = 0
    try:
        cost_per_labour_hour_monday_to_friday_GA_SC = car_count_monday_to_friday_GA_SC/staff_hours_monday_to_friday_GA_SC
        result =cost_per_labour_hour_monday_to_friday_GA_SC
    except Exception as e:
        print(f"Exception in cost_per_labour_hour_monday_to_friday_GA_SC() {e}")

    return result


def cost_per_labour_hour_saturday_to_sunday_GA_SC_function(car_count_saturday_to_sunday_GA_SC,staff_hours_saturday_to_sunday_GA_SC):
    result = 0
    
    try:
        cost_per_labour_hour_saturday_to_sunday_GA_SC = car_count_saturday_to_sunday_GA_SC/staff_hours_saturday_to_sunday_GA_SC
        result = cost_per_labour_hour_saturday_to_sunday_GA_SC
    except Exception as e:
        print(f"Exception in staff_hours_saturday_to_sunday_GA_SC_fucntion() {e}")
        
    return result

def Total_cars_per_man_hour_GA_SC_function(total_cars_GA_SC, staff_hours_monday_to_friday_GA_SC,staff_hours_saturday_to_sunday_GA_SC):
    result =0 
    try:
            Total_cars_per_man_hour_GA_SC = total_cars_GA_SC/sum([
            staff_hours_monday_to_friday_GA_SC,staff_hours_saturday_to_sunday_GA_SC])
            result = Total_cars_per_man_hour_GA_SC
    
    except Exception as e:
        print(f"Exception in  Total_cars_per_man_hour_GA_SC_function() {e}")
    
    return result
 
def Conversion_rate_GA_SC_function(Total_club_plans_sold_GA_SC,retail_car_count_monday_to_friday_GA_SC,
                                   retail_car_count_saturday_to_sunday_GA_SC):
    result = 0 
    try:
        Conversion_rate_GA_SC = Total_club_plans_sold_GA_SC/sum([
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC])     
        
        result =  Conversion_rate_GA_SC
    except Exception as e:
        print(f"Exceptionin Conversion_rate_GA_SC_function() {e}")
    
    return result

def Conversion_rate_Total_function(Total_club_plans_sold_Total,
                                   retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total):
    result = 0
    
    try:
        
        Conversion_rate_Total = Total_club_plans_sold_Total/sum(
        [retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total])  
        
        result = Conversion_rate_Total
    except Exception as e:
        print(f"Exception in Conversion_rate_Total_function()  {e}") 
    
    return result

def Conversion_rate_ILL_function(Total_club_plans_sold_ILL,
                                 retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Conversion_rate_ILL = Total_club_plans_sold_ILL/sum([retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL])         
        result = Conversion_rate_ILL
    except Exception as e:
        print(f"Exception in Conversion_rate_ILL_function() {e}")
        
    return result

def Total_cars_per_man_hour_total_function(Total_cars_Total,
                                           staff_hours_monday_to_friday_Total,
                                           staff_hours_saturday_to_sunday_Total):
    result = 0
    
    try:
        Total_cars_per_man_hour_total = Total_cars_Total/sum(
        [staff_hours_monday_to_friday_Total,staff_hours_saturday_to_sunday_Total])
        
        result = Total_cars_per_man_hour_total
    
    except Exception as e:
        print(f"Exception in Total_cars_per_man_hour_total_function() {e}")
        
    return result


def Total_cars_per_man_hour_ILL_function(total_cars_in_ILL,
                                         staff_hours_monday_to_friday_ILL,
                                         staff_hours_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Total_cars_per_man_hour_ILL = total_cars_in_ILL/(sum([
        staff_hours_monday_to_friday_ILL+staff_hours_saturday_to_sunday_ILL]))

        result = Total_cars_per_man_hour_ILL
    
    except Exception as e:
        print(f"Exception Total_cars_per_man_hour_ILL_function() {e}")
    
    return result


def cost_per_labour_hour_saturday_to_sunday_ILL_function(car_count_saturday_to_sunday_ILL,staff_hours_saturday_to_sunday_ILL):
    result = 0
    
    try:
        cost_per_labour_hour_saturday_to_sunday_ILL = car_count_saturday_to_sunday_ILL/staff_hours_saturday_to_sunday_ILL
        
        result = cost_per_labour_hour_saturday_to_sunday_ILL
    
    except Exception as e:
        print(f"Exception cost_per_labour_hour_saturday_to_sunday_ILL_function() {e}")
        
    return result


def cost_per_labour_hour_monday_to_friday_Total_function(car_count_monday_to_friday_Total,staff_hours_monday_to_friday_Total):
    result = 0
    
    try:
        cost_per_labour_hour_monday_to_friday_Total = car_count_monday_to_friday_Total/staff_hours_monday_to_friday_Total
        result = cost_per_labour_hour_monday_to_friday_Total
    
    except Exception as e:
        print(f"Exception cost_per_labour_hour_monday_to_friday_Total_function() {e}")
    
    return result

def cost_per_labour_hour_monday_to_friday_ILL_function(car_count_monday_to_friday_ILL,staff_hours_monday_to_friday_ILL):
    result = 0
    
    try:
        cost_per_labour_hour_monday_to_friday_ILL = car_count_monday_to_friday_ILL/staff_hours_monday_to_friday_ILL
        result =  cost_per_labour_hour_monday_to_friday_ILL
    
    except Exception as e:
        print(f"Exception cost_per_labour_hour_monday_to_friday_ILL_function() {e}")
        
    return result

def Average_memeber_visit_Total_function(Total_revenue_Total,retail_revenue_monday_to_friday_Total
                        ,retail_revenue_saturday_to_sunday_Total,Total_cars_Total,retail_car_count_monday_to_friday_Total,
                        retail_car_count_saturday_to_sunday_Total):
    result = 0
    
    try:
        Average_memeber_visit_Total = (Total_revenue_Total -sum([retail_revenue_monday_to_friday_Total,retail_revenue_saturday_to_sunday_Total]))/(Total_cars_Total - sum([
        retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total]))
        result =  Average_memeber_visit_Total
    
    except Exception as e:
        print(f"Exception Average_memeber_visit_Total_function() {e}")  
    
    return result    

def Average_memeber_visit_ILL_function(Total_revenue_ILL,retail_revenue_monday_to_friday_ILL,
                                       retail_revenue_saturday_to_sunday_ILL,total_cars_in_ILL,
                                       retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Average_memeber_visit_ILL = (Total_revenue_ILL - sum([retail_revenue_monday_to_friday_ILL,retail_revenue_saturday_to_sunday_ILL]))/(total_cars_in_ILL - sum([
        retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL]))
        
        result = Average_memeber_visit_ILL
    
    except Exception as e:
        print(f"Exception Average_memeber_visit_ILL_function() {e}")
    
    return result

def Average_retail_visit_Total_function(retail_revenue_monday_to_friday_Total,
                                        retail_revenue_saturday_to_sunday_Total,
                                        retail_car_count_monday_to_friday_Total,
                                        retail_car_count_saturday_to_sunday_Total):
    result = 0
    
    try:
        Average_retail_visit_Total = sum([retail_revenue_monday_to_friday_Total,retail_revenue_saturday_to_sunday_Total])/sum([
        retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total])
        
        result = Average_retail_visit_Total
    
    except Exception as e:
        print(f"Exception Average_retail_visit_Total_function() {e}")
        
    return result
            

def Average_retail_visit_IL_function(retail_revenue_monday_to_friday_ILL,
                                     retail_revenue_saturday_to_sunday_ILL,
                                     retail_car_count_monday_to_friday_ILL,
                                     retail_car_count_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Average_retail_visit_IL = sum([retail_revenue_monday_to_friday_ILL,retail_revenue_saturday_to_sunday_ILL])/sum(
        [retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL])

        result = Average_retail_visit_IL
    
    except Exception as e:
        print(f"Exception Average_retail_visit_IL_function() {e}")
    
    return result

def add_commas(value):
    # Check if the value is an integer or can be converted to an integer
    try:
        num = float(value)
        if num >= 1000:
            return "{:,}".format(num)
    except ValueError:
        pass
    
    # Return the original value if it's not an integer or less than 1000
    return value                

#new xl maps functions 
def update_place_to_xlmap(xl_map,place_index,place_dictionary)->list:
    "Will return updates place dictionary"
    xl_map[2][place_index] = place_dictionary.get("car_count_monday_to_friday")
    xl_map[3][place_index]=place_dictionary.get("car_count_saturday_sunday")
    xl_map[4][place_index]=place_dictionary.get("retail_car_count_monday_to_friday")
    xl_map[5][place_index]=place_dictionary.get("retail_car_count_saturday_sunday")
    
    xl_map[7][place_index]=place_dictionary.get("retail_revenue_monday_to_friday")
    xl_map[8][place_index]=place_dictionary.get("retail_revenue_saturday_sunday")
    xl_map[9][place_index]=place_dictionary.get("total_revenue_monday_to_friday")
    xl_map[10][place_index]=place_dictionary.get("total_revenue_saturday_sunday")
    
    xl_map[14][place_index]=place_dictionary.get("labour_hours_monday_to_friday")
    xl_map[15][place_index]=place_dictionary.get("labour_hours_saturday_sunday")
    xl_map[16][place_index]=place_dictionary.get("cars_per_labour_hour_monday_to_friday")
    xl_map[17][place_index]=place_dictionary.get("cars_per_labour_hour_saturday_sunday")
    
    xl_map[19][place_index] = place_dictionary.get("arm_plans_sold_cnt")
    xl_map[20][place_index]= place_dictionary.get("conversion_rate")
    xl_map[21][place_index] = place_dictionary.get("total_arm_planmembers_cnt")
        
    return xl_map

def prepare_xlmap(data,comment="The comment section",filename="test.xlsx",sheet_name="sheet1"):
    # Load the existing workbook using openpyxl
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.create_sheet(sheet_name)
    except Exception as _:
        print(f"creating neww xl file !! {filename}")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
    
    
    xl_map = [
    [""],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ]
    
    cols = ["    ","Totals","ILL",
                "GA / SC","Sudz - Beverly",'Fuller-Calumet',
                "Fuller-Cicero","Fuller-Matteson","Fuller-Elgin",
                "Splash-Peoria","Getaway-Macomb","Getaway-Morton",
                "Getaway-Ottawa","Getaway-Peru","Sparkle-Belair",
                "Sparkle-Evans","Sparkle-Furrys Ferry","Sparkle-Greenwood",
                "Sparkle-Grovetown 1","Sparkle-Grovetown 2","Sparkle-North Augusta",
                "Sparkle-Peach Orchard","Sparkle-Windsor Spring"]

    index_names =["Car Count Mon - Fri","Car Count Sat - Sun","Retail Car Count Mon - Fri",
                "Retail Car Count Sat - Sun","Total Cars","Retail Revenue Mon - Fri",
                "Retail Revenue Sat - Sun","Total Revenue Mon - Fri","Total Revenue Sat - Sun",
                "Total Revenue","Avg. Retail Visit","Avg. Member Visit",
                "Staff Hours Mon - Fri","Staff Hours Sat - Sun","Cars Per Labor Hour Mon - Fri",
                "Cars Per Labor Hour Sat & Sun","Total Cars Per Man Hour","Total Club Plans Sold",
                "Conversion Rate","Total Club Plan Members"]

    xl_map[0][0] = comment #to maps
    
    
    
    for col in range(len(cols)): #column names to maps
        xl_map[1][col]=cols[col]
        
    for index  in range(len(index_names)):  #index names to map
        xl_map[index+2][0]=index_names[index]
    
    Fuller_Cicero = data.get("Fuller-Cicero")  
    
    if Fuller_Cicero:
        Fuller_Cicero_index=6
        update_place_to_xlmap(xl_map,Fuller_Cicero_index,Fuller_Cicero)
    Fuller_Matteson = data.get("Fuller-Matteson")
    
    if Fuller_Matteson:
        Fuller_Matteson_index=7
        update_place_to_xlmap(xl_map,Fuller_Matteson_index,Fuller_Matteson)
    
    Fuller_Elgin = data.get("Fuller-Elgin")
    
    if Fuller_Elgin:
        Fuller_Elgin_index = 8
        update_place_to_xlmap(xl_map,Fuller_Elgin_index,Fuller_Elgin)
        
      
    Sparkle_Belair = data.get("Sparkle-Belair")
    
    if Sparkle_Belair:
        Sparkle_Belair_index=14

        update_place_to_xlmap(xl_map,Sparkle_Belair_index,Sparkle_Belair)
        
    Sparkle_Evans = data.get("Sparkle-Evans")
    
    if Sparkle_Evans:
        Sparkle_Evans_index=15

        update_place_to_xlmap(xl_map,Sparkle_Evans_index,Sparkle_Evans)

    Sparkle_North_Augusta = data.get("Sparkle-North Augusta")
    
    if Sparkle_North_Augusta:
        Sparkle_North_Augusta_index=20

        update_place_to_xlmap(xl_map,Sparkle_North_Augusta_index,Sparkle_North_Augusta)
    
    Sparkle_Greenwood = data.get("Sparkle-Greenwood")
    
    if Sparkle_Greenwood:
        Sparkle_Greenwood_index=17

        update_place_to_xlmap(xl_map,Sparkle_Greenwood_index,Sparkle_Greenwood)
    
    
    Sparkle_Grovetown_1 = data.get("Sparkle-Grovetown 1")
    
    if Sparkle_Grovetown_1:
        Sparkle_Grovetown_1_index=18

        update_place_to_xlmap(xl_map,Sparkle_Grovetown_1_index,Sparkle_Grovetown_1)
        
    Sparkle_Windsor_Spring = data.get("Sparkle-Windsor Spring")    
    
    if Sparkle_Windsor_Spring:
        Sparkle_Windsor_Spring_index=22 
        
        update_place_to_xlmap(xl_map,Sparkle_Windsor_Spring_index,Sparkle_Windsor_Spring)
        
    Sparkle_Furrys_Ferry = data.get("Sparkle-Furrys Ferry")
    
    if Sparkle_Furrys_Ferry:
        Sparkle_Furrys_Ferry_index= 16
        
        update_place_to_xlmap(xl_map,Sparkle_Furrys_Ferry_index,Sparkle_Furrys_Ferry)      

    Sparkle_Peach_Orchard = data.get("Sparkle-Peach Orchard")
    
    if Sparkle_Peach_Orchard:
        Sparkle_Peach_Orchard_index=21
        
        update_place_to_xlmap(xl_map,Sparkle_Peach_Orchard_index,Sparkle_Peach_Orchard)
        
    Sparkle_Grovetown_2 =  data.get("Sparkle-Grovetown 2")   
      
    if Sparkle_Grovetown_2:
        Sparkle_Grovetown_2_index =19
        
        update_place_to_xlmap(xl_map,Sparkle_Grovetown_2_index,Sparkle_Grovetown_2)
        
    Fuller_Calumet = data.get("Fuller-Calumet")
    
    if Fuller_Calumet:
        Fuller_Calumet_index= 5
        
        update_place_to_xlmap(xl_map,Fuller_Calumet_index,Fuller_Calumet)
        
        
    Sudz_Beverly = data.get("Sudz - Beverly")
    
    if Sudz_Beverly:
        Sudz_Beverly_index = 4
        
        update_place_to_xlmap(xl_map,Sudz_Beverly_index,Sudz_Beverly)
        
    ##Hamilton sites 
    
    Getaway_Macomb = data.get("Getaway-Macomb")
    
    if Getaway_Macomb:
        Getaway_Macomb_index = 10
        update_place_to_xlmap(xl_map,Getaway_Macomb_index,Getaway_Macomb)
        
    Getaway_Morton = data.get("Getaway-Morton")
    
    if Getaway_Morton:
        Getaway_Morton_index=11
        update_place_to_xlmap(xl_map,Getaway_Morton_index,Getaway_Morton)
        
    Getaway_Ottawa = data.get("Getaway-Ottawa")
    
    if Getaway_Ottawa:
        Getaway_Ottawa_index = 12
        update_place_to_xlmap(xl_map,Getaway_Ottawa_index,Getaway_Ottawa)
        
    Getaway_Peru = data.get("Getaway-Peru")    
    
    if Getaway_Peru:
        Getaway_Peru_index = 13
        update_place_to_xlmap(xl_map,Getaway_Peru_index,Getaway_Peru)
        
    
    #Hamilton 
    
    Splash_Peoria = data.get("Splash-Peoria")
    
    if Splash_Peoria:
        Splash_Peoria_index = 9
        update_place_to_xlmap(xl_map,Splash_Peoria_index,Splash_Peoria)

    #computation for first three columns 
    # sum([ xl_map[2][i+1] for i in range(3,13)])
    
    #first index
    car_count_monday_to_friday_ILL = do_sum(xl_map,2,range(3,13))
    xl_map[2][2] = car_count_monday_to_friday_ILL
    
    car_count_monday_to_friday_GA_SC = do_sum(xl_map,2,range(13,22))
    xl_map[2][3] = car_count_monday_to_friday_GA_SC
    
    car_count_monday_to_friday_Total = sum([car_count_monday_to_friday_ILL,car_count_monday_to_friday_GA_SC])
    
    xl_map[2][1]=car_count_monday_to_friday_Total
    
    
    
    #secound index 
    car_count_saturday_to_sunday_ILL = do_sum(xl_map,3,range(3,13))
    xl_map[3][2] = car_count_saturday_to_sunday_ILL
    
    car_count_saturday_to_sunday_GA_SC = do_sum(xl_map,3,range(13,22))
    xl_map[3][3] = car_count_saturday_to_sunday_GA_SC
    
    car_count_saturday_to_sunday_Total = sum([car_count_saturday_to_sunday_ILL, car_count_saturday_to_sunday_GA_SC])
    
    xl_map[3][1] = car_count_saturday_to_sunday_Total
    
    #third row 
    
    retail_car_count_monday_to_friday_ILL = do_sum(xl_map,4,range(3,13))
    xl_map[4][2] = retail_car_count_monday_to_friday_ILL
    
    retail_car_count_monday_to_friday_GA_SC = do_sum(xl_map,4,range(13,22))
    xl_map[4][3] = retail_car_count_monday_to_friday_GA_SC
    
    retail_car_count_monday_to_friday_Total = sum([retail_car_count_monday_to_friday_ILL,retail_car_count_monday_to_friday_GA_SC])
    
    xl_map[4][1]=retail_car_count_monday_to_friday_Total
    
    # Reatil car cound satruday to sunday
    
    retail_car_count_saturday_to_sunday_ILL = do_sum(xl_map,5,range(3,13))
    xl_map[5][2] = retail_car_count_saturday_to_sunday_ILL
    
    retail_car_count_saturday_to_sunday_GA_SC = do_sum(xl_map,5,range(13,22))
    xl_map[5][3] = retail_car_count_saturday_to_sunday_GA_SC
    
    retail_car_count_saturday_to_sunday_Total = sum([retail_car_count_saturday_to_sunday_ILL,retail_car_count_saturday_to_sunday_GA_SC])
    
    xl_map[5][1]=retail_car_count_saturday_to_sunday_Total
    
    loc_1 = [[2,3],[3,3]]
    total_cars_GA_SC=do_sum_location(xl_map,loc_1)
    
    xl_map[6][3]=total_cars_GA_SC
    
    loc_2=[[2,2],[3,2]]
    total_cars_in_ILL = do_sum_location(xl_map,loc_2)
    
    xl_map[6][2] = total_cars_in_ILL
    
    loc_3=[[6,3],[6,2]]
    
    Total_cars_Total = do_sum_location(xl_map,loc_3)
    xl_map[6][1]=Total_cars_Total
    
    #Retail Revenue Monday to Friday
    retail_revenue_monday_to_friday_ILL = do_sum(xl_map,7,range(3,13))
    xl_map[7][2] = retail_revenue_monday_to_friday_ILL
    
    retail_revenue_monday_to_friday_GA_SC = do_sum(xl_map,7,range(13,22))
    xl_map[7][3] = retail_revenue_monday_to_friday_GA_SC
    
    retail_revenue_monday_to_friday_Total = sum([retail_revenue_monday_to_friday_ILL,retail_revenue_monday_to_friday_GA_SC])
    
    xl_map[7][1]=retail_revenue_monday_to_friday_Total
    
    
    #Reatil Revenue Saturda to Sunday
    retail_revenue_saturday_to_sunday_ILL = do_sum(xl_map,8,range(3,13))
    xl_map[8][2] = retail_revenue_saturday_to_sunday_ILL
    
    retail_revenue_saturday_to_sunday_GA_SC = do_sum(xl_map,8,range(13,22))
    xl_map[8][3] = retail_revenue_saturday_to_sunday_GA_SC
    
    retail_revenue_saturday_to_sunday_Total = sum([retail_revenue_saturday_to_sunday_ILL,retail_revenue_saturday_to_sunday_GA_SC])
    
    xl_map[8][1]=retail_revenue_saturday_to_sunday_Total
    
    #Total Revnue Monday to Friday
    Total_revenue_monday_to_friday_ILL = do_sum(xl_map,9,range(3,13))
    xl_map[9][2] = Total_revenue_monday_to_friday_ILL
    
    Total_revenue_monday_to_friday_GA_SC = do_sum(xl_map,9,range(13,22))
    xl_map[9][3] = Total_revenue_monday_to_friday_GA_SC
    
    Total_revenue_monday_to_friday = sum([Total_revenue_monday_to_friday_ILL,Total_revenue_monday_to_friday_GA_SC])
    
    xl_map[9][1]=Total_revenue_monday_to_friday
    
    # Total Revenue Saturday to Sunday
    Total_revenue_saturday_to_sunday_ILL = do_sum(xl_map,10,range(3,13))
    xl_map[10][2] = Total_revenue_saturday_to_sunday_ILL
    
    Total_revenue_saturday_to_sunday_GA_SC = do_sum(xl_map,10,range(13,22))
    xl_map[10][3] = Total_revenue_saturday_to_sunday_GA_SC
    
    Total_revenue_saturday_sunday = sum([Total_revenue_saturday_to_sunday_ILL,Total_revenue_saturday_to_sunday_GA_SC])
    
    xl_map[10][1]=Total_revenue_saturday_sunday 
    
    # Total Revenue 
    loc_4=[[9,2],[10,2]]
    Total_revenue_ILL = do_sum_location(xl_map,loc_4)
    xl_map[11][2] =Total_revenue_ILL
    
    loc_5=[[9,3],[10,3]]
    Total_revenue_GA_SC = do_sum_location(xl_map,loc_5)
    xl_map[11][3]= Total_revenue_GA_SC
    
    Total_revenue_Total = sum([Total_revenue_ILL,Total_revenue_GA_SC])
    xl_map[11][1] =Total_revenue_Total
    
    #Average Retail Visit

    
    
    Average_retail_visit_IL_val =Average_retail_visit_IL_function(retail_revenue_monday_to_friday_ILL,
                                     retail_revenue_saturday_to_sunday_ILL,
                                     retail_car_count_monday_to_friday_ILL,
                                     retail_car_count_saturday_to_sunday_ILL)
    xl_map[12][2]=round(Average_retail_visit_IL_val,2)
    
    

    Average_retail_visit__GA_SC_val = Average_retail_visit__GA_SC_fucntion(retail_revenue_monday_to_friday_GA_SC
                                                                            ,retail_revenue_saturday_to_sunday_GA_SC,
                                                                            retail_car_count_monday_to_friday_GA_SC,
                                                                            retail_car_count_saturday_to_sunday_GA_SC)
    xl_map[12][3] = round(Average_retail_visit__GA_SC_val,2)
    
    Average_retail_visit_Total_val =Average_retail_visit_Total_function(retail_revenue_monday_to_friday_Total,
                                        retail_revenue_saturday_to_sunday_Total,
                                        retail_car_count_monday_to_friday_Total,
                                        retail_car_count_saturday_to_sunday_Total)
    xl_map[12][1] = round(Average_retail_visit_Total_val,2)
    
    #Average Member visit 
    
    Average_memeber_visit_ILL_val =Average_memeber_visit_ILL_function(Total_revenue_ILL,retail_revenue_monday_to_friday_ILL,
                                       retail_revenue_saturday_to_sunday_ILL,total_cars_in_ILL,
                                       retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL)
    xl_map[13][2] = round(Average_memeber_visit_ILL_val,2)
    
    Average_memeber_visit_GA_SC_val = Average_memeber_visit_GA_SC_function(Total_revenue_GA_SC,
        retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC,
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC,total_cars_GA_SC)
    
    xl_map[13][3] = round(Average_memeber_visit_GA_SC_val,2)
    
    Average_memeber_visit_Total_val =Average_memeber_visit_Total_function(Total_revenue_Total,retail_revenue_monday_to_friday_Total
                        ,retail_revenue_saturday_to_sunday_Total,Total_cars_Total,retail_car_count_monday_to_friday_Total,
                        retail_car_count_saturday_to_sunday_Total)
    
    xl_map[13][1] = round(Average_memeber_visit_Total_val,2)
    
    #Staff Hours Monday to Friday
    
    staff_hours_monday_to_friday_ILL = do_sum(xl_map,14,range(3,13))
    xl_map[14][2] = staff_hours_monday_to_friday_ILL
    
    staff_hours_monday_to_friday_GA_SC = do_sum(xl_map,14,range(13,22))
    
    xl_map[14][3] = staff_hours_monday_to_friday_GA_SC
    
    staff_hours_monday_to_friday_Total = sum([staff_hours_monday_to_friday_ILL,staff_hours_monday_to_friday_GA_SC])
    
    xl_map[14][1] = staff_hours_monday_to_friday_Total
    
    
    #Staff Hours Saturday to Sunday

    staff_hours_saturday_to_sunday_ILL = do_sum(xl_map,15,range(3,13))
    xl_map[15][2] = staff_hours_saturday_to_sunday_ILL
    
    staff_hours_saturday_to_sunday_GA_SC = do_sum(xl_map,15,range(13,22))
    
    xl_map[15][3] = staff_hours_saturday_to_sunday_GA_SC
    
    staff_hours_saturday_to_sunday_Total = sum([staff_hours_saturday_to_sunday_ILL,staff_hours_saturday_to_sunday_GA_SC])
    
    xl_map[15][1] = staff_hours_saturday_to_sunday_Total  
    
    # Cost per labour hour  Monday to friday
    cost_per_labour_hour_monday_to_friday_ILL_val =cost_per_labour_hour_monday_to_friday_ILL_function(car_count_monday_to_friday_ILL,staff_hours_monday_to_friday_ILL)
    xl_map[16][2] = round(cost_per_labour_hour_monday_to_friday_ILL_val,2)
    
    
    cost_per_labour_hour_monday_to_friday_GA_SC_val = cost_per_labour_hour_monday_to_friday_GA_SC_fucntion(car_count_monday_to_friday_GA_SC,staff_hours_monday_to_friday_GA_SC)
    xl_map[16][3] = round(cost_per_labour_hour_monday_to_friday_GA_SC_val,2)
    
    
    cost_per_labour_hour_monday_to_friday_Total_val =cost_per_labour_hour_monday_to_friday_Total_function(car_count_monday_to_friday_Total,staff_hours_monday_to_friday_Total)
    
    xl_map[16][1] = round(cost_per_labour_hour_monday_to_friday_Total_val,2)
    
    #Cost per laobour hour Saturday and Sunday
    cost_per_labour_hour_saturday_to_sunday_ILL_val = cost_per_labour_hour_saturday_to_sunday_ILL_function(car_count_saturday_to_sunday_ILL,staff_hours_saturday_to_sunday_ILL)
    xl_map[17][2] = round(cost_per_labour_hour_saturday_to_sunday_ILL_val,2)
    
    
    
    cost_per_labour_hour_saturday_to_sunday_GA_SC_val = cost_per_labour_hour_saturday_to_sunday_GA_SC_function(car_count_saturday_to_sunday_GA_SC,staff_hours_saturday_to_sunday_GA_SC)
    xl_map[17][3] = round(cost_per_labour_hour_saturday_to_sunday_GA_SC_val,2)
    
    cost_per_labour_hour_saturday_to_sunday_Total= car_count_saturday_to_sunday_Total/staff_hours_saturday_to_sunday_Total if staff_hours_saturday_to_sunday_Total!=0 else ""
    
    xl_map[17][1] = round(cost_per_labour_hour_saturday_to_sunday_Total,2) if cost_per_labour_hour_saturday_to_sunday_Total else ""
    
    # Total cars per man hour
    

    Total_cars_per_man_hour_ILL_val =Total_cars_per_man_hour_ILL_function(total_cars_in_ILL,
                                         staff_hours_monday_to_friday_ILL,
                                         staff_hours_saturday_to_sunday_ILL)
    xl_map[18][2] = round(Total_cars_per_man_hour_ILL_val,2)
    
    Total_cars_per_man_hour_GA_SC_val = Total_cars_per_man_hour_GA_SC_function(total_cars_GA_SC, staff_hours_monday_to_friday_GA_SC,
                                                                               staff_hours_saturday_to_sunday_GA_SC)
    
    xl_map[18][3] = round(Total_cars_per_man_hour_GA_SC_val,2)
    
    
    Total_cars_per_man_hour_total_val =Total_cars_per_man_hour_total_function(Total_cars_Total,
                                           staff_hours_monday_to_friday_Total,
                                           staff_hours_saturday_to_sunday_Total)
    xl_map[18][1] = round(Total_cars_per_man_hour_total_val,2)
    
    #Total club plans sold 
    Total_club_plans_sold_ILL = do_sum(xl_map,19,range(3,13))
    
    xl_map[19][2] = Total_club_plans_sold_ILL
    
    Total_club_plans_sold_GA_SC = do_sum(xl_map,19,range(13,22))
    
    xl_map[19][3] = Total_club_plans_sold_GA_SC
    
    Total_club_plans_sold_Total = sum([Total_club_plans_sold_ILL,Total_club_plans_sold_GA_SC])
    
    xl_map[19][1] = Total_club_plans_sold_Total
    
    
    #Conversion Rate 
    Conversion_rate_ILL_val =Conversion_rate_ILL_function(Total_club_plans_sold_ILL,
                                                          retail_car_count_monday_to_friday_ILL,
                                                          retail_car_count_saturday_to_sunday_ILL)
    xl_map[20][2] = round((Conversion_rate_ILL_val * 100),2)
    
    Conversion_rate_GA_SC_val =Conversion_rate_GA_SC_function(Total_club_plans_sold_GA_SC,retail_car_count_monday_to_friday_GA_SC,
                                   retail_car_count_saturday_to_sunday_GA_SC)
    
    xl_map[20][3]= round((Conversion_rate_GA_SC_val * 100),2)
    
    

    Conversion_rate_Total_val = Conversion_rate_Total_function(Total_club_plans_sold_Total,
                                                               retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total)
    xl_map[20][1] = round((Conversion_rate_Total_val * 100),2)
    
    
    #Total club plan members 
    Total_club_planmembers_ILL = do_sum(xl_map,21,range(3,13))
    
    xl_map[21][2] = Total_club_planmembers_ILL
    
    Total_club_planmembers_GA_SC = do_sum(xl_map,21,range(13,22))
    
    xl_map[21][3] = Total_club_planmembers_GA_SC 
    
    Total_club_planmembers_Total = sum([Total_club_planmembers_ILL,Total_club_planmembers_GA_SC])
    
    xl_map[21][1] = Total_club_planmembers_Total
    
    #Toatl Cars for all locations 
    total_cars_row = 6 
    for i in range(3,22):
        xl_map[total_cars_row][i+1]= do_sum_location(xl_map,[[2,i+1],[3,i+1]])
    
    
    revenue_row = 11
    
    #Total Revenue calculatio n for all locations 
    for i in range(3,22):
        xl_map[revenue_row][i+1]= do_sum_location(xl_map,[[9,i+1],[10,i+1]])
    
    average_retail_visit_row = 12
    for i in range(3,22):
        retail_revenue_mon_sun = do_sum_location(xl_map,location=[[7,i+1],[8,i+1]])
        retail_car_count_mon_sun = do_sum_location(xl_map,location=[[4,i+1],[5,i+1]])
        
        average_retail_visit_val = retail_revenue_mon_sun/retail_car_count_mon_sun if retail_car_count_mon_sun != 0 else ""
        xl_map[average_retail_visit_row][i+1]=   round(average_retail_visit_val,2) if average_retail_visit_val else ""
    
    
    average_member_visit_row = 13
    for i in range(3,22):
        total_revenue = xl_map[11][i+1]
        total_revenue = total_revenue if isinstance(total_revenue,int) or isinstance(total_revenue,float) else 0
        
        retail_revenue_mon_sun = do_sum_location(xl_map,location=[[7,i+1],[8,i+1]])
        
        total_cars = xl_map[6][i+1]
        total_cars = total_cars if isinstance(total_cars,int) or isinstance(total_cars,float) else 0
        
        retail_car_count_mon_sun  = do_sum_location(xl_map,location=[[4,i+1],[5,i+1]])
        
        average_member_visit_val = (total_revenue-retail_revenue_mon_sun)/(total_cars-retail_car_count_mon_sun) if total_cars-retail_car_count_mon_sun !=0 else ""
        
        xl_map[average_member_visit_row][i+1] = round(average_member_visit_val,2) if average_member_visit_val else ""
    
    #Total cars per man hour 
    total_cars_per_man_hour_row = 18 
    for i in range(3,22):
        total_cars = xl_map[6][i+1]
        total_cars = total_cars if isinstance(total_cars,int) or isinstance(total_cars,float) else 0
        staff_hours_monday_to_saturday = do_sum_location(xl_map,[[14,i+1],[15,i+1]])
        
        total_cars_per_man_hour_val = total_cars/staff_hours_monday_to_saturday if staff_hours_monday_to_saturday !=0 else ""
        
        xl_map[total_cars_per_man_hour_row][i+1] = round(total_cars_per_man_hour_val,2) if total_cars_per_man_hour_val else ""
 
        
      
    # Define cell styles (assuming they are the same as before)
    bg_color = PatternFill(start_color='0b3040', end_color='0b3040', fill_type='solid')
    font_color = Font(color='FFFFFF')

    bg_color_index = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    font_color_index = Font(color='000000')

    darkgreen_format = PatternFill(start_color='0ee85e', end_color='0ee85e', fill_type='solid')
    light_green_format = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    darkred_format = PatternFill(start_color='FC0303', end_color='FC0303', fill_type='solid')
    lightred_format = PatternFill(start_color='D98484', end_color='D98484', fill_type='solid')
    yellow=  PatternFill(start_color='D0D48A', end_color='D0D48A', fill_type='solid')
    
    #writing to  actual sheet
    #first row comment section
    # Write comments in the first row
    # worksheet.append([comment])
    
    for row in range(len(xl_map)):
        for col in range(len(xl_map[row])):
            val = xl_map[row][col]
            cell = worksheet.cell(row=row+1, column=col+1, value=val)  # offset by 0 rows for header and comment
            
            val = add_commas(val)
            
            if row == 1 and col != 0:
                cell.fill = bg_color
                cell.font = font_color

            elif col == 0 and 1 < row < 22:
                cell.fill = bg_color_index
                cell.font = font_color_index

            elif val and row == 12 and col > 0:
                if val >= 10:
                    cell.fill = darkgreen_format
                    
                elif val>=5 and val <10: # [5,9] (inclusive intervals)
                    cell.fill = light_green_format
                    
                elif val>=-5 and val <5:  # [-5,4]
                    cell.fill = lightred_format
        
                elif  val <=-10 and val<-5:
                    cell.fill = darkred_format

            elif val and row in [16, 17, 18, 20] and col > 0:
                if val >= 20:
                    cell.fill = darkgreen_format
                elif val>=10 and val <20:
                    cell.fill = light_green_format
                elif val>=-10 and val <10:
                    cell.fill = lightred_format
                elif val <=-20 or val<-10:
                    cell.fill = darkred_format

            # elif row == 5 and col > 2:
            #     cell.value = ""

            # elif row in [17] and col > 2:
            #     cell.value = ""

            # elif row == 21:
            #     cell.value = ""

            # elif row > 21 and col > 0:
            #     cell.value = ""

    # Add legend or additional information below the table
    legend_start_row = 24

    legend_styles = {
        "Very Concerning": PatternFill(start_color="fc0303", end_color="fc0303", fill_type="solid"),
        "Concerning": PatternFill(start_color="d98484", end_color="d98484", fill_type="solid"),
        "Neutral": PatternFill(start_color="d0d48a", end_color="d0d48a", fill_type="solid"),
        "Positive": PatternFill(start_color="8ad493", end_color="8ad493", fill_type="solid"),
        "Very Positive": PatternFill(start_color="0ee85e", end_color="0ee85e", fill_type="solid"),
    }

    worksheet.cell(row=legend_start_row, column=1, value='Legend')
    worksheet.cell(row=legend_start_row + 1, column=1, value='Very Concerning').fill = legend_styles["Very Concerning"]
    worksheet.cell(row=legend_start_row + 2, column=1, value='Concerning').fill = legend_styles["Concerning"]
    worksheet.cell(row=legend_start_row + 3, column=1, value='Neutral').fill = legend_styles["Neutral"]
    worksheet.cell(row=legend_start_row + 4, column=1, value='Positive').fill = legend_styles["Positive"]
    worksheet.cell(row=legend_start_row + 5, column=1, value='Very Positive').fill = legend_styles["Very Positive"]
    
    #setting cloumn width as deafult 
    column_width = 20  # You can change this value to whatever width you need
    first_col_width =25
    for col in range(1, 24):  # Columns A to W are 1 to 23
        column_letter = get_column_letter(col)
        if col==1:
            worksheet.column_dimensions[column_letter].width = first_col_width
        else:
            worksheet.column_dimensions[column_letter].width = column_width
       
    
    # Define the border style


    thick_border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'))
    
    thick_border_bottom = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    bottom=Side(style='thick'))




    # Apply the border to a range of cells (e.g., A1:C3)
    # for row in worksheet.iter_rows(min_row=3, max_row=21, min_col=2, max_col=4):
    #     for cell in row:
    #         cell.border = thin_border

    for row in range(3,23):
        cell1 = worksheet.cell(row=row,column=2)
        cell2 = worksheet.cell(row=row,column=3)
        cell3 = worksheet.cell(row=row,column=4)
        
        if row==22:
            cell1.border=thick_border_bottom
            cell2.border=thick_border_bottom
            cell3.border=thick_border_bottom
        else:
            cell1.border=thick_border
            cell2.border=thick_border
            cell3.border=thick_border
            
    for row in worksheet.iter_rows():
        for cell in row:
            row_index = cell.row
            if isinstance(cell.value, (int, float)) and row_index in [17,18,19]:
                cell.number_format = '#,##0.0'
            elif isinstance(cell.value, (int, float)): #:
                cell.number_format = '#,##0'
    #Doller sysmbol     
    for row in range(8,13):
        cell1 = worksheet.cell(row=row,column=2)
        cell2 = worksheet.cell(row=row,column=3)
        cell3 = worksheet.cell(row=row,column=4)
        
        cell1.border=thick_border
        cell2.border=thick_border
        cell3.border=thick_border
        
        cells=[cell1,cell2,cell3]
        for cell in cells:
            if isinstance(cell.value, (int, float)) and cell.value >= 1000:
                cell.number_format = '"$"#,##0'     
    


    #applying bold font
    # Define a bold font style
    bold_font = Font(bold=True)

    for row in worksheet.iter_rows(min_row=12, max_row=13, min_col=2, max_col=4):
        for cell in row:
            cell.font = bold_font
            
    for row in worksheet.iter_rows(min_row=19, max_row=19, min_col=2, max_col=4):
        for cell in row:
            cell.font = bold_font
            
    for row in worksheet.iter_rows(min_row=21, max_row=21, min_col=2, max_col=4):
        for cell in row:
            cell.font = bold_font
    # Save the modified workbook
    workbook.save(filename)





# print(data_path)

# Your existing code

def get_week_dates_for_storage():
    "will retun in '%m_%Y' ==> 07-2024"
    # Get the current date
    today = datetime.today()
    
    # Find the current week's Monday date
    current_week_monday = today - timedelta(days=today.weekday())
    
    # Find the current week's Sunday date
    current_week_sunday = current_week_monday + timedelta(days=6)
    
    # Format the dates in dd/mm/yyyy format
    #monday_date_str = current_week_monday.strftime("%m/%d/%Y")
    sunday_date_str = current_week_sunday.strftime("%m_%Y")
    
    #old ret f"{monday_date_str}-{sunday_date_str}".replace('/','_')
    return sunday_date_str



def create_storage_directory(path):
    created_path=None
    try:
        path = os.path.join(data_path,path)
        os.makedirs(path, exist_ok=True)
        print(f"Directory '{path}' created successfully")
        created_path =path
    except OSError as error:
        print(f"Directory '{path}' cannot be created: {error}")
    
    return created_path



if __name__=="__main__":

    load_dotenv()

    class emailConfig:
        env_vars    = os.environ
        FROM_EMAIL   = env_vars.get("FROM_EMAIL")
        FROM_NAME      = env_vars.get("FROM_NAME")
        SMTP_SERVER   = env_vars.get("SMTP_SERVER")
        SMTP_PORT = env_vars.get("SMTP_PORT")
        SMTP_USER=env_vars.get("SMTP_USER")
        SMTP_PASSWORD=env_vars.get("SMTP_PASSWORD")
        TO_EMAIL=env_vars.get("TO_EMAIL")
    
    # Configuration
    subject = 'Weekly reports'
    body = 'This is the body of the email.'
    to_email = emailConfig.TO_EMAIL
    from_email = emailConfig.FROM_EMAIL
    from_name = emailConfig.FROM_NAME
    smtp_server = emailConfig.SMTP_SERVER
    smtp_port = emailConfig.SMTP_PORT
    smtp_user = emailConfig.SMTP_USER
    smtp_password = emailConfig.SMTP_PASSWORD
    
    
    #cc_emails=["CR@SparkleCW.com","FZ@SparkleCW.com","Rick@SparkleStatus.com","Shane@SparkleStatus.com"]
    
    # path = get_week_dates_for_storage()
    # storage_path = create_storage_directory(path)
    
    # wahsify_week_days = washify_week_dates()
    # washify_file_name = 
    # washify_file_path_full = os.path.join(storage_path,washify_file_name)
    # washify_week_report(storage_path,wahsify_week_days[0],wahsify_week_days[1])
    
    # hamilton_week_days = hamilton_week_dates()
    # hamilton_file_name = f"hamilton_{hamilton_week_days[0]}-{hamilton_week_days[-1]}.csv".replace('/','_')
    # hamilton_full_path= os.path.join(storage_path,hamilton_file_name)
    # hamilton_week_report(hamilton_full_path,hamilton_week_days[0],hamilton_week_days[-1])
    
    # sitewatch_week_days = sitewatch_week_dates()
    # site_watch_file_path = storage_path
    # sitewatch_week_report(site_watch_file_path,sitewatch_week_days[0],sitewatch_week_days[-1])
    
    # Directory containing Excel files
    # directory_path = storage_path
    # attachments = get_excel_files(directory_path)
    
    #Sending email to email address
    # send_email(subject, body, to_email, from_email, from_name, smtp_server, smtp_port, smtp_user, smtp_password, attachments)
    
    # ------------------- Test Script start ---------------------#
    
    # # monday_date_str, sunday_date_str = sitewatch_week_dates()
    # # print(monday_date_str,sunday_date_str)
    # monday_date_str="2024-06-24"
    # friday_date_str = "2024-06-28"
    # saturday_date_str = "2024-06-29"
    # sunday_date_str="2024-06-30"  #Y-M-D
    
    # sitewatch_report = sitewatch_week_report("",monday_date_str,friday_date_str,saturday_date_str, sunday_date_str)
    
    # monday_date_str, friday_date_str, saturday_date_str, sunday_date_str =  washify_week_dates()
    
    # #testing dates 
    # monday_date_str =  "06/24/2024"
    # friday_date_str =  "06/28/2024"
    # saturday_date_str = "06/29/2024"
    # sunday_date_str  =  "06/30/2024"  #M/D/Y
    
    # print(monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    # washify_report = washify_week_report("", monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    
    # #for hamilton dates
    # monday_date_str = "2024-06-24"
    # friday_date_str = "2024-06-28"
    # saturday_date_str = "2024-06-29"
    # sunday_date_str  = "2024-06-30"
    
    # hamilton_report = hamilton_week_report(monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    
    # data = sitewatch_report
    
    # data.update(washify_report)
    
    # data.update(hamilton_report)
    # comment =f"Ending {sunday_date_str}"
    # sheet_name= sunday_date_str.replace("/","-")
    
    # with open("all_data_curent.json","w") as f:
    #     json.dump(data,f,indent=4)
    
    # filename="test1.xlsx"
    # file_name_with_fullpath = os.path.join(data_path,filename)
    # prepare_xlmap(data,comment,sheet_name=sheet_name,filename=file_name_with_fullpath)
    
    # # Directory containing Excel files
    # directory_path = data_path
    # attachments = get_excel_files(directory_path)
    
    #Sending email to email address
    #send_email(subject, body, to_email, from_email, from_name, smtp_server, smtp_port, smtp_user, smtp_password, attachments,cc_emails)
    
    
    
    # ------------------- Test Script ends ---------------------#
    
    # # -----------------Actual script  ----------------------------#
    
    path = get_week_dates_for_storage()
    path="test"#"06-2024"
    storage_path = create_storage_directory(path)
    # monday_date_str, friday_date_str, saturday_date_str, sunday_date_str = sitewatch_week_dates()
    # print(monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    monday_date_str="2024-07-01"
    friday_date_str = "2024-07-05"
    saturday_date_str = "2024-07-06"
    sunday_date_str="2024-07-07"  #Y-M-D
    
    sitewatch_report = sitewatch_week_report("",monday_date_str,friday_date_str,saturday_date_str, sunday_date_str)
    
    # monday_date_str, friday_date_str, saturday_date_str, sunday_date_str =  washify_week_dates()
    
    ##testing dates 
    monday_date_str =  "07/01/2024"
    friday_date_str =  "07/05/2024"
    saturday_date_str = "07/06/2024"
    sunday_date_str  =  "07/07/2024"  #M/D/Y
    
    print(monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    washify_report = washify_week_report("", monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    
    ##for hamilton dates
    monday_date_str = "2024-07-01"
    friday_date_str = "2024-07-05"
    saturday_date_str = "2024-07-06"
    sunday_date_str  = "2024-07-07"
    
    # monday_date_str, friday_date_str, saturday_date_str, sunday_date_str = hamilton_week_dates()
    hamilton_report = hamilton_week_report(monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    
    data = sitewatch_report
    
    data.update(washify_report)
    
    data.update(hamilton_report)
    
    comment =f"Ending {sunday_date_str}"
    sheet_name= sunday_date_str.replace("/","-")
    filename=f"{path}.xlsx"
    file_name_with_fullpath = os.path.join(storage_path,filename)
    prepare_xlmap(data,comment,sheet_name=sheet_name,filename=file_name_with_fullpath)
    
    # Directory containing Excel files
    directory_path = storage_path
    attachments = get_excel_files(directory_path)
    
    #Sending email to email address
    body = f'weekly report Ending {sunday_date_str}'
    #send_email(subject, body, to_email, from_email, from_name, smtp_server, smtp_port, smtp_user, smtp_password, attachments)#
    
    # prepare_xlmap(data,comment,sheet_name=sheet_name)
    # # -----------------Actual script  ----------------------------#
    



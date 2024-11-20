
import sys
import os
from dotenv import load_dotenv
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
import json

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import logging
from custom_mailer import send_email,get_excel_files,send_email_on_error


# Add the path to the parent directory of "washify" to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'washify')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'sitewash')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'hamilton')))
# print(sys.path)

from washify_report import generate_report as washify_report
from sitewatch_report import generate_report as sitewatch_report
from hamilton_report import generate_report as hamilton_report


current_file_path = os.path.dirname(os.path.abspath(__file__))
# print(current_file_path)
data_path = os.path.join(current_file_path,"data")


def do_sum(xl_map,start_index,range):
    "This will do row based some "
    total=0
    for i in range:
        val = xl_map[start_index][i+1]
        if isinstance(val,float) or isinstance(val,int):
            total+=val

    return total


def Average_retail_visit_GA_SC_fucntion(retail_revenue_GA_SC,
                                        retail_car_count_GA_SC):

    result = 0
    try:
        Average_retail_visit__GA_SC = retail_revenue_GA_SC /retail_car_count_GA_SC
        result = Average_retail_visit__GA_SC
    except Exception as e:
        print(f"Exception in Average_retail_visit__GA_SC_fucntion() {e}")
        logging.info(f"Exception in Average_retail_visit__GA_SC_fucntion() {e}")

    return result


def Average_memeber_visit_GA_SC_function(Total_revenue_GA_SC,retail_revenue_GA_SC,
                                        retail_car_count_GA_SC,car_count_GA_SC):
    result = 0

    try:
        Average_memeber_visit_GA_SC = (Total_revenue_GA_SC - retail_revenue_GA_SC)/(car_count_GA_SC - retail_car_count_GA_SC)
        result = Average_memeber_visit_GA_SC
    except Exception as e:
        print(f"Exception in Average_memeber_visit_GA_SC_function() {e}")
        logger.info(f"Exception in Average_memeber_visit_GA_SC_function() {e}")

    return result

def cost_per_labour_hour_GA_SC_fucntion(car_count_GA_SC,staff_hours_GA_SC):
    result = 0
    try:
        cost_per_labour_hour_GA_SC = car_count_GA_SC/staff_hours_GA_SC
        result =cost_per_labour_hour_GA_SC
    except Exception as e:
        print(f"Exception in cost_per_labour_hour_GA_SC() {e}")
        logger.info(f"Exception in cost_per_labour_hour_GA_SC() {e}")

    return result


def Total_cars_per_man_hour_GA_SC_function(car_count_GA_SC, staff_hours_GA_SC):
    result =0
    try:
            Total_cars_per_man_hour_GA_SC = car_count_GA_SC/staff_hours_GA_SC
            result = Total_cars_per_man_hour_GA_SC

    except Exception as e:
        print(f"Exception in  Total_cars_per_man_hour_GA_SC_function() {e}")
        logger.info(f"Exception in  Total_cars_per_man_hour_GA_SC_function() {e}")

    return result

def Conversion_rate_GA_SC_function(Total_club_plans_sold_GA_SC, retail_car_count_GA_SC):
    result = 0
    try:
        Conversion_rate_GA_SC = Total_club_plans_sold_GA_SC/retail_car_count_GA_SC
        result =  Conversion_rate_GA_SC
    except Exception as e:
        print(f"Exceptionin Conversion_rate_GA_SC_function() {e}")
        logger.info(f"Exceptionin Conversion_rate_GA_SC_function() {e}")

    return result

def Conversion_rate_Total_function(Total_club_plans_sold_Total, retail_car_count_Total):
    result = 0

    try:
        Conversion_rate_Total = Total_club_plans_sold_Total/retail_car_count_Total
        result = Conversion_rate_Total
    except Exception as e:
        print(f"Exception in Conversion_rate_Total_function()  {e}")
        logger.info(f"Exception in Conversion_rate_Total_function()  {e}")

    return result

def Conversion_rate_ILL_function(Total_club_plans_sold_ILL, retail_car_count_ILL):
    result = 0

    try:
        Conversion_rate_ILL = Total_club_plans_sold_ILL/retail_car_count_ILL
        result = Conversion_rate_ILL
    except Exception as e:
        print(f"Exception in Conversion_rate_ILL_function() {e}")
        logger.info(f"Exception in Conversion_rate_ILL_function() {e}")

    return result

def Total_cars_per_man_hour_total_function(Total_cars_Total,
                                           staff_hours_Total):
    result = 0

    try:
        Total_cars_per_man_hour_total = Total_cars_Total/staff_hours_Total
        result = Total_cars_per_man_hour_total

    except Exception as e:
        print(f"Exception in Total_cars_per_man_hour_total_function() {e}")
        logger.info(f"Exception in Total_cars_per_man_hour_total_function() {e}")

    return result


def Total_cars_per_man_hour_ILL_function(car_count_ILL,staff_hours_ILL):

    result = 0
    try:
        Total_cars_per_man_hour_ILL = car_count_ILL/staff_hours_ILL
        result = Total_cars_per_man_hour_ILL

    except Exception as e:
        print(f"Exception Total_cars_per_man_hour_ILL_function() {e}")
        logger.info(f"Exception Total_cars_per_man_hour_ILL_function() {e}")

    return result


def Average_memeber_visit_Total_function(Total_revenue_Total,retail_revenue_Total,
                                        Total_cars_Total,retail_car_count_Total):
    result = 0

    try:
        Average_memeber_visit_Total = (Total_revenue_Total - retail_revenue_Total)/(Total_cars_Total - retail_car_count_Total)
        result =  Average_memeber_visit_Total

    except Exception as e:
        print(f"Exception Average_memeber_visit_Total_function() {e}")
        logger.info(f"Exception Average_memeber_visit_Total_function() {e}")

    return result

def Average_memeber_visit_ILL_function(Total_revenue_ILL,retail_revenue_ILL,
                                    car_count_ILL, retail_car_count_ILL):

    result = 0
    try:
        Average_memeber_visit_ILL = (Total_revenue_ILL - retail_revenue_ILL)/(car_count_ILL - retail_car_count_ILL)
        result = Average_memeber_visit_ILL

    except Exception as e:
        print(f"Exception Average_memeber_visit_ILL_function() {e}")
        logger.info(f"Exception Average_memeber_visit_ILL_function() {e}")

    return result

def Average_retail_visit_Total_function(retail_revenue_Total,retail_car_count_Total):

    result = 0
    try:
        Average_retail_visit_Total = retail_revenue_Total / retail_car_count_Total
        result = Average_retail_visit_Total

    except Exception as e:
        print(f"Exception Average_retail_visit_Total_function() {e}")
        logger.info(f"Exception Average_retail_visit_Total_function() {e}")

    return result


def Average_retail_visit_IL_function(retail_revenue_ILL,retail_car_count_ILL):

    result = 0
    try:
        Average_retail_visit_IL = retail_revenue_ILL / retail_car_count_ILL
        result = Average_retail_visit_IL

    except Exception as e:
        print(f"Exception Average_retail_visit_IL_function() {e}")
        logger.info(f"Exception Average_retail_visit_IL_function() {e}")

    return result

def do_percentage(current_year_data, last_year_data):
    '''
    Calculate percentage based on current year and last year data
    '''
    try:
        if isinstance(current_year_data,(int, float)) and isinstance(last_year_data, (int,float)):
            percentage =((current_year_data - last_year_data) / last_year_data)*100 if last_year_data!=0 else 0
            # print("percentage: ", percentage)
            return percentage
        else:
            return ""
    except Exception as e:
        print(f"Exception do_percentage() {e}")
        logger.info(f"Exception do_percentage {e}")


#new xl maps functions
def update_place_to_xlmap(xl_map,place_index,place_dictionary)->list:
    "Will return updates place dictionary"
    xl_map[2][place_index] = place_dictionary.get("car_count_current_year")
    xl_map[3][place_index] = place_dictionary.get("car_count_last_year")

    xl_map[6][place_index]=place_dictionary.get("retail_car_count_current_year")
    xl_map[7][place_index]=place_dictionary.get("retail_car_count_last_year")

    xl_map[14][place_index]=place_dictionary.get("retail_revenue_current_year")
    xl_map[15][place_index]=place_dictionary.get("retail_revenue_last_year")

    xl_map[18][place_index]=place_dictionary.get("total_revenue_current_year")
    xl_map[19][place_index]=place_dictionary.get("total_revenue_last_year")

    xl_map[33][place_index]=place_dictionary.get("labour_hours_current_year")

    xl_map[35][place_index]=place_dictionary.get("cars_per_labour_hour_current_year")
    xl_map[36][place_index]=place_dictionary.get("cars_per_labour_hour_last_year")

    xl_map[39][place_index] = place_dictionary.get("arm_plans_sold_cnt_current_year")

    xl_map[40][place_index]= place_dictionary.get("conversion_rate_current_year")
    xl_map[41][place_index]= place_dictionary.get("conversion_rate_last_year")

    xl_map[44][place_index] = place_dictionary.get("total_arm_planmembers_cnt_current_year")

    return xl_map

def set_colour(val,row,col,worksheet,colours):
    darkgreen_format,light_green_format,darkred_format,lightred_format =colours
    cell=worksheet.cell(row,col)
    if val>=20:
       cell.fill = darkgreen_format

    elif val >=10:
        cell.fill = light_green_format


    elif val>=-10:
        cell.fill = darkred_format
    else : #val>=-20
       cell.fill = lightred_format

def set_colour_for_avg_retail(current_year,last_year, row, col, worksheet, colours):
    """This will do colur coding of for the xl sheet 10 5 0 -5 - 10

    Args:
        current_year (_type_): _description_
        last_year (_type_): _description_
    """

    if not all([current_year, last_year]):
        return

    darkgreen_format,light_green_format,darkred_format,lightred_format =colours
    cell=worksheet.cell(row,col)

    percentage = ((current_year - last_year) / last_year)*100 if last_year!=0 else 0
    logger.info(f" curent index : {row},{col}")
    logger.info(f"current year : {current_year} last year: {last_year}")
    logger.info(f"percentage : {percentage}")
    print("percentage :",percentage)
    if percentage >=10:
        #Dark green
        print("Dark green ")
        cell.fill = darkgreen_format
        logger.info("Dark green ")
    elif percentage >= 5:
        print("light green")
        cell.fill = light_green_format
        logger.info("light green")

    elif percentage > -5:
        print("neutral")
        logger.info("neutral")

    elif percentage > -10:
        print("light red ")
        cell.fill = lightred_format
        logger.info("light red ")

    elif percentage <=-10:
        print("Dark red")
        cell.fill = darkred_format
        logger.info("Dark red")

def set_colour_new(current_year_data,last_year_data,row,col,worksheet,colours):
    """This will do colur coding of for the xl sheet 10 5 0 -5 - 10

    Args:
        current_year (_type_): _description_
        last_year (_type_): _description_
    """

    if not all([current_year_data, last_year_data]):
        return
    print(f"location on xl : {row},{col}")
    logger.info(f"location on xl : {row},{col}")
    print(f"current year: {current_year_data} , last year: { last_year_data}")
    logger.info(f"current year: {current_year_data} , last year: { last_year_data}")
    darkgreen_format,light_green_format,darkred_format,lightred_format =colours
    cell=worksheet.cell(row,col)

    percentage = ((current_year_data - last_year_data) / last_year_data)*100 if last_year_data!=0 else 0

    print("percentage :",percentage)
    logger.info(f"percentage : {percentage}")
    if percentage >=10:
        #Dark green
        print("Dark green ")
        cell.fill = darkgreen_format
        logger.info("Dark green ")
    elif percentage >= 5:
        print("light green")
        cell.fill = light_green_format
        logger.info("light green")

    elif percentage > -5:
        print("neutral")
        logger.info("neutral")

    elif percentage > -10:
        print("light red ")
        cell.fill = lightred_format
        logger.info("light red ")

    elif percentage <=-10:
        print("Dark red")
        cell.fill = darkred_format
        logger.info("Dark red")


def prepare_xlmap(data,comment="The comment section",filename="test.xlsx",sheet_name="sheet1"):
    # Load the existing workbook using openpyxl
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.create_sheet(sheet_name)
    except Exception as _:
        print(f"creating new xl file !! {filename}")
        logger.info(f"creating new xl file !! {filename}")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name


    xl_map = [
    [""],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ["","","","","","","","","","","","","","","","","","","","","","","",],
    ]

    cols = ["    ","Totals","ILL",
                "GA / SC","Sudz - Beverly",'Fuller-Calumet',
                "Fuller-Cicero","Fuller-Matteson","Fuller-Elgin",
                "Splash-Peoria","Getaway-Macomb","Getaway-Morton",
                "Getaway-Ottawa","Getaway-Peru","Sparkle-Belair",
                "Sparkle-Evans","Sparkle-Furrys Ferry","Sparkle-Greenwood",
                "Sparkle-Grovetown 1","Sparkle-Grovetown 2","Sparkle-North Augusta",
                "Sparkle-Peach Orchard","Sparkle-Windsor Spring"]

    index_names =["Current Year Car Count","Last Year Car Count", "Percentage","",
                  "Current Year Retail Car Count","Last Year Retail Car Count","Percentage","",
                  "Current Year Total Cars", "Last Year Total Cars", "Percentage","",
                  "Current Year Retail Revenue","Last Year Retail Revenue", "Percentage","",
                  "Current Year Revenue","Last Year Revenue", "Percentage","",
                  "Current Year Total Revenue", "Last Year Total Revenue", "Percentage","",
                  "Current Year Avg. Retail Visit","Last Year Avg. Retail Visit","Percentage","",
                  "Current Year Avg. Member Visit","Last Year Avg. Member Visit","",
                  "Staff Hours","",
                  "Current Year Total Cars Per Man Hour","Last Year Total Cars Per Man Hour","Percentage","",
                  "Total Club Plans Sold",
                  "Current Year Conversion Rate","Last Year Conversion Rate","Percentage","",
                  "Total Club Plan Members"
                ]

    xl_map[0][0] = comment #to maps



    for col in range(len(cols)): #column names to maps
        xl_map[1][col]=cols[col]

    for index  in range(len(index_names)):  #index names to map
        xl_map[index+2][0]=index_names[index]

    Fuller_Cicero = data.get("Fuller-Cicero")

    if Fuller_Cicero:
        Fuller_Cicero_index=6
        update_place_to_xlmap(xl_map,Fuller_Cicero_index,Fuller_Cicero)
    Fuller_Matteson = data.get("Fuller-Matteson")

    if Fuller_Matteson:
        Fuller_Matteson_index=7
        update_place_to_xlmap(xl_map,Fuller_Matteson_index,Fuller_Matteson)

    Fuller_Elgin = data.get("Fuller-Elgin")

    if Fuller_Elgin:
        Fuller_Elgin_index = 8
        update_place_to_xlmap(xl_map,Fuller_Elgin_index,Fuller_Elgin)


    Sparkle_Belair = data.get("Sparkle-Belair")

    if Sparkle_Belair:
        Sparkle_Belair_index=14

        update_place_to_xlmap(xl_map,Sparkle_Belair_index,Sparkle_Belair)

    Sparkle_Evans = data.get("Sparkle-Evans")

    if Sparkle_Evans:
        Sparkle_Evans_index=15

        update_place_to_xlmap(xl_map,Sparkle_Evans_index,Sparkle_Evans)

    Sparkle_North_Augusta = data.get("Sparkle-North Augusta")

    if Sparkle_North_Augusta:
        Sparkle_North_Augusta_index=20

        update_place_to_xlmap(xl_map,Sparkle_North_Augusta_index,Sparkle_North_Augusta)

    Sparkle_Greenwood = data.get("Sparkle-Greenwood")

    if Sparkle_Greenwood:
        Sparkle_Greenwood_index=17

        update_place_to_xlmap(xl_map,Sparkle_Greenwood_index,Sparkle_Greenwood)


    Sparkle_Grovetown_1 = data.get("Sparkle-Grovetown 1")

    if Sparkle_Grovetown_1:
        Sparkle_Grovetown_1_index=18

        update_place_to_xlmap(xl_map,Sparkle_Grovetown_1_index,Sparkle_Grovetown_1)

    Sparkle_Windsor_Spring = data.get("Sparkle-Windsor Spring")

    if Sparkle_Windsor_Spring:
        Sparkle_Windsor_Spring_index=22

        update_place_to_xlmap(xl_map,Sparkle_Windsor_Spring_index,Sparkle_Windsor_Spring)

    Sparkle_Furrys_Ferry = data.get("Sparkle-Furrys Ferry")

    if Sparkle_Furrys_Ferry:
        Sparkle_Furrys_Ferry_index= 16

        update_place_to_xlmap(xl_map,Sparkle_Furrys_Ferry_index,Sparkle_Furrys_Ferry)

    Sparkle_Peach_Orchard = data.get("Sparkle-Peach Orchard")

    if Sparkle_Peach_Orchard:
        Sparkle_Peach_Orchard_index=21

        update_place_to_xlmap(xl_map,Sparkle_Peach_Orchard_index,Sparkle_Peach_Orchard)

    Sparkle_Grovetown_2 =  data.get("Sparkle-Grovetown 2")

    if Sparkle_Grovetown_2:
        Sparkle_Grovetown_2_index =19

        update_place_to_xlmap(xl_map,Sparkle_Grovetown_2_index,Sparkle_Grovetown_2)

    Fuller_Calumet = data.get("Fuller-Calumet")

    if Fuller_Calumet:
        Fuller_Calumet_index= 5

        update_place_to_xlmap(xl_map,Fuller_Calumet_index,Fuller_Calumet)


    Sudz_Beverly = data.get("Sudz - Beverly")

    if Sudz_Beverly:
        Sudz_Beverly_index = 4

        update_place_to_xlmap(xl_map,Sudz_Beverly_index,Sudz_Beverly)

    ##Hamilton sites
    Getaway_Macomb = data.get("Getaway-Macomb")

    if Getaway_Macomb:
        Getaway_Macomb_index = 10
        update_place_to_xlmap(xl_map,Getaway_Macomb_index,Getaway_Macomb)

    Getaway_Morton = data.get("Getaway-Morton")

    if Getaway_Morton:
        Getaway_Morton_index=11
        update_place_to_xlmap(xl_map,Getaway_Morton_index,Getaway_Morton)

    Getaway_Ottawa = data.get("Getaway-Ottawa")

    if Getaway_Ottawa:
        Getaway_Ottawa_index = 12
        update_place_to_xlmap(xl_map,Getaway_Ottawa_index,Getaway_Ottawa)

    Getaway_Peru = data.get("Getaway-Peru")

    if Getaway_Peru:
        Getaway_Peru_index = 13
        update_place_to_xlmap(xl_map,Getaway_Peru_index,Getaway_Peru)


    #Hamilton
    Splash_Peoria = data.get("Splash-Peoria")

    if Splash_Peoria:
        Splash_Peoria_index = 9
        update_place_to_xlmap(xl_map,Splash_Peoria_index,Splash_Peoria)


    all_locations = [Sudz_Beverly,Fuller_Calumet,
                    Fuller_Cicero,Fuller_Matteson,
                    Fuller_Elgin,Splash_Peoria,Getaway_Macomb,Getaway_Morton,
                    Getaway_Ottawa,Getaway_Peru,Sparkle_Belair,
                    Sparkle_Evans,Sparkle_Furrys_Ferry,Sparkle_Greenwood,
                    Sparkle_Grovetown_1,Sparkle_Grovetown_2,Sparkle_North_Augusta,
                    Sparkle_Peach_Orchard,Sparkle_Windsor_Spring]

    ga_sc = all_locations[10:]
    ill = all_locations[0:10]

    #current year car count - ILL
    car_count_current_year_ILL = do_sum(xl_map,2,range(3,13))
    xl_map[2][2] = car_count_current_year_ILL

    #last year car count - ILL
    car_count_last_year_ILL = do_sum(xl_map,3,range(3,13))
    xl_map[3][2] = car_count_last_year_ILL

    car_count_percentage_ILL = do_percentage(car_count_current_year_ILL, car_count_last_year_ILL)
    xl_map[4][2] = car_count_percentage_ILL


    #current year car count - GA_SC
    car_count_current_year_GA_SC = do_sum(xl_map,2,range(13,22))
    xl_map[2][3] = car_count_current_year_GA_SC

    #last year car count - GA_SC
    car_count_last_year_GA_SC = do_sum(xl_map,3,range(13,22))
    xl_map[3][3] = car_count_last_year_GA_SC

    car_count_percentage_GA_SC = do_percentage(car_count_current_year_GA_SC, car_count_last_year_GA_SC)
    xl_map[4][3] = car_count_percentage_GA_SC


    #current year car count total
    car_count_current_year_Total = sum([car_count_current_year_ILL,car_count_current_year_GA_SC])
    xl_map[2][1]=car_count_current_year_Total

    #last year car count total
    car_count_last_year_Total = sum([car_count_last_year_ILL,car_count_last_year_GA_SC])
    xl_map[3][1]=car_count_last_year_Total

    car_count_total_percentage = do_percentage(car_count_current_year_Total, car_count_last_year_Total)
    xl_map[4][1] = car_count_total_percentage


    # Retail car count current year -IIL
    retail_car_count_current_year_ILL = do_sum(xl_map,6,range(3,13))
    xl_map[6][2] = retail_car_count_current_year_ILL

    # Retail car count last year -IIL
    retail_car_count_last_year_ILL = do_sum(xl_map,7,range(3,13))
    xl_map[7][2] = retail_car_count_last_year_ILL

    retail_car_percentage_ILL = do_percentage(retail_car_count_current_year_ILL, retail_car_count_last_year_ILL)
    xl_map[8][2] = retail_car_percentage_ILL


    # Retail car count current year - GA_SC
    retail_car_count_current_year_GA_SC = do_sum(xl_map,6,range(13,22))
    xl_map[6][3] = retail_car_count_current_year_GA_SC

    # Retail car count last year - GA_SC
    retail_car_count_last_year_GA_SC = do_sum(xl_map,7,range(13,22))
    xl_map[7][3] = retail_car_count_last_year_GA_SC

    retail_car_percentage_GA_SC = do_percentage(retail_car_count_current_year_GA_SC, retail_car_count_last_year_GA_SC)
    xl_map[8][3] = retail_car_percentage_GA_SC


    # reatail car total current year
    retail_car_count_current_year_Total = sum([retail_car_count_current_year_ILL,retail_car_count_current_year_GA_SC])
    xl_map[6][1]=retail_car_count_current_year_Total

    # reatail car total last year
    retail_car_count_last_year_Total = sum([retail_car_count_last_year_ILL,retail_car_count_last_year_GA_SC])
    xl_map[7][1]=retail_car_count_last_year_Total

    retail_car_total_percentage = do_percentage(retail_car_count_current_year_Total, retail_car_count_last_year_Total)
    xl_map[8][1] = retail_car_total_percentage


    #tota cars will be same as car count
    xl_map[10][3]=car_count_current_year_GA_SC
    xl_map[11][3]=car_count_last_year_GA_SC
    xl_map[12][3]=car_count_percentage_GA_SC

    xl_map[10][2] = car_count_current_year_ILL
    xl_map[11][2] = car_count_last_year_ILL
    xl_map[12][2] = car_count_percentage_ILL

    xl_map[10][1]= car_count_current_year_Total
    xl_map[11][1]= car_count_last_year_Total
    xl_map[12][1]= car_count_total_percentage


    # Retail Revenue current year - ILL
    retail_revenue_current_year_ILL = do_sum(xl_map,14,range(3,13))
    xl_map[14][2] = retail_revenue_current_year_ILL

    # Retail Revenue last year - ILL
    retail_revenue_last_year_ILL = do_sum(xl_map,15,range(3,13))
    xl_map[15][2] = retail_revenue_last_year_ILL

    retail_revenue_percentage_ILL = do_percentage(retail_revenue_current_year_ILL, retail_revenue_last_year_ILL)
    xl_map[16][2] = retail_revenue_percentage_ILL


    # Retail Revenue current year - GA_SC
    retail_revenue_current_year_GA_SC = do_sum(xl_map,14,range(13,22))
    xl_map[14][3] = retail_revenue_current_year_GA_SC

    # Retail Revenue last year - GA_SC
    retail_revenue_last_year_GA_SC = do_sum(xl_map,15,range(13,22))
    xl_map[15][3] = retail_revenue_last_year_GA_SC

    retail_revenue_percentage_GA_SC = do_percentage(retail_revenue_current_year_GA_SC, retail_revenue_last_year_GA_SC)
    xl_map[16][3] = retail_revenue_percentage_GA_SC


    # # reatail car total current year
    retail_revenue_current_year_Total = sum([retail_revenue_current_year_ILL,retail_revenue_current_year_GA_SC])
    xl_map[14][1]=retail_revenue_current_year_Total

    # reatail car total last year
    retail_revenue_last_year_Total = sum([retail_revenue_last_year_ILL,retail_revenue_last_year_GA_SC])
    xl_map[15][1]=retail_revenue_last_year_Total

    retail_revenue_percentage_GA_SC = do_percentage(retail_revenue_current_year_Total, retail_revenue_last_year_Total)
    xl_map[16][1] = retail_revenue_percentage_GA_SC


    #Revnue current year - ILL
    Total_revenue_current_year_ILL = do_sum(xl_map,18,range(3,13))
    xl_map[18][2] = Total_revenue_current_year_ILL

    #Revnue last year - ILL
    Total_revenue_last_year_ILL = do_sum(xl_map,19,range(3,13))
    xl_map[19][2] = Total_revenue_last_year_ILL

    Total_revenue_percentage_ILL = do_percentage(Total_revenue_current_year_ILL, Total_revenue_last_year_ILL)
    xl_map[20][2] = Total_revenue_percentage_ILL


    #Revnue current year - GA_SC
    Total_revenue_current_year_GA_SC = do_sum(xl_map,18,range(13,22))
    xl_map[18][3] = Total_revenue_current_year_GA_SC

    #Revnue last year - GA_SC
    Total_revenue_last_year_GA_SC = do_sum(xl_map,19,range(13,22))
    xl_map[19][3] = Total_revenue_last_year_GA_SC

    Total_revenue_percentage_GA_SC = do_percentage(Total_revenue_current_year_GA_SC, Total_revenue_last_year_GA_SC)
    xl_map[20][3] = Total_revenue_percentage_GA_SC


    # Revebue Total current year
    Total_revenue_current_year = sum([Total_revenue_current_year_ILL,Total_revenue_current_year_GA_SC])
    xl_map[18][1]=Total_revenue_current_year

    # Revebue Total last year
    Total_revenue_last_year = sum([Total_revenue_last_year_ILL,Total_revenue_last_year_GA_SC])
    xl_map[19][1]=Total_revenue_last_year

    Total_revenue_percentage = do_percentage(Total_revenue_current_year, Total_revenue_last_year)
    xl_map[20][1] = Total_revenue_percentage


    # Revenue and Total revenue are same
    xl_map[22][2] =Total_revenue_current_year_ILL
    xl_map[23][2] =Total_revenue_last_year_ILL
    xl_map[24][2] =Total_revenue_percentage_ILL

    xl_map[22][3]= Total_revenue_current_year_GA_SC
    xl_map[23][3]= Total_revenue_last_year_GA_SC
    xl_map[24][3]= Total_revenue_percentage_GA_SC

    xl_map[22][1] =Total_revenue_current_year
    xl_map[23][1] =Total_revenue_last_year
    xl_map[24][1] =Total_revenue_percentage


    #Average Retail Visit current year - IIL
    Current_average_retail_visit_IIL =Average_retail_visit_IL_function(retail_revenue_current_year_ILL,
                                     retail_car_count_current_year_ILL)
    xl_map[26][2]=round(Current_average_retail_visit_IIL,2)

    #Average Retail Visit last year - IIL
    last_average_retail_visit_IIL =Average_retail_visit_IL_function(retail_revenue_last_year_ILL,
                                     retail_car_count_last_year_ILL)
    xl_map[27][2]=round(last_average_retail_visit_IIL,2)

    average_retail_visit_percentage_IIL = do_percentage(Current_average_retail_visit_IIL,
                                                        last_average_retail_visit_IIL)
    xl_map[28][2] = average_retail_visit_percentage_IIL


    #Average Retail Visit current year - GA_SC
    current_average_retail_visit_GA_SC = Average_retail_visit_GA_SC_fucntion(retail_revenue_current_year_GA_SC,
                                                                            retail_car_count_current_year_GA_SC)
    xl_map[26][3] = round(current_average_retail_visit_GA_SC,2)

    #Average Retail Visit current year - GA_SC
    last_average_retail_visit_GA_SC = Average_retail_visit_GA_SC_fucntion(retail_revenue_last_year_GA_SC,
                                                                            retail_car_count_last_year_GA_SC)
    xl_map[27][3] = round(last_average_retail_visit_GA_SC,2)

    average_retail_visit_percentage_GA_SC = do_percentage(current_average_retail_visit_GA_SC,
                                                        last_average_retail_visit_GA_SC)
    xl_map[28][3] = average_retail_visit_percentage_GA_SC


    #Avergae Retail visit Total - current year
    current_average_retail_visit_Total_val =Average_retail_visit_Total_function(retail_revenue_current_year_Total,
                                                                                retail_car_count_current_year_Total)
    xl_map[26][1] = round(current_average_retail_visit_Total_val,2)

    #Avergae Retail visit Total - last year
    last_average_retail_visit_Total_val =Average_retail_visit_Total_function(retail_revenue_last_year_Total,
                                                                        retail_car_count_last_year_Total)
    xl_map[27][1] = round(last_average_retail_visit_Total_val,2)

    average_retail_visit_total_percentage = do_percentage(current_average_retail_visit_Total_val,
                                                        last_average_retail_visit_Total_val)
    xl_map[28][1] = average_retail_visit_total_percentage


    #Average Member visit current year - IIL
    current_average_memeber_visit_ILL =Average_memeber_visit_ILL_function(Total_revenue_current_year_ILL,retail_revenue_current_year_ILL,
                                                                      car_count_current_year_ILL, retail_car_count_current_year_ILL)
    xl_map[30][2] = round(current_average_memeber_visit_ILL,2)

    #Average Member visit last year - IIL
    last_average_memeber_visit_ILL =Average_memeber_visit_ILL_function(Total_revenue_last_year_ILL,retail_revenue_last_year_ILL,
                                                                      car_count_last_year_ILL, retail_car_count_last_year_ILL)
    xl_map[31][2] = round(last_average_memeber_visit_ILL,2)


    #Average Member visit current year - GA_SC
    current_average_memeber_visit_GA_SC = Average_memeber_visit_GA_SC_function(Total_revenue_current_year_GA_SC,retail_revenue_current_year_GA_SC,
                                                                            retail_car_count_current_year_GA_SC,car_count_current_year_GA_SC)
    xl_map[30][3] = round(current_average_memeber_visit_GA_SC,2)

    #Average Member visit last year - GA_SC
    last_average_memeber_visit_GA_SC = Average_memeber_visit_GA_SC_function(Total_revenue_last_year_GA_SC,retail_revenue_last_year_GA_SC,
                                                                            retail_car_count_last_year_GA_SC,car_count_last_year_GA_SC)
    xl_map[31][3] = round(last_average_memeber_visit_GA_SC,2)


    #Average Member visit Total - current Year
    current_average_memeber_visit_Total = Average_memeber_visit_Total_function(Total_revenue_current_year,retail_revenue_current_year_Total,
                                                                          car_count_current_year_Total,retail_car_count_current_year_Total)
    xl_map[30][1] = round(current_average_memeber_visit_Total,2)

    #Average Member visit Total - last Year
    last_average_memeber_visit_Total =Average_memeber_visit_Total_function(Total_revenue_last_year,retail_revenue_last_year_Total,
                                                                          car_count_last_year_Total,retail_car_count_last_year_Total)
    xl_map[31][1] = round(last_average_memeber_visit_Total,2)


    #Staff Hours current year
    staff_hours_current_year_ILL = do_sum(xl_map,33,range(3,13))
    xl_map[33][2] = staff_hours_current_year_ILL

    staff_hours_current_year_GA_SC = do_sum(xl_map,33,range(13,22))
    xl_map[33][3] = staff_hours_current_year_GA_SC

    staff_hours_current_year_Total = sum([staff_hours_current_year_ILL,staff_hours_current_year_GA_SC])
    xl_map[33][1] = staff_hours_current_year_Total

    # Staff_hours last year
    ill_last_year_labour_hours = []
    for loc_data in ill:
        if loc_data:
            val = loc_data.get("labour_hours_last_year", 0)
            if isinstance(val,float) or isinstance(val,int):
                ill_last_year_labour_hours.append(val)

    staff_hours_last_year_ILL = sum(ill_last_year_labour_hours)

    ga_sc_last_year_labour_hours = []
    for loc_data in ga_sc:
        if loc_data:
            val = loc_data.get("labour_hours_last_year", 0)
            if isinstance(val,float) or isinstance(val,int):
                ga_sc_last_year_labour_hours.append(val)

    staff_hours_last_year_GA_SC = sum(ga_sc_last_year_labour_hours)

    staff_hours_last_year_Total = sum([staff_hours_last_year_ILL, staff_hours_last_year_GA_SC])


    # Total cars per man hour current year - IIL
    current_total_cars_per_man_hour_ILL = Total_cars_per_man_hour_ILL_function(car_count_current_year_ILL,staff_hours_current_year_ILL)
    xl_map[35][2] = round(current_total_cars_per_man_hour_ILL,2)

    # Total cars per man hour last year - IIL
    last_total_cars_per_man_hour_ILL =Total_cars_per_man_hour_ILL_function(car_count_last_year_ILL,staff_hours_last_year_ILL)
    xl_map[36][2] = round(last_total_cars_per_man_hour_ILL,2)

    total_cars_per_man_hour_percentage_ILL = do_percentage(current_total_cars_per_man_hour_ILL,
                                                        last_total_cars_per_man_hour_ILL)
    xl_map[37][2] = total_cars_per_man_hour_percentage_ILL


    # Total cars per man hour current year - GA_SC
    current_total_cars_per_man_hour_GA_SC = Total_cars_per_man_hour_GA_SC_function(car_count_current_year_GA_SC, staff_hours_current_year_GA_SC)
    xl_map[35][3] = round(current_total_cars_per_man_hour_GA_SC,2)

    # Total cars per man hour current year - GA_SC
    last_total_cars_per_man_hour_GA_SC = Total_cars_per_man_hour_GA_SC_function(car_count_last_year_GA_SC, staff_hours_last_year_GA_SC)
    xl_map[36][3] = round(last_total_cars_per_man_hour_GA_SC,2)

    total_cars_per_man_hour_percentage_GA_SC = do_percentage(current_total_cars_per_man_hour_GA_SC,
                                                        last_total_cars_per_man_hour_GA_SC)
    xl_map[37][3] = total_cars_per_man_hour_percentage_GA_SC


    # Total cars per man hour current year total
    current_total_cars_per_man_hour_total =Total_cars_per_man_hour_total_function(car_count_current_year_Total,
                                                                            staff_hours_current_year_Total)
    xl_map[35][1] = round(current_total_cars_per_man_hour_total,2)

    # Total cars per man hour last year total
    last_total_cars_per_man_hour_total =Total_cars_per_man_hour_total_function(car_count_last_year_Total,
                                                                            staff_hours_last_year_Total)
    xl_map[36][1] = round(last_total_cars_per_man_hour_total,2)

    total_cars_per_man_hour_percentage = do_percentage(current_total_cars_per_man_hour_total,
                                                        last_total_cars_per_man_hour_total)
    xl_map[37][1] = total_cars_per_man_hour_percentage


    #Total club plans sold - current year
    Total_club_plans_sold_current_year_ILL = do_sum(xl_map,39,range(3,13))
    xl_map[39][2] = Total_club_plans_sold_current_year_ILL

    Total_club_plans_sold_current_year_GA_SC = do_sum(xl_map,39,range(13,22))
    xl_map[39][3] = Total_club_plans_sold_current_year_GA_SC

    Total_club_plans_sold_current_year_Total = sum([Total_club_plans_sold_current_year_ILL,Total_club_plans_sold_current_year_GA_SC])
    xl_map[39][1] = Total_club_plans_sold_current_year_Total


    # Total club plans sold
    ill_last_year_total_club_sold = [loc_data.get("arm_plans_sold_cnt_last_year", 0) for loc_data in ill if loc_data]
    Total_club_plans_sold_last_year_ILL = sum(ill_last_year_total_club_sold)

    ga_sc_last_year_total_club_sold = [loc_data.get("arm_plans_sold_cnt_last_year", 0) for loc_data in ga_sc if loc_data]
    Total_club_plans_sold_last_year_GA_SC = sum(ga_sc_last_year_total_club_sold)

    Total_club_plans_sold_last_year_Total = sum([Total_club_plans_sold_last_year_ILL, Total_club_plans_sold_last_year_GA_SC])


    #Conversion Rate current year - IIL
    current_conversion_rate_ILL = Conversion_rate_ILL_function(Total_club_plans_sold_current_year_ILL,
                                                          retail_car_count_current_year_ILL)
    xl_map[40][2] = round((current_conversion_rate_ILL * 100),2)

    #Conversion Rate last year - IIL
    last_conversion_rate_ILL = Conversion_rate_ILL_function(Total_club_plans_sold_last_year_ILL,
                                                          retail_car_count_last_year_ILL)
    xl_map[41][2] = round((last_conversion_rate_ILL * 100),2)

    conversion_rate_percentage_IIL = do_percentage(current_conversion_rate_ILL,
                                                   last_conversion_rate_ILL)
    xl_map[42][2] = conversion_rate_percentage_IIL


    #Conversion Rate current year - GA_SC
    current_conversion_rate_GA_SC = Conversion_rate_GA_SC_function(Total_club_plans_sold_current_year_GA_SC,
                                                                  retail_car_count_current_year_GA_SC)
    xl_map[40][3]= round((current_conversion_rate_GA_SC * 100),2)

    #Conversion Rate last year - GA_SC
    last_conversion_rate_GA_SC = Conversion_rate_GA_SC_function(Total_club_plans_sold_last_year_GA_SC,
                                                                  retail_car_count_last_year_GA_SC)
    xl_map[41][3]= round((last_conversion_rate_GA_SC * 100),2)

    conversion_rate_percentage_GA_SC = do_percentage(current_conversion_rate_GA_SC,
                                                     last_conversion_rate_GA_SC)
    xl_map[42][3] = conversion_rate_percentage_GA_SC


    # Total Conversion Rate -current year
    current_conversion_rate_Total = Conversion_rate_Total_function(Total_club_plans_sold_current_year_Total,
                                                                    retail_car_count_current_year_Total)
    xl_map[40][1] = round((current_conversion_rate_Total * 100),2)

    # Total Conversion Rate -current year
    last_conversion_rate_Total = Conversion_rate_Total_function(Total_club_plans_sold_last_year_Total,
                                                                    retail_car_count_last_year_Total)
    xl_map[41][1] = round((last_conversion_rate_Total * 100),2)

    conversion_rate_percentage = do_percentage(current_conversion_rate_Total,
                                                last_conversion_rate_Total)
    xl_map[42][1] = conversion_rate_percentage


    #Total club plan members
    Total_club_planmembers_ILL = do_sum(xl_map,44,range(3,13))
    xl_map[44][2] = Total_club_planmembers_ILL

    Total_club_planmembers_GA_SC = do_sum(xl_map,44,range(13,22))
    xl_map[44][3] = Total_club_planmembers_GA_SC

    Total_club_planmembers_Total = sum([Total_club_planmembers_ILL,Total_club_planmembers_GA_SC])
    xl_map[44][1] = Total_club_planmembers_Total

    # car count percetage for all locations
    car_count_row = 2
    for i in range(3,22):
        car_count_current = xl_map[car_count_row][i+1]
        car_count_last = xl_map[car_count_row+1][i+1]
        xl_map[car_count_row+2][i+1] = do_percentage(car_count_current,car_count_last)

    # Retail car count percetage for all locations
    retail_car_count_row = 6
    for i in range(3,22):
        retail_car_count_current = xl_map[retail_car_count_row][i+1]
        retail_car_count_last = xl_map[retail_car_count_row+1][i+1]
        xl_map[retail_car_count_row+2][i+1] = do_percentage(retail_car_count_current,retail_car_count_last)

    #Total Cars for all locations
    total_cars_row = 10
    for i in range(3,22):
        xl_map[total_cars_row][i+1] = xl_map[2][i+1] # current year
        xl_map[total_cars_row+1][i+1] = xl_map[3][i+1] # last year
        xl_map[total_cars_row+2][i+1] = xl_map[4][i+1] # percentage

    # Retail Revenue percentage for all location
    retail_revenue_row = 14
    for i in range(3,22):
        retail_revenue_current = xl_map[retail_revenue_row][i+1] # current year
        retail_revenue_last = xl_map[retail_revenue_row+1][i+1] # last year
        xl_map[retail_revenue_row+2][i+1] = do_percentage(retail_revenue_current, retail_revenue_last) # percentage

    # Revenue percentage for all location
    revenue_row = 18
    for i in range(3,22):
        revenue_current = xl_map[revenue_row][i+1] # current year
        revenue_last = xl_map[revenue_row+1][i+1] # last year
        xl_map[revenue_row+2][i+1] = do_percentage(revenue_current, revenue_last) # percentage

    # Total Revenue calculation for all locations
    total_revenue_row = 22
    for i in range(3,22):
        xl_map[total_revenue_row][i+1] = xl_map[18][i+1] # current year
        xl_map[total_revenue_row+1][i+1] = xl_map[19][i+1] # last year
        xl_map[total_revenue_row+2][i+1] = xl_map[20][i+1] # percentage

    # coversion rate calculation for all locations
    conversion_rate_row = 40
    for i in range(3,22):
        conversion_rate_current = xl_map[conversion_rate_row][i+1] # current year
        conversion_rate_last = xl_map[conversion_rate_row+1][i+1] # last year
        xl_map[conversion_rate_row+2][i+1] = do_percentage(conversion_rate_current, conversion_rate_last) # percentage


    #Average_retail_visit
    average_retail_visit_row = 26
    for i in range(3,22):
        #current year
        retail_revenue_curent_year = xl_map[14][i+1]
        retail_car_count_curent_year = xl_map[6][i+1]

        current_average_retail_visit_val = retail_revenue_curent_year/retail_car_count_curent_year if retail_car_count_curent_year != 0 else 0
        xl_map[average_retail_visit_row][i+1]= round(current_average_retail_visit_val, 2) if current_average_retail_visit_val else 0

        #last year
        retail_revenue_last_year = xl_map[15][i+1]
        retail_car_count_last_year = xl_map[7][i+1]

        last_average_retail_visit_val = retail_revenue_last_year/retail_car_count_last_year if retail_car_count_last_year != 0 else 0
        xl_map[average_retail_visit_row+1][i+1]= round(last_average_retail_visit_val,2) if last_average_retail_visit_val else 0

        # percentage
        xl_map[average_retail_visit_row+2][i+1] = do_percentage(current_average_retail_visit_val,last_average_retail_visit_val)


    #Average member visit
    average_member_visit_row = 30
    for i in range(3,22):

        #current year
        total_revenue_cuurent_year = xl_map[18][i+1]
        total_revenue_cuurent_year = total_revenue_cuurent_year if isinstance(total_revenue_cuurent_year,int) or isinstance(total_revenue_cuurent_year,float) else 0

        retail_revenue_current_year_1 = xl_map[14][i+1]

        total_cars_current_year = xl_map[10][i+1]
        total_cars_current_year = total_cars_current_year if isinstance(total_cars_current_year,int) or isinstance(total_cars_current_year,float) else 0

        retail_car_count_current_year_1  = xl_map[6][i+1]

        current_average_member_visit_val = (total_revenue_cuurent_year-retail_revenue_current_year_1)/(total_cars_current_year-retail_car_count_current_year_1) if total_cars_current_year-retail_car_count_current_year_1 !=0 else ""
        xl_map[average_member_visit_row][i+1] = round(current_average_member_visit_val,2) if current_average_member_visit_val else ""

        #last year
        total_revenue_last_year = xl_map[19][i+1]
        total_revenue_last_year = total_revenue_last_year if isinstance(total_revenue_last_year,int) or isinstance(total_revenue_last_year,float) else 0

        retail_revenue_last_year_1 = xl_map[15][i+1]

        total_cars_last_year = xl_map[11][i+1]
        total_cars_last_year = total_cars_last_year if isinstance(total_cars_last_year,int) or isinstance(total_cars_last_year,float) else 0

        retail_car_count_last_year_1  = xl_map[7][i+1]

        last_average_member_visit_val = (total_revenue_last_year-retail_revenue_last_year_1)/(total_cars_last_year-retail_car_count_last_year_1) if total_cars_last_year-retail_car_count_last_year_1 !=0 else ""

        xl_map[average_member_visit_row+1][i+1] = round(last_average_member_visit_val,2) if last_average_member_visit_val else ""

        #percentage


    #Total cars per man hour - last year for all locations
    total_cars_per_man_hour_row = 35
    for index, place_dict in enumerate(all_locations):
        if not place_dict:
            continue

        #current year
        total_cars_current_year_1 = xl_map[10][index+4]
        total_cars_current_year_1 = total_cars_current_year_1 if isinstance(total_cars_current_year_1,int) or isinstance(total_cars_current_year_1,float) else 0
        staff_hours_current_year = xl_map[33][index+4]
        staff_hours_current_year = staff_hours_current_year if isinstance(staff_hours_current_year,int) or isinstance(staff_hours_current_year,float) else 0

        current_total_cars_per_man_hour_val = (total_cars_current_year_1 / staff_hours_current_year) if staff_hours_current_year !=0 else ""
        xl_map[total_cars_per_man_hour_row][i+1] = round(current_total_cars_per_man_hour_val,2) if current_total_cars_per_man_hour_val else ""

        #last year
        total_cars_last_year_1 = xl_map[11][index+4]
        total_cars_last_year_1 = total_cars_last_year_1 if isinstance(total_cars_last_year_1,int) or isinstance(total_cars_last_year_1,float) else 0
        staff_hours_last_year = place_dict.get('labour_hours_last_year',0)
        staff_hours_last_year = staff_hours_last_year if isinstance(staff_hours_last_year,int) or isinstance(staff_hours_last_year,float) else 0

        last_total_cars_per_man_hour_val = (total_cars_last_year_1 / staff_hours_last_year) if staff_hours_last_year !=0 else ""
        xl_map[total_cars_per_man_hour_row+1][index+4] = round(last_total_cars_per_man_hour_val,2) if last_total_cars_per_man_hour_val else ""

        #percentage
        xl_map[total_cars_per_man_hour_row+2][index+4] = do_percentage(current_total_cars_per_man_hour_val,last_total_cars_per_man_hour_val)


    # Define cell styles (assuming they are the same as before)
    bg_color = PatternFill(start_color='0b3040', end_color='0b3040', fill_type='solid')
    font_color = Font(color='FFFFFF')

    bg_color_index = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    font_color_index = Font(color='000000')

    darkgreen_format = PatternFill(start_color='0ee85e', end_color='0ee85e', fill_type='solid')
    light_green_format = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    darkred_format = PatternFill(start_color='FC0303', end_color='FC0303', fill_type='solid')
    lightred_format = PatternFill(start_color='D98484', end_color='D98484', fill_type='solid')
    yellow=  PatternFill(start_color='D0D48A', end_color='D0D48A', fill_type='solid')

    for row in range(len(xl_map)):
        for col in range(len(xl_map[row])):
            val = xl_map[row][col]
            cell = worksheet.cell(row=row+1, column=col+1, value=val)  # offset by 0 rows for header and comment

            if row == 1 and col != 0:
                cell.fill = bg_color
                cell.font = font_color

            elif col == 0 and row in [2,6,10,14,18,22,26,30,33,35,39,40,44]:
                cell.fill = bg_color_index
                cell.font = font_color_index

            elif col==0 and 1 < row > 45:
                cell.font = font_color_index

    # Add legend or additional information below the table
    legend_start_row = 47

    legend_styles = {
        "Very Concerning": PatternFill(start_color="fc0303", end_color="fc0303", fill_type="solid"),
        "Concerning": PatternFill(start_color="d98484", end_color="d98484", fill_type="solid"),
        "Neutral": PatternFill(start_color="d0d48a", end_color="d0d48a", fill_type="solid"),
        "Positive": PatternFill(start_color="8ad493", end_color="8ad493", fill_type="solid"),
        "Very Positive": PatternFill(start_color="0ee85e", end_color="0ee85e", fill_type="solid"),
    }

    worksheet.cell(row=legend_start_row, column=1, value='Legend')
    worksheet.cell(row=legend_start_row + 1, column=1, value='Very Concerning').fill = legend_styles["Very Concerning"]
    worksheet.cell(row=legend_start_row + 2, column=1, value='Concerning').fill = legend_styles["Concerning"]
    worksheet.cell(row=legend_start_row + 3, column=1, value='Neutral').fill = legend_styles["Neutral"]
    worksheet.cell(row=legend_start_row + 4, column=1, value='Positive').fill = legend_styles["Positive"]
    worksheet.cell(row=legend_start_row + 5, column=1, value='Very Positive').fill = legend_styles["Very Positive"]

    #setting cloumn width as deafult
    column_width = 25  # You can change this value to whatever width you need
    first_col_width = 35
    for col in range(1, 24):  # Columns A to W are 1 to 23
        column_letter = get_column_letter(col)
        if col==1:
            worksheet.column_dimensions[column_letter].width = first_col_width
        else:
            worksheet.column_dimensions[column_letter].width = column_width


    for row in range(3,46):
        # cell0 = worksheet.cell(row=row,column=1)
        cell1 = worksheet.cell(row=row,column=2)
        cell2 = worksheet.cell(row=row,column=3)
        cell3 = worksheet.cell(row=row,column=4)

        if row == 3:
            # cell0.border=thick_border_bottom
            cell1.border= Border(
                left=Side(style='thick'),
                top=Side(style='thick')
            )
            cell2.border= Border(
                top=Side(style='thick')
            )
            cell3.border=Border(
                right=Side(style='thick'),
                top=Side(style='thick')
            )
        elif row == 45:
            cell1.border= Border(
                left=Side(style='thick'),
                bottom=Side(style='thick')
            )
            cell2.border= Border(
                bottom=Side(style='thick')
            )
            cell3.border=Border(
                right=Side(style='thick'),
                bottom=Side(style='thick')
            )
        else:
            # cell0.border=thick_border
            cell1.border=Border(
                left=Side(style='thick')
            )

            cell3.border=Border(
                right=Side(style='thick')
            )

    #number formates 3 dots
    for row in worksheet.iter_rows():
        for cell in row:
            row_index = cell.row
            if isinstance(cell.value, (int, float)) and row_index in [5,9,13,17,21,25,29,36,37,38,43]:
                cell.number_format = '#,##0.0'
            elif isinstance(cell.value, (int, float)) and row_index in [27,28,31,32,41,42]:
                cell.number_format = '#,##0.00'
            elif isinstance(cell.value, (int, float)): #:
                cell.number_format = '#,##0'
    #Doller sysmbol
    for row in [15,16,19,20,23,24]:
        cell1 = worksheet.cell(row=row,column=2)
        cell2 = worksheet.cell(row=row,column=3)
        cell3 = worksheet.cell(row=row,column=4)

        cells=[cell1,cell2,cell3]
        for cell in cells:
            if isinstance(cell.value, (int, float)) and cell.value >= 1000:
                cell.number_format = '"$"#,##0'


    colours = darkgreen_format,light_green_format,darkred_format,lightred_format

    print("Conversion Rate")
    logger.info("Conversion Rate")
    set_colour_new(current_conversion_rate_ILL,last_conversion_rate_ILL,41,3,worksheet,colours)  #conversation rate
    set_colour_new(current_conversion_rate_GA_SC,last_conversion_rate_GA_SC,41,4,worksheet,colours)
    set_colour_new(current_conversion_rate_Total,last_conversion_rate_Total,41,2,worksheet,colours)

    print('Car count')
    logger.info('Car count')
    set_colour_new(car_count_current_year_ILL, car_count_last_year_ILL, 3, 3, worksheet, colours) #car count ill
    set_colour_new(car_count_current_year_GA_SC, car_count_last_year_GA_SC, 3, 4, worksheet, colours) #car count ga-sc
    set_colour_new(car_count_current_year_Total, car_count_last_year_Total, 3, 2, worksheet, colours) # total car count

    print('Retail count')
    logger.info('Retail count')
    set_colour_new(retail_car_count_current_year_ILL, retail_car_count_last_year_ILL, 7, 3, worksheet, colours) # retail count ill
    set_colour_new(retail_car_count_current_year_GA_SC, retail_car_count_last_year_GA_SC, 7, 4, worksheet, colours) # retail count ga_sc
    set_colour_new(retail_car_count_current_year_Total, retail_car_count_last_year_Total, 7, 2, worksheet, colours) # retail count

    print('Total car count')
    logger.info('Total car count')
    set_colour_new(car_count_current_year_ILL, car_count_last_year_ILL, 11, 3, worksheet, colours) # total car count ill
    set_colour_new(car_count_current_year_GA_SC, car_count_last_year_GA_SC, 11, 4, worksheet, colours) # total car count ga-sc
    set_colour_new(car_count_current_year_Total, car_count_last_year_Total, 11, 2, worksheet, colours) # total car count

    print('Retail Revenue')
    logger.info('Retail Revenue count')
    set_colour_new(retail_revenue_current_year_ILL, retail_revenue_last_year_ILL, 15, 3, worksheet, colours)
    set_colour_new(retail_revenue_current_year_GA_SC, retail_revenue_last_year_GA_SC, 15, 4, worksheet, colours)
    set_colour_new(retail_revenue_current_year_Total, retail_revenue_last_year_Total, 15, 2, worksheet, colours)

    print('Revenue count')
    logger.info('Revenue count')
    set_colour_new(Total_revenue_current_year_ILL, Total_revenue_last_year_ILL, 19, 3, worksheet, colours)
    set_colour_new(Total_revenue_current_year_GA_SC, Total_revenue_last_year_GA_SC, 19, 4, worksheet, colours)
    set_colour_new(Total_revenue_current_year, total_revenue_last_year, 19, 2, worksheet, colours)

    print('Total Revenue count')
    logger.info('Total Revenue count')
    set_colour_new(Total_revenue_current_year_ILL, Total_revenue_last_year_ILL, 23, 3, worksheet, colours)
    set_colour_new(Total_revenue_current_year_GA_SC, Total_revenue_last_year_GA_SC, 23, 4, worksheet, colours)
    set_colour_new(Total_revenue_current_year, total_revenue_last_year, 23, 2, worksheet, colours)

    print('Avg retail visit')
    logger.info('Avg retail visit')
    set_colour_for_avg_retail(Current_average_retail_visit_IIL, last_average_retail_visit_IIL, 27, 3, worksheet, colours)
    set_colour_for_avg_retail(current_average_retail_visit_GA_SC, last_average_retail_visit_GA_SC, 27, 4, worksheet, colours)
    set_colour_for_avg_retail(current_average_retail_visit_Total_val, last_average_retail_visit_Total_val, 27, 2, worksheet, colours)

    print('Total Cars Per Man Hour')
    logger.info('Total Cars Per Man Hour')
    set_colour_new(current_total_cars_per_man_hour_ILL, last_total_cars_per_man_hour_ILL, 36, 3, worksheet, colours)
    set_colour_new(current_total_cars_per_man_hour_GA_SC, last_total_cars_per_man_hour_GA_SC, 36, 4, worksheet, colours)
    set_colour_new(current_total_cars_per_man_hour_total, last_total_cars_per_man_hour_total, 36, 2, worksheet, colours)

    print()

    loc_names = ["Sudz - Beverly",'Fuller-Calumet',
                "Fuller-Cicero","Fuller-Matteson","Fuller-Elgin",
                "Splash-Peoria","Getaway-Macomb","Getaway-Morton",
                "Getaway-Ottawa","Getaway-Peru","Sparkle-Belair",
                "Sparkle-Evans","Sparkle-Furrys Ferry","Sparkle-Greenwood",
                "Sparkle-Grovetown 1","Sparkle-Grovetown 2","Sparkle-North Augusta",
                "Sparkle-Peach Orchard","Sparkle-Windsor Spring"]

    for i in range(len(loc_names)):
        #conversion rate
        print("conversion rate")
        logger.info("conversion rate")
        current_year_conversation_rate = xl_map[40][i+4]
        last_year_conversation_rate =  xl_map[41][i+4]
        set_colour_new(current_year_conversation_rate,last_year_conversation_rate,41,i+5,worksheet,colours) #conversation rate colours

        #Total Revenue
        print("Total Revenue")
        logger.info("Total Revenue")
        current_year_revenue_total = xl_map[22][i+4]
        last_year_revenue_total = xl_map[23][i+4]
        set_colour_new(current_year_revenue_total, last_year_revenue_total, 23, i+5,worksheet,colours)

        #Revenue
        print("Revenue")
        logger.info("Revenue")
        current_year_revenue_total = xl_map[18][i+4]
        last_year_revenue_total = xl_map[19][i+4]
        set_colour_new(current_year_revenue_total, last_year_revenue_total, 19, i+5,worksheet,colours)

        #Car Count
        print("Car Count")
        logger.info("Car Count")
        last_year_car_count = xl_map[3][i+4]
        current_year_car_count = xl_map[2][i+4]
        set_colour_new(current_year_car_count, last_year_car_count, 3, i+5, worksheet, colours)

        # Reatil car count
        print("Reatil car count")
        logger.info("Reatil car count")
        last_year_retail_car_count = xl_map[7][i+4]
        curr_year_retail_car_count = xl_map[6][i+4]
        set_colour_new(curr_year_retail_car_count, last_year_retail_car_count, 7, i+5, worksheet, colours)

        #Total Car Count
        print("Total Car Count")
        logger.info("Total Car Count")
        last_year_car_count = xl_map[11][i+4]
        current_year_car_count = xl_map[10][i+4]
        set_colour_new(current_year_car_count, last_year_car_count, 11, i+5, worksheet, colours)

        # Retail Revenue
        print("Retail Revenue")
        logger.info("Retail Revenue")
        last_year_retail_revenue = xl_map[15][i+4]
        current_year_retail_revenue = xl_map[14][i+4]
        set_colour_new(current_year_retail_revenue, last_year_retail_revenue, 15, i+5, worksheet, colours)

        #Avg. Retail Visit
        print("Avg. Retail Visit")
        logger.info("Avg. Retail Visit")
        last_year_avg_ratail_visit = xl_map[27][i+4]
        current_year_avg_ratail_visit = xl_map[26][i+4]
        set_colour_for_avg_retail(current_year_avg_ratail_visit, last_year_avg_ratail_visit, 27, i+5, worksheet, colours)

        # Total Cars Per Man Hour
        print("Total Cars Per Man Hour")
        logger.info("Total Cars Per Man Hour")
        last_year_cars_labour_hours = xl_map[36][i+4]
        if xl_map[35][i+4] != '':
            current_year_labour_hours = xl_map[35][i+4]
            set_colour_new(current_year_labour_hours, last_year_cars_labour_hours, 36, i+5, worksheet, colours)

        print("\n"*2)

    #applying bold font
    # Define a bold font style
    bold_font = Font(bold=True)

    for xl_row in [3,7,11,15,19,23,27,31,34,36,40,41,45]:
        for row in worksheet.iter_rows(min_row=xl_row, max_row=xl_row, min_col=1, max_col=4):
            for cell in row:
                cell.font = bold_font

    # Save the modified workbook
    workbook.save(filename)


def get_year_for_storage():
    "will retun in '%Y' ==> 2024"

    today = datetime.today()

    return today.strftime("%Y")


def create_storage_directory(path):
    created_path=None
    try:
        path = os.path.join(data_path,path)
        os.makedirs(path, exist_ok=True)
        print(f"Directory '{path}' created successfully")
        created_path = path
    except OSError as error:
        print(f"Directory '{path}' cannot be created: {error}")
        logger.info(f"Directory '{path}' cannot be created: {error}")

    return created_path


if __name__=="__main__":
    from dates_generator import *
    from zero_value_check import check_zero_values
    import logging
    from logging_config import setup_logging
    import traceback
    setup_logging()
    logger = logging.getLogger(__name__)
    logger.info("started main script")

    load_dotenv()

    class emailConfig:
        # env_vars    = os.environ
        # FROM_EMAIL   = env_vars.get("FROM_EMAIL")
        # FROM_NAME      = env_vars.get("FROM_NAME")
        # SMTP_SERVER   = env_vars.get("SMTP_SERVER")
        # SMTP_PORT = env_vars.get("SMTP_PORT")
        # SMTP_USER=env_vars.get("SMTP_USER")
        # SMTP_PASSWORD=env_vars.get("SMTP_PASSWORD")
        # TO_EMAIL=env_vars.get("TO_EMAIL")
        FROM_EMAIL = 'sushilvarma@reluconsultancy.in'
        FROM_NAME = 'Sushil'
        SMTP_SERVER = 'smtp-mail.outlook.com'
        SMTP_PORT = 587
        SMTP_USER = 'sushilvarma@reluconsultancy.in'
        SMTP_PASSWORD = 'JaggaDaku<<162357>>'
        TO_EMAIL = 'varmasushil004@gmail.com'

    # Configuration
    subject = 'Sample Comparison report'
    # body = 'This is the body of the email.'
    to_email = emailConfig.TO_EMAIL
    from_email = emailConfig.FROM_EMAIL
    from_name = emailConfig.FROM_NAME
    smtp_server = emailConfig.SMTP_SERVER
    smtp_port = emailConfig.SMTP_PORT
    smtp_user = emailConfig.SMTP_USER
    smtp_password = emailConfig.SMTP_PASSWORD


    # cc_emails=["CR@SparkleCW.com","FZ@SparkleCW.com","Rick@SparkleStatus.com","Shane@SparkleStatus.com", "mgiamalis@firmament.com","tech@reluconsultancy.in"]
    # cc_emails=[]

    # # -----------------Actual script  ----------------------------#

    start_date_c_year="2024-11-01"
    end_date_c_year="2024-11-14"

    start_date_l_year="2023-11-01"
    end_date_l_year="2023-11-14"

    path = get_year_for_storage()
    storage_path = create_storage_directory(path)
    current_dates = get_dates_for_current_year((start_date_c_year, end_date_c_year))

    if current_dates[-1]:
        last_year_dates = get_dates_for_last_year((start_date_l_year, end_date_l_year))

        # for sitewatch
        start_date_current_year, end_date_current_year, start_date_last_year, end_date_last_year = format_dates_sitewatch(
            current_dates + last_year_dates
        )
        # print(f"sitewatch dates : {start_date_current_year} {end_date_current_year} {start_date_last_year} {end_date_last_year}")
        logger.info(f"sitewatch dates : {start_date_current_year} {end_date_current_year} {start_date_last_year} {end_date_last_year}")
        report_sitewatch = sitewatch_report("",start_date_current_year, end_date_current_year, start_date_last_year, end_date_last_year)


        # for washify
        start_date_current_year, end_date_current_year, start_date_last_year, end_date_last_year = format_dates_washify(
            current_dates + last_year_dates
        )
        # print(f"washify dates : {start_date_current_year} {end_date_current_year} {start_date_last_year} {end_date_last_year}")
        logger.info(f"washify dates: {start_date_current_year} {end_date_current_year} {start_date_last_year} {end_date_last_year}")
        report_washify = washify_report("",start_date_current_year, end_date_current_year, start_date_last_year, end_date_last_year)


        # for hamilton
        start_date_current_year, end_date_current_year, start_date_last_year, end_date_last_year = format_dates_hamilton(
            current_dates + last_year_dates
        )
        # print(f"hamilton dates : {start_date_current_year} {end_date_current_year} {start_date_last_year} {end_date_last_year}")
        logger.info(f"hamilton dates : {start_date_current_year} {end_date_current_year} {start_date_last_year} {end_date_last_year}")
        report_hamilton = hamilton_report(start_date_current_year, end_date_current_year, start_date_last_year, end_date_last_year)

        # Complete Data
        data = report_sitewatch
        data.update(report_washify)
        data.update(report_hamilton)

        with open("comparison_report_data.json", 'w') as f:
            json.dump(data, f, indent=4)

        # with open("comparison_report_data.json", 'r') as f:
        #     data = json.load(f)

        comment =f"Ending {end_date_current_year}"
        sheet_name= end_date_current_year.replace("/","-")
        filename=f"comparison_report_{path}.xlsx"
        # filename=f"comparison_report.xlsx"
        file_name_with_fullpath = os.path.join(storage_path,filename)
        prepare_xlmap(data,comment,sheet_name=sheet_name,filename=file_name_with_fullpath)

        print("final data:")
        print(data)

        # Directory containing Excel files
        directory_path = storage_path
        attachments = get_excel_files(directory_path)

        #Sending email to email address
        body = f'Comparison Report Ending {end_date_current_year}'
        # cc_emails = ['sushilvarma@reluconsultancy.in',"CR@SparkleCW.com","tech@reluconsultancy.in"]

        zero_val_check = check_zero_values(file_name_with_fullpath,sheet_name)
        # zero_val_check = False

        if zero_val_check:
            body = f'Error in Comparison report  Ending {end_date_current_year}'
            cc_emails=["CR@SparkleCW.com","FZ@SparkleCW.com","Rick@SparkleStatus.com","Shane@SparkleStatus.com", "mgiamalis@firmament.com","tech@reluconsultancy.in"]
            relu_emails= ["abhishekmeher@reluconsultancy.in","namangupta@reluconsultancy.in","vijaykumarmanthena@reluconsultancy.in"]
            cc_emails = cc_emails.extend(relu_emails)
            send_email_on_error(subject, body, to_email, from_email, from_name, smtp_server, smtp_port, smtp_user, smtp_password,cc_emails)
            logger.info(f"Error in Comparison Report: {traceback.print_exc()} ")
        else:
            body = f'Comparison report Ending {end_date_current_year}'
            send_email(subject, body, to_email, from_email, from_name, smtp_server, smtp_port, smtp_user, smtp_password, attachments, cc_emails)
    else:
        logger.info(f" End Date of current year {current_dates[-1]}. skipped report generation due to end date is None")

        # prepare_xlmap(data,comment,sheet_name=sheet_name)
        # # -----------------Actual script  ----------------------------#

        # -----------------cron jaob info ------------ ------------#
        # #Below script will run on every sunday 12 after noon
        #0 10 * * 1 /home/ubuntu/CAR_WASH_2/carwash_weekly/weekly_run.sh
        #sudo timedatectl set-timezone America/Chicago

    logger.info("!!! completed main script !!!!")

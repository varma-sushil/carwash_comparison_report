from datetime import datetime, timedelta
import os
import json
import time

import pandas as pd 
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import xlsxwriter

from sitewatch4 import sitewatchClient
from sitewatch4 import generate_past_4_weeks_days
from sitewatch4 import generate_past_4_week_days_full
import traceback 
import random

import sys
# Add the carwash directory to the sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))



tg_message=[]

current_folder_path = os.path.dirname(os.path.abspath(__file__))

cookies_path = os.path.join(current_folder_path,"cookies")

xlfile_path = os.path.join(current_folder_path,"sitewash_data.xlsx")

site_watch_latest_json = os.path.join(current_folder_path,"sitewatch_data_latest.json")

differenec_json = os.path.join(current_folder_path,"difference.json")

sites_df = pd.read_excel(xlfile_path)

orginations_lst = ['SPKLUS-001', 'SPKLUS-002', 
                   'SPKLUS-003', 'SPKLUS-004', 
                   'SPKLUS-005', 'SPKLUS-006', 
                   'SPKLUS-007', 'SPKLUS-008', 
                   'SPKLUS-009', 'SPKLUS-012', 
                   'SPKLUS-013', 'SPKLUS-014', 
                   'SPKLUS-015', 'SUDZWL-002']

client_names = ['Belair', 'Evans', 'North Augusta', 
                'Greenwood', 'Grovetown 1', 'Windsor Springs',
                'Furys Ferry (Martinez)', 'Peach Orchard Rd.', 
                'Grovetown 2', 'Cicero', 'Matteson', 'Sparkle Express ', 
                "Fuller's Calumet City "]

# wash_sales_lst= []
# wash_packages_lst =[]
# wash_extra_service_lst =[]
# gross_wash_sales_lst = []
# less_wash_sales_rdmd_lst =[]
# less_wash_discounts_lst=[]
# less_loyality_disc_lst = []

# net_site_sales_lst = []

# arm_plans_sold_lst  = []

# arm_plans_recharged_lst = []

# arm_plans_reedemed_lst = []

# arm_plans_terminated_lst = []

# prepaid_sold_lst = []

# less_prepaid_reedemed_lst = []

# online_sold_lst = []

# less_online_reedemed_lst = []

# free_washes_issued_lst = []

# less_paidouts_lst = []

# total_to_account_for_lst =[]

# deposits_lst = []

# total_xpt_cash_lst = []

# house_accounts_lst =[]

# over_short_lst = []

# cash_lst = []

# xpt_acceptors_lst = []

# xpt_dispensers_lst = []

# total_lst = []

# credit_card_list = []

# other_tenders_lst = []

# xpt_balancing_lst = []

# report_balance_lst = []

# picture_mismatch_lst = []

def wash_sales(section):
    wash_sales_lst = []
    
    # reports = section.get("reports")
    # for report in reports:
    #     wash_sales_structure ={
    #         "Wash_sales_Description":report.get("description"),
    #         "Wash_sales_price":report.get("price"),
    #         "Wash_sales_quantity":report.get("quantity"),
    #         "Wash_sales_amount":report.get("amount")
    #     }
    #     wash_sales_lst.append(wash_sales_structure)
    
    subtotals = section.get("subtotals")[0]
    
    # for subtotal in subtotals:
    #     wash_sales_structure ={
    #         "Wash_sales_Description":subtotal.get("description"),
    #         "Wash_sales_price":subtotal.get("price"),
    #         "Wash_sales_quantity":subtotal.get("quantity"),
    #         "Wash_sales_amount":subtotal.get("amount")
    #     }
    #     wash_sales_lst.append(wash_sales_structure)
    return subtotals.get("quantity")
        
    return wash_sales_lst

def wash_packages(section):
    wash_packages_lst = []

    reports = section.get("reports")
    subtotals = section.get("subtotals")
    
    for report in reports:
        wash_package_structure = {
            "Wash_packages_Description":report.get("description"),
            "Wash_packages_price":report.get("price"),
            "Wash_packages_quantity":report.get("quantity"),
            "Wash_packages_amount":report.get("amount"),
        }
        wash_packages_lst.append(wash_package_structure)
    
    for subtotal in subtotals:
        wash_package_structure = {
            "Wash_packages_Description":subtotal.get("description"),
            "Wash_packages_price":subtotal.get("price"),
            "Wash_packages_quantity":subtotal.get("quantity"),
            "Wash_packages_amount":subtotal.get("amount"),
        }
        wash_packages_lst.append(wash_package_structure)

    return wash_packages_lst
 
def wash_extra_services(section):
    wash_extra_service_lst = []
    reports = section.get("reports")
    subtotals  = section.get("subtotals")
    
    for report in reports:
        wash_extra_structure = {
            "Wash_Extra_Services_Description":report.get("description"),
            "Wash_Extra_Services_price":report.get("price"),
            "Wash_Extra_Services_quantity":report.get("quantity"),
            "Wash_Extra_Services_amout":report.get("amount")
        }
        wash_extra_service_lst.append(wash_extra_structure)
        
    for total in subtotals:
        wash_extra_structure = {
            "Wash_Extra_Services_Description":total.get("description"),
            "Wash_Extra_Services_price":total.get("price"),
            "Wash_Extra_Services_quantity":total.get("quantity"),
            "Wash_Extra_Services_amout":total.get("amount")
        }
        wash_extra_service_lst.append(wash_extra_structure)  
        
    return wash_extra_service_lst  

def gross_wash_sales(section):
    gross_wash_sales_lst = []
    gross_wash_sale_structure = {
            "Gross_Wash_Sales":section.get("totalAmount")
        }
        
    gross_wash_sales_lst.append(gross_wash_sale_structure)
    
    return gross_wash_sales_lst

def less_free_wash_rdmd(section):
    less_wash_sales_rdmd_lst = []
    reports = section.get("reports")
    subtotals = section.get("subtotals")
    
    for report in reports:
        less_free_wash_rmd_structure = {
            "Less_free_wash_rdmd_Description":report.get("description"),
            "Less_free_wash_rdmd_price":report.get("price"),
            "Less_free_wash_rdmd_quantity":report.get("quantity"),
            "Less_free_wash_rdmd_amount":report.get("amount"),
        }
        less_wash_sales_rdmd_lst.append(less_free_wash_rmd_structure)
    
    for subtotal in subtotals:
        less_free_wash_rmd_structure = {
            "Less_free_wash_rdmd_Description":subtotal.get("description"),
            "Less_free_wash_rdmd_price":subtotal.get("price"),
            "Less_free_wash_rdmd_quantity":subtotal.get("quantity"),
            "Less_free_wash_rdmd_amount":subtotal.get("amount"),
        }
        less_wash_sales_rdmd_lst.append(less_free_wash_rmd_structure)
    return less_wash_sales_rdmd_lst

def less_wash_discounts(section):
    less_wash_discounts_lst = []
    reports  = section.get("reports")  
    subtotals  = section.get("subtotals")  
    for report in reports:
        less_wash_discount_structure = {
            "Less_Wash_Discounts_Description":report.get("description"),
            "Less_Wash_Discounts_Price":report.get("price"),
            "Less_Wash_Discounts_quantity":report.get("quantity"),
            "Less_Wash_Discounts_amount":report.get("amount")
        }
        less_wash_discounts_lst.append(less_wash_discount_structure)

    for subtotal in subtotals:
        less_wash_discount_structure = {
            "Less_Wash_Discounts_Description":subtotal.get("description"),
            "Less_Wash_Discounts_Price":subtotal.get("price"),
            "Less_Wash_Discounts_quantity":subtotal.get("quantity"),
            "Less_Wash_Discounts_amount":subtotal.get("amount")
        }
        less_wash_discounts_lst.append(less_wash_discount_structure)    

    return less_wash_discounts_lst
    

def less_loyality_disc(section):
    less_loyality_disc_lst = []

    reports = section.get("reports")
    subtotals = section.get("subtotals")
    
    for report in reports:
        less_loyality_disc_structure = {
            "Less_Loyalty_disc_description":report.get("description"),
            "Less_Loyalty_disc_price":report.get("price"),
            "Less_Loyalty_disc_quantity":report.get("quantity"),
            "Less_Loyalty_disc_amount":report.get("amount")
        }
        less_loyality_disc_lst.append(less_loyality_disc_structure)
        
    for subtotal in subtotals:
        less_loyality_disc_structure = {
            "Less_Loyalty_disc_description":subtotal.get("description"),
            "Less_Loyalty_disc_price":subtotal.get("price"),
            "Less_Loyalty_disc_quantity":subtotal.get("quantity"),
            "Less_Loyalty_disc_amount":subtotal.get("amount")
        }
        less_loyality_disc_lst.append(less_loyality_disc_structure)

    return less_loyality_disc_lst

def net_site_sales(section):
    # net_site_sales_lst = []
    # net_site_sales_structue={
    #             "Net_site_sales":section.get("totalAmount")
    #         }
    # net_site_sales_lst.append(net_site_sales_structue)

    return section.get("totalAmount")
 
def arm_plans_sold(section):
    # arm_plans_sold_lst = []
    # reports = section.get("reports")
    # subtotals  = section.get("subtotals")
    
    # for report in reports:
    #     arm_plans_sold_structure = {
    #         "Arm_plan_sold_description":report.get("description"),
    #         "Arm_plan_sold_price":report.get("price"),
    #         "Arm_plan_sold_quantity":report.get("quantity"),
    #         "Arm_plan_sold_amount":report.get("amount")
    #     }
    #     arm_plans_sold_lst.append(arm_plans_sold_structure)
        
    # for subtotal in subtotals:
    #     arm_plans_sold_structure = {
    #         "Arm_plan_sold_description":subtotal.get("description"),
    #         "Arm_plan_sold_price":subtotal.get("price"),
    #         "Arm_plan_sold_quantity":subtotal.get("quantity"),
    #         "Arm_plan_sold_amount":subtotal.get("amount")
    #     }
    #     arm_plans_sold_lst.append(arm_plans_sold_structure)
        
    return section.get("totalQuantity",0.0)   


def arm_plans_recharged(section):
    data=None
    arm_plans_recharged_lst = []
    reports  = section.get("reports")   
    subtotals  = section.get("subtotals")   
    
    # for report in reports:
    #     arm_plans_recharged_structure = {
    #         "Arm_plan_recharged_description":report.get("description"),
    #         "Arm_plan_recharged_price":report.get("price"),
    #         "Arm_plan_recharged_quantity":report.get("quantity"),
    #         "Arm_plan_recharged_amount":report.get("amount")
    #     } 
    #     arm_plans_recharged_lst.append(arm_plans_recharged_structure)
    data = subtotals[0].get("amount")
    
    return data
    
        
    # for subtotal in subtotals:
    #     arm_plans_recharged_structure = {
    #         "Arm_plan_recharged_description":subtotal.get("description"),
    #         "Arm_plan_recharged_price":subtotal.get("price"),
    #         "Arm_plan_recharged_quantity":subtotal.get("quantity"),
    #         "Arm_plan_recharged_amount":subtotal.get("amount")
    #     } 
    #     arm_plans_recharged_lst.append(arm_plans_recharged_structure)
    
    # return arm_plans_recharged_lst

def arm_planes_reedemed(section):
    data={}
    
    # arm_plans_reedemed_lst = []
    # reports  = section.get("reports")
    # subtotals = section.get("subtotals")     
    
    # for report in reports:
    #     arm_plans_reedemed_structure = {
    #         "Arm_plan_redeemed_description":report.get("description"),
    #         "Arm_plan_redeemed_price":report.get("price"),
    #         "Arm_plan_redeemed_quantity":report.get("quantity"),
    #         "Arm_plan_redeemed_amount":report.get("amount"),

    #     }
    
    #     arm_plans_reedemed_lst.append(arm_plans_reedemed_structure)
    # for subtotal in subtotals:
    #     arm_plans_reedemed_structure = {
    #         "Arm_plan_redeemed_description":subtotal.get("description"),
    #         "Arm_plan_redeemed_price":subtotal.get("price"),
    #         "Arm_plan_redeemed_quantity":subtotal.get("quantity"),
    #         "Arm_plan_redeemed_amount":subtotal.get("amount"),

    #     }
    
    #     arm_plans_reedemed_lst.append(arm_plans_reedemed_structure)    

    data["arm_plans_reedemed_cnt"] = section.get("totalQuantity")
    data["arm_plans_reedemed_amt"] = section.get("totalAmount",0.0)*(-1)
    return data
    
    
def arm_plans_terminated(section):
    arm_plans_terminated_lst = []
    reports = section.get("reports")
    subtotals = section.get("subtotals")
    for report in reports:
        arm_plan_terminated_structure ={
            "Arm_plans_terminated_description":report.get("description"),
            "Arm_plans_terminated_price":report.get("price"),
            "Arm_plans_terminated_quantity":report.get("quantity"),
            "Arm_plans_terminated_amount":report.get("amount")
        }
        arm_plans_terminated_lst.append(arm_plan_terminated_structure)
    for subtotal in subtotals:
        arm_plan_terminated_structure ={
            "Arm_plans_terminated_description":subtotal.get("description"),
            "Arm_plans_terminated_price":subtotal.get("price"),
            "Arm_plans_terminated_quantity":subtotal.get("quantity"),
            "Arm_plans_terminated_amount":subtotal.get("amount")
        }
        arm_plans_terminated_lst.append(arm_plan_terminated_structure)
           
    return arm_plans_terminated_lst


def prepaid_sold(section):
    prepaid_sold_lst = []
    reports = section.get("reports")
    subtotals  = section.get("subtotals")
    
    for report in reports:
        prepaid_sold_structure = {
            "Prepaids_Sold_description":report.get("description"),
            "Prepaids_Sold_price":report.get("price"),
            "Prepaids_Sold_quantity":report.get("quantity"),
            "Prepaids_Sold_amount":report.get("amount"),
        }
        prepaid_sold_lst.append(prepaid_sold_structure)
        
    for subtotal in subtotals:
        prepaid_sold_structure = {
            "Prepaids_Sold_description":subtotal.get("description"),
            "Prepaids_Sold_price":subtotal.get("price"),
            "Prepaids_Sold_quantity":subtotal.get("quantity"),
            "Prepaids_Sold_amount":subtotal.get("amount"),
        }
        prepaid_sold_lst.append(prepaid_sold_structure)
        
    return prepaid_sold_lst
    
def less_prepaid_reedemed(section):
    less_prepaid_reedemed_lst = []
    
    reports  = section.get("reports")
    subtotals = section.get("subtotals")
    for report in reports:
        less_prepaid__reedemed_structure ={
            "Less_prepaids_redeemed_description":report.get("description"),
            "Less_prepaids_redeemed_price":report.get("price"),
            "Less_prepaids_redeemed_quantity":report.get("quantity"),
            "Less_prepaids_redeemed_amount":report.get("amount")
        }
        
        less_prepaid_reedemed_lst.append(less_prepaid__reedemed_structure)
        
    for subtotal in subtotals:
        less_prepaid__reedemed_structure ={
            "Less_prepaids_redeemed_description":subtotal.get("description"),
            "Less_prepaids_redeemed_price":subtotal.get("price"),
            "Less_prepaids_redeemed_quantity":subtotal.get("quantity"),
            "Less_prepaids_redeemed_amount":subtotal.get("amount")
        }
        
        less_prepaid_reedemed_lst.append(less_prepaid__reedemed_structure)
    
    return less_prepaid_reedemed_lst
 
def online_sold(section):
    data = []
    reports = section.get("reports")
    subtotals= section.get("subtotals")
    
    for report in  reports:
        online_sold_structure ={
            "online_sold_description":report.get("description"),
            "online_sold_price":report.get("price"),
            "online_sold_quantity":report.get("quantity"),
            "online_sold_amount":report.get("amount")
        }
        data.append(online_sold_structure)
        
    for subtotal in  subtotals:
        online_sold_structure ={
            "online_sold_description":subtotal.get("description"),
            "online_sold_price":subtotal.get("price"),
            "online_sold_quantity":subtotal.get("quantity"),
            "online_sold_amount":subtotal.get("amount")
        }
        data.append(online_sold_structure)
    
    return data
        
def less_online_reedemed(section):
    data = [] 
    reports =section.get("reports")
    subtotals= section.get("subtotals")
    
    for report in reports:
        less_online_reedemed_strcture = {
            "Less_online_redeemed_description":report.get("description"),
            "Less_online_redeemed_price":report.get("price"),
            "Less_online_redeemed_quantity":report.get("quantity"),
            "Less_online_redeemed_amount":report.get("amount")
        }
        data.append(less_online_reedemed_strcture)
    for subtotal in subtotals:
        less_online_reedemed_strcture = {
            "Less_online_redeemed_description":subtotal.get("description"),
            "Less_online_redeemed_price":subtotal.get("price"),
            "Less_online_redeemed_quantity":subtotal.get("quantity"),
            "Less_online_redeemed_amount":subtotal.get("amount")
        }
        data.append(less_online_reedemed_strcture)
    return data
    

def free_wash_issued(section):
    data=[]
    reports = section.get("reports")
    subtotals = section.get("subtotals")
    for report in reports:
        free_wash_issued_structure = {
            "Free_washes_issued_description":report.get("description"),
            "Free_washes_issued_price":report.get("price"),
            "Free_washes_issued_quantity":report.get("quantity"),
            "Free_washes_issued_amount":report.get("amount")
        }
        
        data.append(free_wash_issued_structure)
    for subtotal in subtotals:
        free_wash_issued_structure = {
            "Free_washes_issued_description":subtotal.get("description"),
            "Free_washes_issued_price":subtotal.get("price"),
            "Free_washes_issued_quantity":subtotal.get("quantity"),
            "Free_washes_issued_amount":subtotal.get("amount")
        }
        
        data.append(free_wash_issued_structure)
        
    return data

def less_paidouts(section):
    data=[]
    reports = section.get("reports")
    subtotals = section.get("subtotals")
    
    for report in reports:
        less_paidouts_structure = {
            "Less_paidouts_description":report.get("description"),
            "Less_paidouts_price":report.get("price"),
            "Less_paidouts_quantity":report.get("quantity"),
            "Less_paidouts_amount":report.get("amount")
        }
        data.append(less_paidouts_structure)
    for total in subtotals:
        less_paidouts_structure = {
            "Less_paidouts_description":total.get("description"),
            "Less_paidouts_price":total.get("price"),
            "Less_paidouts_quantity":total.get("quantity"),
            "Less_paidouts_amount":total.get("amount")
        }
        data.append(less_paidouts_structure)
    
    return data

def total_to_account_for(section):
    # data = []
    

    # total_to_account_for_structure = {
    #     "TOTAL_TO_ACCOUNT_FOR:":section.get("totalAmount")
    # }

    # data.append(total_to_account_for_structure)
    
    return section.get("totalAmount",0.0)

def deposits(section):
    data = []
    reports  = section.get("reports")
    subtotals  = section.get("subtotals")
    for report in reports:
        deposit_structure = {
            "Deposits_description":report.get("description"),
            "Deposits_price":report.get("price"),
            "Deposits_quantity":report.get("quantity"),
            "Deposits_amount":report.get("amount")
        }
        data.append(deposit_structure)
    
    for subtotal in subtotals:
        deposit_structure = {
            "Deposits_description":subtotal.get("description"),
            "Deposits_price":subtotal.get("price"),
            "Deposits_quantity":subtotal.get("quantity"),
            "Deposits_amount":subtotal.get("amount")
        }
        data.append(deposit_structure)
        
    return data

def total_xpt_cash(section):
    data =[]
    total_xpt_structure ={
        "TOTAL XPT CASH:":section.get("totalAmount")
    }
    data.append(total_xpt_structure)
    
    return data

def house_accounts(section):
    data = []
    reports = section.get('reports')
    subtotals  = section.get('subtotals')
    for report in reports:
        house_accounts_structure ={
            "House_accounts_description":report.get("description"),
            "House_accounts_price":report.get("price"),
            "House_accounts_quantity":report.get("quantity"),
            "House_accounts_amount":report.get("amount")
        }
        data.append(house_accounts_structure)
        
    for subtotal in subtotals:
        house_accounts_structure ={
            "House_accounts_description":subtotal.get("description"),
            "House_accounts_price":subtotal.get("price"),
            "House_accounts_quantity":subtotal.get("quantity"),
            "House_accounts_amount":subtotal.get("amount")
        }
        data.append(house_accounts_structure)
        
    return data
    
# def over_short(section):

def cash(section):
    data=[]
    cash_structure ={
        "CASH:":section.get("totalAmount")
    }
    data.append(cash_structure)

    return data

def xpt_acceptors(section):
    data = []
    xpt_acceptor_structure = {
        "XPT ACCEPTORS:":section.get("totalAmount")
    }
    
    data.append(xpt_acceptor_structure)
    
    return data
  
def xpt_dispensers(section):
    data = []
    xpt_dispenser_structure ={
        "XPT DISPENSERS:":section.get("totalAmount")
    }  
    
    data.append(xpt_dispenser_structure)
    
    return data

def total_function(section):
    data = []
    total_structure = {
        "TOTAL:":section.get("totalAmount")
    }
    data.append(total_structure)
    
    return data

def credit_card(section):
    data = []
    reports = section.get("reports")
    subtotals = section.get("subtotals")
    
    for report in reports:
        credit_card_structure = {
            "Credit_Card_description":report.get("description"),
            "Credit_Card_price":report.get("price"),
            "Credit_Card_quantity":report.get("quantity"),
            "Credit_Card_amount":report.get("amount")
        }
        data.append(credit_card_structure)
    
    for subtotal in subtotals:
        credit_card_structure = {
            "Credit_Card_description":subtotal.get("description"),
            "Credit_Card_price":subtotal.get("price"),
            "Credit_Card_quantity":subtotal.get("quantity"),
            "Credit_Card_amount":subtotal.get("amount")
        }
        data.append(credit_card_structure)
    return data

def other_tenders(section):
    data = []
    reports = section.get("reports")
    subtotals = section.get("subtotals")
    for report in reports:
        other_tender_structure = {
            "Other_tenders_description":report.get("description"),
            "Other_tenders_price":report.get("price"),
            "Other_tenders_quantity":report.get("quantity"),
            "Other_tenders_amount":report.get("amount")
        }
        data.append(other_tender_structure)
    for subtotal in subtotals:
        other_tender_structure = {
            "Other_tenders_description":subtotal.get("description"),
            "Other_tenders_price":subtotal.get("price"),
            "Other_tenders_quantity":subtotal.get("quantity"),
            "Other_tenders_amount":subtotal.get("amount")
        }
        data.append(other_tender_structure)
        
    return data

def xpt_balancing(section):
    data = []
    xpt_balancing_structure = {
        "XPT BALANCING:":section.get("totalAmount")
    }
    data.append(xpt_balancing_structure)
    
    return data


def report_balance(section):
    data = []
    report_balance_structure = {
        "REPORT BALANCE:":section.get("totalAmount")
    }  
    
    data.append(report_balance_structure)
    
    return data

def picture_mismatch(section):
    data = []
    picture_mismatch_structure = {
        "PICTURE MISMATCH:":section.get("totalCount")
    }
    data.append(picture_mismatch_structure)
    
    return data


def append_dict_to_excel(file_path, data, num_lines,add_headers=True):
    try:
        # Try to load an existing workbook
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
    else:
        # If the file exists, get the active sheet
        sheet = workbook.active

    # Append blank lines
    for _ in range(num_lines):
        sheet.append([])

    if add_headers:
        # Append the header row with bold keys
        bold_font = Font(bold=True)
        header_row = list(data.keys())
        sheet.append(header_row)
        for cell in sheet[sheet.max_row]:
            cell.font = bold_font

    # Append the data row
    data_row = list(data.values())
    sheet.append(data_row)

    # Save the workbook
    workbook.save(file_path)

def write_dictionary_to_xlshet(dictionary_data,file_name,is_first_dictionary=True):
    return ""
    for index,data in enumerate(dictionary_data):
                if index == 0:
                    column_gap = 0 if is_first_dictionary else 2
                    append_dict_to_excel(file_name,data,column_gap)
                else:
                    append_dict_to_excel(file_name,data,0,False)


def report_data_extractor(report_data):
    single_site_report = {}
    if report_data:
        gsviews =  report_data.get("gsviews")
        gsviews_0 = gsviews[0]

        sections = gsviews_0.get("sections")
        for section in sections:
            text        = section.get("text")
            
            
            if text=="WASH SALES-":
                wash_sales_ret = wash_sales(section)#,file_path
                print(f"Washsales:{wash_sales_ret}")
                single_site_report["car_count"] = wash_sales_ret #car count
                
                    
                
            elif text=="WASH PACKAGES-":
                wash_packages(section)
                    
            elif text=="WASH EXTRA SERVICES-":
                wash_extra_services(section)
                    
            elif text=="GROSS WASH SALES-":
                gross_wash_sales(section)
                
            
            elif text=="LESS FREE WASH RDMD-":
                less_free_wash_rdmd(section)
                    
            
            elif text=="LESS WASH DISCOUNTS-":
                less_wash_discounts(section)
                    
            elif text=="LESS LOYALTY DISC-":
                less_loyality_disc(section)
            
            elif text == "NET SITE SALES:":      
                net_site_sales_value = net_site_sales(section)
                single_site_report['net_sales']=net_site_sales_value
            
            elif text=="ARM PLANS SOLD-":
                arm_plans_sold_cnt= arm_plans_sold(section)
                single_site_report["arm_plans_sold_cnt"] = arm_plans_sold_cnt
                
                
            elif text=="ARM PLANS RECHARGED-":
                arm_reachrged_amt =arm_plans_recharged(section)
                print("arm plans rechanged:",arm_reachrged_amt)
                single_site_report["arm_plans_recharged_amt"]=arm_reachrged_amt
                
            elif text=="ARM PLANS REDEEMED-":
                arm_plans_reedemed_value= arm_planes_reedemed(section)
                single_site_report["arm_plans_reedemed_cnt"] = arm_plans_reedemed_value.get("arm_plans_reedemed_cnt")
                single_site_report["arm_plans_reedemed_amt"] =arm_plans_reedemed_value.get("arm_plans_reedemed_amt")
                    
            elif text=="ARM PLANS TERMINATED-":
                arm_plans_terminated(section)
            
            elif text=="PREPAIDS SOLD-":
                prepaid_sold(section)
            
            elif text=="LESS PREPAIDS REDEEMED-":
                less_prepaid_reedemed(section)
                
            elif text == "ONLINE SOLD-":
                online_sold(section)
                
            elif text == "LESS ONLINE REDEEMED-":
                less_online_reedemed(section)
                
            elif text=="FREE WASHES ISSUED-":
                free_wash_issued(section)
                
            elif text=="LESS PAIDOUTS:":
                less_paidouts(section)
                
            elif text=="TOTAL TO ACCOUNT FOR:":
                total_revenue_val = total_to_account_for(section)
                single_site_report['total_revenue'] = total_revenue_val
            
            elif text=="DEPOSITS-":
                deposits(section)
                
            elif text=="TOTAL XPT CASH:":
                total_xpt_cash(section)
                
            elif text=="HOUSE ACCOUNTS-":
                house_accounts(section)
                
            # elif text =="OVER / SHORT (-)":
            #     over_short_lst = over_short(section)
            
            elif text=="CASH:":
                cash(section)
                
            elif text=="XPT ACCEPTORS:":
                xpt_acceptors(section)
                
            elif text =="XPT DISPENSERS:":
                xpt_dispensers(section)
                
            elif text =="TOTAL:":
                total_function(section)
                
            elif text=="CREDIT CARD:":
                credit_card(section)
                
                
            elif text=="OTHER TENDERS:":
                other_tenders(section)
                
            elif text=="XPT BALANCING:":
                xpt_balancing(section)
                
            elif text=="REPORT BALANCE:":
                report_balance(section)
                
            elif text=="PICTURE MISMATCH:":
                picture_mismatch(section)
            
    return single_site_report

def get_week_dates():
    # Get the current date
    today = datetime.today()
    
    # Find the current week's Monday date
    current_week_monday = today - timedelta(days=today.weekday())
    
    # Find the current week's Friday, Saturday, and Sunday dates
    current_week_friday = current_week_monday + timedelta(days=4)
    current_week_saturday = current_week_monday + timedelta(days=5)
    current_week_sunday = current_week_monday + timedelta(days=6)
    
    # Format the dates in yyyy-mm-dd format
    monday_date_str = current_week_monday.strftime("%Y-%m-%d")
    friday_date_str = current_week_friday.strftime("%Y-%m-%d")
    saturday_date_str = current_week_saturday.strftime("%Y-%m-%d")
    sunday_date_str = current_week_sunday.strftime("%Y-%m-%d")
    
    return monday_date_str, friday_date_str, saturday_date_str, sunday_date_str



def generate_weekly_report(path,monday_date_str,friday_date_str,saturday_date_str, sunday_date_str):
    site_watch_report={}
    
    is_location_code_success=False
    success_location_code=None,None
    location_codes_a = ['SPKLUS-001', 'SPKLUS-002', 'SPKLUS-003', 'SPKLUS-004', 'SPKLUS-005', 'SPKLUS-006', 'SPKLUS-007', 'SPKLUS-008', 'SPKLUS-009']
    location_codes_b = ['SPKLUS-012', 'SPKLUS-013', 'SPKLUS-014', 'SPKLUS-015']
    for index,site in sites_df.iterrows():
        # all_dictionaries_lst = []
        site_dict = site.to_dict()
        slno=site_dict.get("slno")
        locationCode = site_dict.get("Organization")
        
        
        while True:
                      
            try:
            
                combined_data = {}
                
                
                cookiesfile_name = f"{(site_dict.get('Organization')).strip().replace('-','_')}.pkl"
                # print(cookiesfile_name)
                print(site_dict)
                

                cookies_file = os.path.join(cookies_path,cookiesfile_name)
            
                
                client = sitewatchClient(cookies_file=cookies_file)
                employCode = site_dict.get("employee")
                password = site_dict.get("password")
                
                
                if is_location_code_success:
                    locationCode_old,slno_old =success_location_code
                    location_a_range =range(1,10)
                    location_b_range =range(10,14)
                    if slno_old in location_a_range and slno in location_a_range: 
                        locationCode=locationCode_old
                    
                    elif slno_old in location_b_range and slno in location_b_range:
                        locationCode=locationCode_old
                        
                print(f"\n location code used :{locationCode}")
                client_name = site_dict.get("client_name2")
                remember = 1
                # file_path=f"sitewatch_{client_name.strip().replace(' ','_')}_{monday_date_str}_{sunday_date_str}.xlsx"
                
                # file_path = os.path.join(path,file_path)
                session_chek = client.check_session_auth(timeout=60)

                if not session_chek:
                    token = client.login(employeeCode=employCode,password=password,locationCode=locationCode,remember=1)
                    print(token)
                
                session_chek = client.check_session_auth(timeout=60)
                if session_chek:
                    reportOn=site_dict.get("site")
                    id=site_dict.get("id")
                    idname=site_dict.get("id_name")
                    request_id1 = client.get_general_sales_report_request_id(reportOn,id,idname,monday_date_str,friday_date_str)
                    request_id1_1 =client.get_activity_by_date_proft_request_id(reportOn,monday_date_str,friday_date_str) #for labour hours 

                    if request_id1 and request_id1_1:
                        report_data = client.get_report(reportOn,request_id1)
                        # print(f"report data: {report_data}")
                        extracted_data1= report_data_extractor(report_data)
                        car_count_monday_to_friday = extracted_data1.get("car_count",0)
                        arm_plans_reedemed_monday_to_friday_cnt = extracted_data1.get("arm_plans_reedemed_cnt",0)
                        arm_plans_reedemed_monday_to_friday_amt = abs(extracted_data1.get("arm_plans_reedemed_amt",0))#need abs because will nee to mek it postive 
                        arm_plans_recharged_amt_monday_to_friday_amt =extracted_data1.get("arm_plans_recharged_amt",0)
                        
                        retail_car_count_monday_to_friday=(car_count_monday_to_friday - arm_plans_reedemed_monday_to_friday_cnt)
                        
                        net_sales_amt= extracted_data1.get("net_sales",0.0)
                        # retail_revenue__monday_to_friday = round((net_sales_amt - arm_plans_reedemed_monday_to_friday_amt),2)
                        total_revenue_val = round(extracted_data1.get("total_revenue",0.0),2)
                        
                        if client_name=="Sudz - Beverly":
                            retail_revenue__monday_to_friday = round((total_revenue_val - arm_plans_recharged_amt_monday_to_friday_amt),2)
                        else:
                            retail_revenue__monday_to_friday = round((net_sales_amt - arm_plans_reedemed_monday_to_friday_amt),2)
                            
                            
                        arm_plans_sold_cnt1 = extracted_data1.get("arm_plans_sold_cnt")
                        labour_hours_monday_to_friday=client.get_labour_hours(reportOn,request_id1_1)
                        cars_per_labour_hour_monday_to_friday = round((car_count_monday_to_friday/labour_hours_monday_to_friday),2) if labour_hours_monday_to_friday !=0 else "" 
                        
                        mon_fri_data = {
                            "car_count_monday_to_friday":car_count_monday_to_friday,
                            "arm_plans_reedemed_monday_to_friday_cnt":arm_plans_reedemed_monday_to_friday_cnt,
                            "retail_car_count_monday_to_friday":retail_car_count_monday_to_friday,
                            "retail_revenue_monday_to_friday":retail_revenue__monday_to_friday,
                            "total_revenue_monday_to_friday":total_revenue_val,
                            "labour_hours_monday_to_friday":labour_hours_monday_to_friday,
                            "cars_per_labour_hour_monday_to_friday":cars_per_labour_hour_monday_to_friday
                        }
                        combined_data.update(mon_fri_data)
                    
                    request_id2 = client.get_general_sales_report_request_id(reportOn,id,idname,saturday_date_str,sunday_date_str)
                    request_id2_2 =client.get_activity_by_date_proft_request_id(reportOn,saturday_date_str,sunday_date_str) #for labour hours
                    
                    if request_id2 and request_id2_2:
                        report_data = client.get_report(reportOn,request_id2)
                        # print(f"data2:{report_data}")
                        extracted_data2= report_data_extractor(report_data)
                        car_count_saturday_sunday = extracted_data2.get("car_count",0)
                        arm_plans_reedemed_saturday_sunday_cnt = extracted_data2.get("arm_plans_reedemed_cnt",0)
                        arm_plans_reedemed_saturday_sunday_amt = abs(extracted_data2.get("arm_plans_reedemed_amt",0))
                        arm_plans_recharged_amt_saturday_sunday_amt =extracted_data2.get("arm_plans_recharged_amt",0)
                        
                        retail_car_count_saturday_sunday =(car_count_saturday_sunday- arm_plans_reedemed_saturday_sunday_cnt)
                        
                        net_sales_amt2= extracted_data2.get("net_sales",0.0)
                        
                        total_revenue_val2 = round(extracted_data2.get("total_revenue",0.0),2)
                        
                        #print(f"retail rev : total rev {total_revenue_val2}-{arm_plans_recharged_amt_saturday_sunday_amt} ={total_revenue_val2-arm_plans_recharged_amt_saturday_sunday_amt}")
                        
                        if client_name=="Sudz - Beverly":
                            retail_revenue__saturday_sunday = round((total_revenue_val2 - arm_plans_recharged_amt_saturday_sunday_amt),2)
                        else:
                            retail_revenue__saturday_sunday = round((net_sales_amt2- arm_plans_reedemed_saturday_sunday_amt),2)
                            
                        arm_plans_sold_cnt2 = extracted_data2.get("arm_plans_sold_cnt")
                        
                        labour_hours_saturday_sunday=client.get_labour_hours(reportOn,request_id2_2)
                        cars_per_labour_hour_saturday_sunday = round((car_count_saturday_sunday/labour_hours_saturday_sunday),2) if labour_hours_saturday_sunday !=0 else ""
                        
                        sat_sun_data = {
                            "car_count_saturday_sunday":car_count_saturday_sunday,
                            "arm_plans_reedemed_saturday_sunday":arm_plans_reedemed_saturday_sunday_cnt,
                            "retail_car_count_saturday_sunday":retail_car_count_saturday_sunday,
                            "retail_revenue_saturday_sunday":retail_revenue__saturday_sunday,
                            "total_revenue_saturday_sunday":total_revenue_val2,
                            "labour_hours_saturday_sunday":labour_hours_saturday_sunday,
                            "cars_per_labour_hour_saturday_sunday":cars_per_labour_hour_saturday_sunday
                        }
                        
                        combined_data.update(sat_sun_data)
                        
                        arm_plans_sold_total_cnt = sum([arm_plans_sold_cnt1,arm_plans_sold_cnt2])
                        # total_arm_planmembers_cnt = sum([arm_plans_sold_cnt1,arm_plans_sold_cnt2,
                        #                                                 arm_plans_reedemed_monday_to_friday_cnt,
                        #                                                 arm_plans_reedemed_saturday_sunday_cnt])
                        
                        total_retail_car_cnt = sum([retail_car_count_monday_to_friday,retail_car_count_saturday_sunday])
                        conversion_rate  = round((arm_plans_sold_total_cnt/total_retail_car_cnt)*100,2) if total_retail_car_cnt !=0 else ""
                        #combines values update 
                        #combined_data["total_cars"] = sum([car_count_monday_to_friday,car_count_saturday_sunday])
                        total_members_req_id = client.get_plan_analysis_request_id(sunday_date_str,reportOn)
                        
                        total_arm_planmembers_cnt = client.get_total_plan_members(total_members_req_id,reportOn)
                        
                        #Past 4 weeks logic 
                        past_week_day1, past_week_day2= generate_past_4_weeks_days(monday_date_str)
                        
                        request_id3 = client.get_general_sales_report_request_id(reportOn,id,idname,past_week_day1, past_week_day2)
                        report_data3 = client.get_report(reportOn,request_id3)
                        extracted_data3= report_data_extractor(report_data3)
                        past_4_week_cnt = extracted_data3.get("car_count",0)
                        past_4_week_arm_plans_sold =extracted_data3.get("arm_plans_sold_cnt")
                        past_4_weeks_arm_plans_reedemed_cnt = extracted_data3.get("arm_plans_reedemed_cnt",0)
                        past_4_weeks_retail_car_count = past_4_week_cnt - past_4_weeks_arm_plans_reedemed_cnt
                        past_4_weeks_total_revenue = extracted_data3.get("total_revenue")
                        
                        print(f"past week cnt : {past_4_week_cnt}")
                        past_4_weeks_conversion_rate = (past_4_week_arm_plans_sold/past_4_weeks_retail_car_count)*100
                        
                        #Past 4 weeks splieed days mon-day st -sun
                        full_weeks_lst = generate_past_4_week_days_full(monday_date_str)
                        
                        combined_data["past_4_week_car_cnt_mon_fri"]=0
                        combined_data["past_4_week_car_cnt_sat_sun"]=0

                        combined_data["past_4_week_labour_hours_mon_fri"]=0                     
                        combined_data["past_4_week_labour_hours_sat_sun"]=0

                        combined_data["past_4_week_retail_car_count_mon_fri"]=0
                        combined_data["past_4_week_retail_car_count_sat_sun"]=0

                        combined_data['past_4_week_retail_revenue_mon_fri'] = 0
                        combined_data['past_4_week_retail_revenue_sat_sun'] = 0

                        combined_data['past_4_week_total_revenue_mon_fri'] = 0
                        combined_data['past_4_week_total_revenue_sat_sun'] = 0

                        cnt=0
                        for single_week in full_weeks_lst:
                            mon = single_week[0]
                            fri = single_week[1]
                            sat =single_week[2]
                            sun = single_week[3]

                            request_id4 = client.get_general_sales_report_request_id(reportOn,id,idname,mon, fri)
                            report_data4 = client.get_report(reportOn,request_id4)
                            extracted_data4 = report_data_extractor(report_data4)
                            
                            request_id4_2 =client.get_activity_by_date_proft_request_id(reportOn,mon, fri) #for labour hours
                            labour_hours_mon_fri=client.get_labour_hours(reportOn,request_id4_2)
                            
                            request_id5_2 =client.get_activity_by_date_proft_request_id(reportOn,sat, sun) #for labour hours
                            labour_hours_sat_sun=client.get_labour_hours(reportOn,request_id5_2)
                            
                            request_id5 = client.get_general_sales_report_request_id(reportOn,id,idname,sat, sun)
                            report_data5 = client.get_report(reportOn,request_id5)
                            extracted_data5 = report_data_extractor(report_data5)
                            
                            # Mon-Friday
                            car_cnt_mon_fri = extracted_data4.get("car_count",0)
                            print(f"car cnt 1 ; {car_cnt_mon_fri}")
                            arm_plans_reedemed_cnt_mon_fri = extracted_data4.get("arm_plans_reedemed_cnt", 0)
                            net_sales_amt_mon_fri = extracted_data4.get("net_sales",0.0)
                            total_revenue_val_mon_fri = round(extracted_data4.get("total_revenue",0.0),2)
                            retail_car_count_mon_fri = (car_cnt_mon_fri-arm_plans_reedemed_cnt_mon_fri)
                            arm_plans_reedemed_mon_fri_amt = abs(extracted_data4.get("arm_plans_reedemed_amt",0))
                            arm_plans_recharged_amt_mon_fri_amt = extracted_data4.get("arm_plans_recharged_amt",0)
                            total_revenue_mon_fri = round(extracted_data4.get("total_revenue",0.0),2)

                            # Sat-Sun
                            car_cnt_sat_sun    = extracted_data5.get("car_count",0)
                            print(f"car count2 : {car_cnt_sat_sun}")
                            arm_plans_reedemed_cnt_sat_sun = extracted_data5.get("arm_plans_reedemed_cnt", 0)
                            net_sales_amt_sat_sun = extracted_data5.get("net_sales",0.0)
                            total_revenue_val_sat_sun = round(extracted_data5.get("total_revenue",0.0),2)
                            retail_car_count_sat_sun = (car_cnt_sat_sun-arm_plans_reedemed_cnt_sat_sun)
                            arm_plans_reedemed_sat_sun_amt = abs(extracted_data5.get("arm_plans_reedemed_amt",0))
                            arm_plans_recharged_amt_sat_sun_amt = extracted_data5.get("arm_plans_recharged_amt",0)
                            total_revenue_sat_sun = round(extracted_data5.get("total_revenue",0.0),2)

                            if client_name=="Sudz - Beverly":
                                past_4_week_retail_revenue_mon_fri = round((total_revenue_val_mon_fri - arm_plans_recharged_amt_mon_fri_amt),2) #arm_plans_recharged_amt_mon_fri_amt
                                past_4_week_retail_revenue_sat_sun = round((total_revenue_val_sat_sun - arm_plans_recharged_amt_sat_sun_amt),2) # arm_plans_recharged_amt_sat_sun_amt
                                combined_data['past_4_week_retail_revenue_mon_fri'] = combined_data.get("past_4_week_retail_revenue_mon_fri",0)+ past_4_week_retail_revenue_mon_fri
                                combined_data['past_4_week_retail_revenue_sat_sun'] = combined_data.get("past_4_week_retail_revenue_sat_sun",0) + past_4_week_retail_revenue_sat_sun
                                # print(f"retail revenue mon fri : {combined_data['past_4_week_retail_revenue_mon_fri']}")
                                # print(f" retail revenue sat sun : {combined_data['past_4_week_retail_revenue_sat_sun']}")
                                combined_data[f"past_4_week_retail_revenue_mon_fri_week_{cnt+1}"] = round((total_revenue_val_mon_fri - arm_plans_recharged_amt_mon_fri_amt),2)
                                combined_data[f"past_4_week_retail_revenue_sat_sun_week_{cnt+1}"] = round((total_revenue_val_sat_sun - arm_plans_recharged_amt_sat_sun_amt),2)
                            else:
                                past_4_week_retail_revenue_mon_fri  = round((net_sales_amt_mon_fri - arm_plans_reedemed_mon_fri_amt),2) #arm_plans_reedemed_mon_fri_amt
                                past_4_week_retail_revenue_sat_sun  =  round((net_sales_amt_sat_sun - arm_plans_reedemed_sat_sun_amt),2) # arm_plans_reedemed_sat_sun_amt
                                combined_data['past_4_week_retail_revenue_mon_fri'] = combined_data.get("past_4_week_retail_revenue_mon_fri",0) + past_4_week_retail_revenue_mon_fri
                                combined_data['past_4_week_retail_revenue_sat_sun'] = combined_data.get("past_4_week_retail_revenue_sat_sun",0) + past_4_week_retail_revenue_sat_sun

                                combined_data[f"past_4_week_retail_revenue_mon_fri_week_{cnt+1}"]  = round((net_sales_amt_mon_fri -arm_plans_reedemed_mon_fri_amt),2)
                                combined_data[f"past_4_week_retail_revenue_sat_sun_week_{cnt+1}"]  = round((net_sales_amt_sat_sun - arm_plans_reedemed_sat_sun_amt),2)
                            retail_car_count_sat_sun = (car_cnt_sat_sun-arm_plans_reedemed_cnt_sat_sun)

                            combined_data["past_4_week_car_cnt_mon_fri"] = combined_data.get("past_4_week_car_cnt_mon_fri",0)+car_cnt_mon_fri
                            combined_data["past_4_week_car_cnt_sat_sun"] = combined_data.get("past_4_week_car_cnt_sat_sun",0)+ car_cnt_sat_sun

                            combined_data["past_4_week_labour_hours_mon_fri"] = combined_data.get("past_4_week_labour_hours_mon_fri",0) + labour_hours_mon_fri
                            combined_data["past_4_week_labour_hours_sat_sun"] = combined_data.get("past_4_week_labour_hours_sat_sun",0) + labour_hours_sat_sun

                            combined_data["past_4_week_retail_car_count_mon_fri"] = combined_data.get("past_4_week_retail_car_count_mon_fri", 0)+retail_car_count_mon_fri
                            combined_data["past_4_week_retail_car_count_sat_sun"] = combined_data.get("past_4_week_retail_car_count_sat_sun", 0)+retail_car_count_sat_sun
                            combined_data[f"past_4_week_retail_car_count_mon_fri_week_{cnt+1}"]  = retail_car_count_mon_fri
                            combined_data[f"past_4_week_retail_car_count_sat_sun_week_{cnt+1}"]  = retail_car_count_sat_sun

                            combined_data['past_4_week_total_revenue_mon_fri'] += total_revenue_mon_fri
                            combined_data['past_4_week_total_revenue_sat_sun'] += total_revenue_sat_sun
                            
                            cnt+=1    
                            
                            
                            
                            
                            
                        
                        combined_data["total_revenue"] = sum([total_revenue_val,total_revenue_val2])
                        combined_data["arm_plans_sold_cnt"] = arm_plans_sold_total_cnt
                        combined_data["total_arm_planmembers_cnt"] = total_arm_planmembers_cnt
                        combined_data["conversion_rate"]= conversion_rate
                        
                        combined_data["past_4_week_cnt"]=past_4_week_cnt #total car count past 4weeks 
                        combined_data["past_4_week_conversion_rate"] = past_4_weeks_conversion_rate
                        combined_data["past_4_weeks_total_revenue"] = past_4_weeks_total_revenue
                        combined_data["past_4_weeks_arm_plans_sold_cnt"]= past_4_week_arm_plans_sold
                        combined_data["past_4_weeks_retail_car_count"] = past_4_weeks_retail_car_count

                    print(f"combined data:{combined_data}")
                    site_watch_report[client_name]=combined_data
                    if combined_data:
                        is_location_code_success=True
                        success_location_code = locationCode,slno
                    if is_location_code_success:
                        break #break loop
                
                # if not is_location_code_success:#session not success
                else:
                    location_a_range =range(1,10)
                    location_b_range =range(10,14)
                    print("\n switching lcoation code")
                    if slno in location_a_range: 
                        locationCode=random.choice(location_codes_a)
                        print(f"\n retrying with new location code {locationCode} for {slno}")
                        success_location_code = locationCode,slno
                    
                    elif  slno in location_b_range:
                        locationCode=random.choice(location_codes_b)
                        success_location_code = locationCode,slno
                        print(f"\n retrying with new location code {locationCode} for {slno}")
                        
                    print("sleep for 5 secound before next retry")
                    time.sleep(15)
                    
                    
            except Exception as e:
                print(f"Excetion for this loctaion {client_name} {e} {traceback.print_exc() }")                 
    
    return site_watch_report

#xl maps fucntions 

def do_sum(xl_map,start_index,range):
    "This will do row based some "
    total=0
    for i in range:
        val = xl_map[start_index][i+1]
        if isinstance(val,float) or isinstance(val,int):
            total+=val
    
    return total

def do_sum_location(xl_map,location:list):
    "This will take array of row ,col"
    total =0
    for row,col in location:
        val = xl_map[row][col]
        if isinstance(val,float) or isinstance(val,int):
            total+=val
    return total

#Xl maps fucntions ends

def Average_retail_visit__GA_SC_fucntion(retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC,
                                         retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC):    
    
    result = 0
    try:
        Average_retail_visit__GA_SC = sum([retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC])/sum(
        [retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC ]
    )
        result = Average_retail_visit__GA_SC
    except Exception as e:
        print(f"Exception in Average_retail_visit__GA_SC_fucntion() {e}")
        

    return result 

def Average_memeber_visit_GA_SC_function(Total_revenue_GA_SC,
        retail_revenue_monday_to_friday_GA_SC,retail_revenue_saturday_to_sunday_GA_SC,
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC,
        total_cars_GA_SC
        ):
    result = 0
    
    try:
        
        Average_memeber_visit_GA_SC = (Total_revenue_GA_SC - sum([retail_revenue_monday_to_friday_GA_SC , retail_revenue_saturday_to_sunday_GA_SC]))/(total_cars_GA_SC - sum([
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC]))
        
        result = Average_memeber_visit_GA_SC
        
    except Exception as e:
        print(f"Exception in Average_memeber_visit_GA_SC_function() {e}")
    
    return result

def cost_per_labour_hour_monday_to_friday_GA_SC_fucntion(car_count_monday_to_friday_GA_SC,staff_hours_monday_to_friday_GA_SC):
    result = 0
    try:
        cost_per_labour_hour_monday_to_friday_GA_SC = car_count_monday_to_friday_GA_SC/staff_hours_monday_to_friday_GA_SC
        result =cost_per_labour_hour_monday_to_friday_GA_SC
    except Exception as e:
        print(f"Exception in cost_per_labour_hour_monday_to_friday_GA_SC() {e}")

    return result


def cost_per_labour_hour_saturday_to_sunday_GA_SC_function(car_count_saturday_to_sunday_GA_SC,staff_hours_saturday_to_sunday_GA_SC):
    result = 0
    
    try:
        cost_per_labour_hour_saturday_to_sunday_GA_SC = car_count_saturday_to_sunday_GA_SC/staff_hours_saturday_to_sunday_GA_SC
        result = cost_per_labour_hour_saturday_to_sunday_GA_SC
    except Exception as e:
        print(f"Exception in staff_hours_saturday_to_sunday_GA_SC_fucntion() {e}")
        
    return result

def Total_cars_per_man_hour_GA_SC_function(total_cars_GA_SC, staff_hours_monday_to_friday_GA_SC,staff_hours_saturday_to_sunday_GA_SC):
    result =0 
    try:
            Total_cars_per_man_hour_GA_SC = total_cars_GA_SC/sum([
            staff_hours_monday_to_friday_GA_SC,staff_hours_saturday_to_sunday_GA_SC])
            result = Total_cars_per_man_hour_GA_SC
    
    except Exception as e:
        print(f"Exception in  Total_cars_per_man_hour_GA_SC_function() {e}")
    
    return result
 
def Conversion_rate_GA_SC_function(Total_club_plans_sold_GA_SC,retail_car_count_monday_to_friday_GA_SC,
                                   retail_car_count_saturday_to_sunday_GA_SC):
    result = 0 
    try:
        Conversion_rate_GA_SC = Total_club_plans_sold_GA_SC/sum([
        retail_car_count_monday_to_friday_GA_SC,retail_car_count_saturday_to_sunday_GA_SC])     
        
        result =  Conversion_rate_GA_SC
    except Exception as e:
        print(f"Exceptionin Conversion_rate_GA_SC_function() {e}")
    
    return result

def Conversion_rate_Total_function(Total_club_plans_sold_Total,
                                   retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total):
    result = 0
    
    try:
        
        Conversion_rate_Total = Total_club_plans_sold_Total/sum(
        [retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total])  
        
        result = Conversion_rate_Total
    except Exception as e:
        print(f"Exception in Conversion_rate_Total_function()  {e}") 
    
    return result

def Conversion_rate_ILL_function(Total_club_plans_sold_ILL,
                                 retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Conversion_rate_ILL = Total_club_plans_sold_ILL/sum([retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL])         
        result = Conversion_rate_ILL
    except Exception as e:
        print(f"Exception in Conversion_rate_ILL_function() {e}")
        
    return result

def Total_cars_per_man_hour_total_function(Total_cars_Total,
                                           staff_hours_monday_to_friday_Total,
                                           staff_hours_saturday_to_sunday_Total):
    result = 0
    
    try:
        Total_cars_per_man_hour_total = Total_cars_Total/sum(
        [staff_hours_monday_to_friday_Total,staff_hours_saturday_to_sunday_Total])
        
        result = Total_cars_per_man_hour_total
    
    except Exception as e:
        print(f"Exception in Total_cars_per_man_hour_total_function() {e}")
        
    return result


def Total_cars_per_man_hour_ILL_function(total_cars_in_ILL,
                                         staff_hours_monday_to_friday_ILL,
                                         staff_hours_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Total_cars_per_man_hour_ILL = total_cars_in_ILL/(sum([
        staff_hours_monday_to_friday_ILL+staff_hours_saturday_to_sunday_ILL]))

        result = Total_cars_per_man_hour_ILL
    
    except Exception as e:
        print(f"Exception Total_cars_per_man_hour_ILL_function() {e}")
    
    return result


def cost_per_labour_hour_saturday_to_sunday_ILL_function(car_count_saturday_to_sunday_ILL,staff_hours_saturday_to_sunday_ILL):
    result = 0
    
    try:
        cost_per_labour_hour_saturday_to_sunday_ILL = car_count_saturday_to_sunday_ILL/staff_hours_saturday_to_sunday_ILL
        
        result = cost_per_labour_hour_saturday_to_sunday_ILL
    
    except Exception as e:
        print(f"Exception cost_per_labour_hour_saturday_to_sunday_ILL_function() {e}")
        
    return result


def cost_per_labour_hour_monday_to_friday_Total_function(car_count_monday_to_friday_Total,staff_hours_monday_to_friday_Total):
    result = 0
    
    try:
        cost_per_labour_hour_monday_to_friday_Total = car_count_monday_to_friday_Total/staff_hours_monday_to_friday_Total
        result = cost_per_labour_hour_monday_to_friday_Total
    
    except Exception as e:
        print(f"Exception cost_per_labour_hour_monday_to_friday_Total_function() {e}")
    
    return result

def cost_per_labour_hour_monday_to_friday_ILL_function(car_count_monday_to_friday_ILL,staff_hours_monday_to_friday_ILL):
    result = 0
    
    try:
        cost_per_labour_hour_monday_to_friday_ILL = car_count_monday_to_friday_ILL/staff_hours_monday_to_friday_ILL
        result =  cost_per_labour_hour_monday_to_friday_ILL
    
    except Exception as e:
        print(f"Exception cost_per_labour_hour_monday_to_friday_ILL_function() {e}")
        
    return result

def Average_memeber_visit_Total_function(Total_revenue_Total,retail_revenue_monday_to_friday_Total
                        ,retail_revenue_saturday_to_sunday_Total,Total_cars_Total,retail_car_count_monday_to_friday_Total,
                        retail_car_count_saturday_to_sunday_Total):
    result = 0
    
    try:
        Average_memeber_visit_Total = (Total_revenue_Total -sum([retail_revenue_monday_to_friday_Total,retail_revenue_saturday_to_sunday_Total]))/(Total_cars_Total - sum([
        retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total]))
        result =  Average_memeber_visit_Total
    
    except Exception as e:
        print(f"Exception Average_memeber_visit_Total_function() {e}")  
    
    return result    

def Average_memeber_visit_ILL_function(Total_revenue_ILL,retail_revenue_monday_to_friday_ILL,
                                       retail_revenue_saturday_to_sunday_ILL,total_cars_in_ILL,
                                       retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Average_memeber_visit_ILL = (Total_revenue_ILL - sum([retail_revenue_monday_to_friday_ILL,retail_revenue_saturday_to_sunday_ILL]))/(total_cars_in_ILL - sum([
        retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL]))
        
        result = Average_memeber_visit_ILL
    
    except Exception as e:
        print(f"Exception Average_memeber_visit_ILL_function() {e}")
    
    return result

def Average_retail_visit_Total_function(retail_revenue_monday_to_friday_Total,
                                        retail_revenue_saturday_to_sunday_Total,
                                        retail_car_count_monday_to_friday_Total,
                                        retail_car_count_saturday_to_sunday_Total):
    result = 0
    
    try:
        Average_retail_visit_Total = sum([retail_revenue_monday_to_friday_Total,retail_revenue_saturday_to_sunday_Total])/sum([
        retail_car_count_monday_to_friday_Total,retail_car_count_saturday_to_sunday_Total])
        
        result = Average_retail_visit_Total
    
    except Exception as e:
        print(f"Exception Average_retail_visit_Total_function() {e}")
        
    return result
            

def Average_retail_visit_IL_function(retail_revenue_monday_to_friday_ILL,
                                     retail_revenue_saturday_to_sunday_ILL,
                                     retail_car_count_monday_to_friday_ILL,
                                     retail_car_count_saturday_to_sunday_ILL):
    result = 0
    
    try:
        Average_retail_visit_IL = sum([retail_revenue_monday_to_friday_ILL,retail_revenue_saturday_to_sunday_ILL])/sum(
        [retail_car_count_monday_to_friday_ILL,retail_car_count_saturday_to_sunday_ILL])

        result = Average_retail_visit_IL
    
    except Exception as e:
        print(f"Exception Average_retail_visit_IL_function() {e} {traceback.print_exc() }")
    
    return result






if __name__=="__main__":
    # import pandas as pd
    # monday_date_str, sunday_date_str = get_week_dates()
    # print(monday_date_str,sunday_date_str)
    monday_date_str="2024-06-10"
    friday_date_str = "2024-06-14"
    saturday_date_str = "2024-06-15"
    sunday_date_str="2024-06-16"  #YMD
    
    monday_date_str="2024-07-22"
    friday_date_str = "2024-07-26"
    saturday_date_str = "2024-07-27"
    sunday_date_str="2024-07-28"  #Y-M-D
    
    report = generate_weekly_report("",monday_date_str,friday_date_str,saturday_date_str, sunday_date_str)
    print("\n"*6)
    print(report)
    with open("sitewatch_report_full.json","w") as f:
        json.dump(report,f,indent=4)
    # with open("SPKLUS-002.json",'r') as f:
    #     data = json.load(f)
        
    # df = pd.read_json("SPKLUS-002.json")
    
    # df.to_excel("sitewatch.xlsx")
    
    # with open("sitewatch_report_old.json",'r') as f:
    #     data=json.load(f)
    # data = report
    #prepare_xlmap(data)
    
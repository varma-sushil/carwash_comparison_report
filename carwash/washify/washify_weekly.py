from washify import washifyClient
import os 
import json
import sys
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import traceback

from washify import generate_past_4_weeks_days
from washify import generate_past_4_week_days_full
import logging

# Add the path to the parent directory of "washify" to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'washify')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'sitewash')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'hamilton')))



# Add the carwash directory to the sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__))))

username = 'Cameron'
password = 'Password1'
companyName = 'cleangetawayexpress'
userType = 'CWA'

file_path="washify_test.xlsx"


def get_week_dates():
    # Get the current date
    today = datetime.today()
    
    # Find the current week's Monday date
    current_week_monday = today - timedelta(days=today.weekday())
    
    # Find the current week's Sunday date
    current_week_sunday = current_week_monday + timedelta(days=6)
    
    # Find the current week's Friday date
    current_week_friday = current_week_monday + timedelta(days=4)
    
    # Find the current week's Saturday date
    current_week_saturday = current_week_monday + timedelta(days=5)
    
    # Format the dates in mm/dd/yyyy format
    monday_date_str = current_week_monday.strftime("%m/%d/%Y")
    friday_date_str = current_week_friday.strftime("%m/%d/%Y")
    saturday_date_str = current_week_saturday.strftime("%m/%d/%Y")
    sunday_date_str = current_week_sunday.strftime("%m/%d/%Y")
    
    return monday_date_str, friday_date_str, saturday_date_str, sunday_date_str


def append_dict_to_excel(file_path, data, num_lines,add_headers=True):
    try:
        # Try to load an existing workbook
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
    else:
        # If the file exists, get the active sheet
        sheet = workbook.active

    # Append blank lines
    for _ in range(num_lines):
        sheet.append([])

    if add_headers:
        # Append the header row with bold keys
        bold_font = Font(bold=True)
        header_row = list(data.keys())
        sheet.append(header_row)
        for cell in sheet[sheet.max_row]:
            cell.font = bold_font

    # Append the data row
    data_row = list(data.values())
    sheet.append(data_row)

    # Save the workbook
    workbook.save(file_path)


def write_dictionary_to_xlshet(dictionary_data,file_name,is_first_dictionary=True):
    for index,data in enumerate(dictionary_data):
                if index == 0:
                    column_gap = 0 if is_first_dictionary else 2
                    append_dict_to_excel(file_name,data,column_gap)
                else:
                    append_dict_to_excel(file_name,data,0,False)
                    
def write_place_report(file_path,client,client_locations_number_codes,monday,sunday):
        wash_packages_response = client.GetRevenuReportFinancialWashPackage(client_locations_number_codes, monday,sunday)
        wash_packages_data = client.GetRevenuReportFinancialWashPackage_formatter(wash_packages_response)  #first table 
        # print(wash_packages_response)
        write_dictionary_to_xlshet(wash_packages_data,file_path) #first dictionary
        
        wash_package_discount_response = client.GetRevenuReportFinancialWashDiscounts(client_locations_number_codes,monday,sunday)
        washpack_discount_data = client.GetRevenuReportFinancialWashDiscounts_formatter(wash_package_discount_response)
        
        write_dictionary_to_xlshet(washpack_discount_data,file_path,False)
        
        wash_extra_response = client.GetRevenuReportFinancialPackagesDiscount(client_locations_number_codes,monday,sunday)
        wash_extra_data = client.GetRevenuReportFinancialPackagesDiscount_formatter(wash_extra_response)
        
        write_dictionary_to_xlshet(wash_extra_data,file_path,False)
        
        unlimited_sales_response = client.GetRevenuReportFinancialUnlimitedSales(client_locations_number_codes,monday,sunday)
        unlimited_sales_data  = client.GetRevenuReportFinancialUnlimitedSales_formatter(unlimited_sales_response) 
        
        write_dictionary_to_xlshet(unlimited_sales_data,file_path,False)
        
        giftcard_sales_response = client.GetRevenuReportFinancialGiftcardsale(client_locations_number_codes,monday,sunday)
        giftcards_sales_data = client.GetRevenuReportFinancialGiftcardsale_formatter(giftcard_sales_response) 
        
        write_dictionary_to_xlshet(giftcards_sales_data,file_path,False)

        discount_discount_response = client.GetRevenuReportFinancialWashDiscounts(client_locations_number_codes,monday,sunday)
        discount_discount_data = client.GetRevenuReportFinancialWashDiscounts_formatter2(discount_discount_response) 
        
        write_dictionary_to_xlshet(discount_discount_data,file_path,False)       
        
        giftcard_reedemed_response = client.GetRevenuReportFinancialRevenueSummary(client_locations_number_codes,monday,sunday)
        giftcard_reedemed_data = client.GetRevenuReportFinancialRevenueSummary_formatted(giftcard_reedemed_response)
        
        write_dictionary_to_xlshet(giftcard_reedemed_data,file_path,False)
        
        payment_response  = client.GetRevenuReportFinancialPaymentNew(client_locations_number_codes,monday,sunday)
        payment_data = client.GetRevenuReportFinancialPaymentNew_formatter(payment_response) 
        
        write_dictionary_to_xlshet(payment_data,file_path,False)


def conversion_rate_washify(arm_plans_sold,retail_car_count_monday_to_friday,retail_car_count_saturday_to_sunday):
    rate = 0.0
    
    try:
        rate = arm_plans_sold/sum([retail_car_count_monday_to_friday,retail_car_count_saturday_to_sunday])
        rate= rate*100
        rate = round(rate,2)
    
    except Exception as e:
        print(f"Exception in conversion_rate_washify() {e}")
    return rate

def generate_weekly_report(file_path, monday_date_str, friday_date_str, saturday_date_str, sunday_date_str):
    logger = logging.getLogger(__name__)
    logger.info("started main washify")
    "This will generate weekly report"
    final_report = {}
    
    try:
        client  = washifyClient()
        is_logged_in = client.check_login(proxy=None)

        if not is_logged_in:
            login = client.login(username=username,password=password,companyName=companyName,userType=userType)
            print(f"doing relogin : {login}")
            logger.info(f"doing relogin : {login}")
        client_locations = client.get_user_locations()
        
        # client_locations_number_codes =list(client_locations.values())
        print(f"client lcoations {client_locations.items()}")
        logger.info(f"client lcoations {client_locations.items()}")
        
        
        
        if client_locations:
            
            for location_name,location_code in client_locations.items():
                single_site_report = {}
                print(location_code,location_name)
                ## -----------Monday  to ---- Friday  ----------------##
                car_count_report_mon_fri_report = client.get_car_count_report([location_code],monday_date_str,friday_date_str)
                retail_revenue_summary_report_mon_fri = client.GetRevenuReportFinancialRevenueSummary([location_code],monday_date_str,friday_date_str)
                total_arm_plans1 = client.GetRevenuReportFinancialUnlimitedSales([location_code],monday_date_str,friday_date_str)
                retail_revenue_monday_fri = retail_revenue_summary_report_mon_fri.get("netPrice",0)
                total_revenue_monday_fri = retail_revenue_summary_report_mon_fri.get("total",0.0)
                labour_hours_monday_to_friday = car_count_report_mon_fri_report.get("totalhrs")
                
                car_count_monday_to_friday_cnt = car_count_report_mon_fri_report.get("car_count")
                print(car_count_report_mon_fri_report)
                print("retail revenue  report :",retail_revenue_summary_report_mon_fri)
                
                cars_per_labour_hour_monday_to_friday = round((car_count_monday_to_friday_cnt/labour_hours_monday_to_friday),2) if labour_hours_monday_to_friday !=0 else ""
                
                retail_car_count_monday_to_friday = car_count_report_mon_fri_report.get("retail_car_count")
                 
                single_site_report["car_count_monday_to_friday"]=car_count_monday_to_friday_cnt
               
                #single_site_report["arm_plans_reedemed_monday_to_friday_cnt"] = ""  #update
                single_site_report["retail_car_count_monday_to_friday"] = retail_car_count_monday_to_friday
                single_site_report["retail_revenue_monday_to_friday"] = retail_revenue_monday_fri
                single_site_report["total_revenue_monday_to_friday"] = total_revenue_monday_fri
                single_site_report["labour_hours_monday_to_friday"] = labour_hours_monday_to_friday
                single_site_report["cars_per_labour_hour_monday_to_friday"] = cars_per_labour_hour_monday_to_friday
                
                ## -----------Monday  to ---- Friday  ----------------##
                
                car_count_report_sat_sun_report = client.get_car_count_report([location_code],saturday_date_str, sunday_date_str)
                retail_revenue_summary_report_sat_sun = client.GetRevenuReportFinancialRevenueSummary([location_code],saturday_date_str, sunday_date_str)
                total_arm_plans2 = client.GetRevenuReportFinancialUnlimitedSales([location_code],saturday_date_str, sunday_date_str)
                retail_revenue_sat_sun = retail_revenue_summary_report_sat_sun.get("netPrice",0)
                total_revenue_sat_sun = retail_revenue_summary_report_sat_sun.get("total",0.0)
                #print(car_count_report_sat_sun_report)
                car_count_saturday_sunday_cnt = car_count_report_sat_sun_report.get("car_count",0)
                labour_hours_saturday_sunday = car_count_report_sat_sun_report.get("totalhrs")
                
                cars_per_labour_hour_saturday_sunday = round((car_count_saturday_sunday_cnt/labour_hours_saturday_sunday),2) if labour_hours_saturday_sunday != 0 else ""
                
                print("retail revenue  report :",retail_revenue_summary_report_mon_fri)
                
                retail_car_count_saturday_to_sunday = car_count_report_sat_sun_report.get("retail_car_count")
                
                single_site_report["car_count_saturday_sunday"]=car_count_report_sat_sun_report.get("car_count")
                #single_site_report["arm_plans_reedemed_saturday_sunday"] = "" #update
                single_site_report["retail_car_count_saturday_sunday"] = retail_car_count_saturday_to_sunday
                single_site_report["retail_revenue_saturday_sunday"] = retail_revenue_sat_sun
                single_site_report["total_revenue_saturday_sunday"] = total_revenue_sat_sun
                single_site_report["labour_hours_saturday_sunday"] = labour_hours_saturday_sunday
                single_site_report["cars_per_labour_hour_saturday_sunday"] = cars_per_labour_hour_saturday_sunday
                
                
                total_arm_planmembers_cnt = client.get_club_plan_members(location_code)
                
                #past week logis 
                past_4_week_day1,past_4_week_day2 = generate_past_4_weeks_days(monday_date_str)
                past_4_week_cnt_report = client.get_car_count_report([location_code],past_4_week_day1,past_4_week_day2)
                total_arm_plans3 = client.GetRevenuReportFinancialUnlimitedSales([location_code],past_4_week_day1,past_4_week_day2)
                past_4_weeks_reveunu_summary = client.GetRevenuReportFinancialRevenueSummary([location_code],past_4_week_day1,past_4_week_day2)
                past_4_week_cnt = past_4_week_cnt_report.get("car_count")
                past_4_weeks_retail_car_count = past_4_week_cnt_report.get("retail_car_count")
                
                past_4_week_conversion_rate  = conversion_rate_washify(total_arm_plans3,past_4_weeks_retail_car_count,0)
                past_4_weeks_total_revenue =past_4_weeks_reveunu_summary.get("total",0.0)
                
                single_site_report["past_4_week_car_cnt_mon_fri"]=0
                single_site_report["past_4_week_labour_hours_mon_fri"]=0
                
                single_site_report["past_4_week_car_cnt_sat_sun"]=0
                single_site_report["past_4_week_labour_hours_sat_sun"]=0

                single_site_report["past_4_week_retail_car_count_mon_fri"]=0
                single_site_report["past_4_week_retail_car_count_sat_sun"]=0

                single_site_report['past_4_week_retail_revenue_mon_fri'] = 0
                single_site_report['past_4_week_retail_revenue_sat_sun'] = 0

                single_site_report['past_4_week_total_revenue_mon_fri'] = 0
                single_site_report['past_4_week_total_revenue_sat_sun'] = 0

                full_weeks_lst = generate_past_4_week_days_full(monday_date_str)
                cnt=0
                for single_week in full_weeks_lst:
                    mon = single_week[0]
                    fri = single_week[1]
                    sat =single_week[2]
                    sun = single_week[3]
                    past_4_week_car_cnt_mon_fri_report = client.get_car_count_report([location_code],mon,fri)
                    past_4_week_retail_revenue_summary_report_mon_fri = client.GetRevenuReportFinancialRevenueSummary([location_code],mon,fri)

                    past_4_week_car_cnt_sat_sun_report = client.get_car_count_report([location_code],sat,sun)
                    past_4_week_retail_revenue_summary_report_sat_sun = client.GetRevenuReportFinancialRevenueSummary([location_code],sat,sun)
                    
                    past_4_week_car_cnt_mon_fri = past_4_week_car_cnt_mon_fri_report.get("car_count", 0)
                    past_4_week_car_cnt_sat_sun  = past_4_week_car_cnt_sat_sun_report.get("car_count", 0)
                    
                    past_4_week_labour_hours_mon_fri = past_4_week_car_cnt_mon_fri_report.get("totalhrs", 0)
                    past_4_week_labour_hours_sat_sun = past_4_week_car_cnt_sat_sun_report.get("totalhrs", 0)

                    past_4_week_retail_car_count_mon_fri = past_4_week_car_cnt_mon_fri_report.get('retail_car_count', 0)
                    past_4_week_retail_car_count_sat_sun = past_4_week_car_cnt_sat_sun_report.get('retail_car_count', 0)

                    past_4_week_retail_revenue_mon_fri = past_4_week_retail_revenue_summary_report_mon_fri.get('netPrice', 0)
                    past_4_week_retail_revenue_sat_sun = past_4_week_retail_revenue_summary_report_sat_sun.get('netPrice', 0)

                    past_4_week_total_revenue_mon_fri = past_4_week_retail_revenue_summary_report_mon_fri.get('total', 0)
                    past_4_week_total_revenue_sat_sun = past_4_week_retail_revenue_summary_report_sat_sun.get('total', 0)

                    single_site_report["past_4_week_car_cnt_mon_fri"] = single_site_report.get("past_4_week_car_cnt_mon_fri",0) +  past_4_week_car_cnt_mon_fri 
                    single_site_report["past_4_week_car_cnt_sat_sun"] = single_site_report.get("past_4_week_car_cnt_sat_sun",0) + past_4_week_car_cnt_sat_sun
                        
                    single_site_report["past_4_week_labour_hours_mon_fri"] = single_site_report.get("past_4_week_labour_hours_mon_fri",0) + past_4_week_labour_hours_mon_fri
                    single_site_report["past_4_week_labour_hours_sat_sun"] = single_site_report.get("past_4_week_labour_hours_sat_sun",0) + past_4_week_labour_hours_sat_sun

                    single_site_report["past_4_week_retail_car_count_mon_fri"] += past_4_week_retail_car_count_mon_fri
                    single_site_report[f"past_4_week_retail_car_count_mon_fri_week_{cnt+1}"] = past_4_week_retail_car_count_mon_fri
                    single_site_report["past_4_week_retail_car_count_sat_sun"] += past_4_week_retail_car_count_sat_sun
                    single_site_report[f"past_4_week_retail_car_count_sat_sun_week_{cnt+1}"] = past_4_week_retail_car_count_sat_sun
                    
                    single_site_report['past_4_week_retail_revenue_mon_fri'] += past_4_week_retail_revenue_mon_fri
                    single_site_report['past_4_week_retail_revenue_sat_sun'] += past_4_week_retail_revenue_sat_sun
                    
                    single_site_report[f'past_4_week_retail_revenue_mon_fri_week_{cnt+1}'] = past_4_week_retail_revenue_mon_fri
                    single_site_report[f'past_4_week_retail_revenue_sat_sun_week_{cnt+1}'] = past_4_week_retail_revenue_sat_sun

                    single_site_report['past_4_week_total_revenue_mon_fri'] += past_4_week_total_revenue_mon_fri
                    single_site_report['past_4_week_total_revenue_sat_sun'] += past_4_week_total_revenue_sat_sun
                    
                    print(f"past_4_week_retail_revenue_mon_fri : {past_4_week_retail_revenue_mon_fri}")
                    print(f"past_4_week_retail_revenue_sat_sun : {past_4_week_retail_revenue_sat_sun}")
                    
                    cnt+=1
                
                print(f"past week cnt :{past_4_week_cnt}")
                arm_plans_sold = sum([total_arm_plans1,total_arm_plans2])
                single_site_report["total_revenue"] = sum([total_revenue_monday_fri,total_revenue_sat_sun])
                single_site_report["arm_plans_sold_cnt"] = arm_plans_sold
                single_site_report["total_arm_planmembers_cnt"] = total_arm_planmembers_cnt
                single_site_report["conversion_rate"] = conversion_rate_washify(arm_plans_sold,retail_car_count_monday_to_friday,retail_car_count_saturday_to_sunday)
                single_site_report["past_4_week_cnt"] = past_4_week_cnt
                single_site_report["past_4_week_conversion_rate"] = past_4_week_conversion_rate
                single_site_report["past_4_weeks_total_revenue"]  = past_4_weeks_total_revenue
                single_site_report["past_4_weeks_arm_plans_sold_cnt"] = total_arm_plans3
                single_site_report["past_4_weeks_retail_car_count"]  = past_4_weeks_retail_car_count
               
                if "1631" in location_name: # 1631 E Jackson St
                    final_report["Getaway-Macomb"] = single_site_report
                elif "1821" in location_name:
                    final_report["Getaway-Morton"]=  single_site_report
                elif "2950" in location_name:
                    final_report["Getaway-Ottawa"] = single_site_report
                elif "4234" in location_name:
                    final_report["Getaway-Peru"]   = single_site_report
                    
                #final_report[location_name]=single_site_report
                
        
    except Exception as e:
        print(f"Exception generate_weeklyrepoer washify {e},{traceback.format_exc()}")
        logger.info(f"Exception generate_weeklyrepoer washify {e},{traceback.format_exc()}")
    logger.info(f"{final_report}")
    
    return final_report

if __name__=="__main__":
    # monday_date_str, friday_date_str, saturday_date_str, sunday_date_str =  get_week_dates()
        #testing dates 
    monday_date_str =  "06/03/2024"
    friday_date_str =  "06/07/2024"
    saturday_date_str = "06/08/2024"
    sunday_date_str  =  "06/09/2024"  #M/D/Y
    
    monday_date_str =  "07/22/2024"
    friday_date_str =  "07/26/2024"
    saturday_date_str = "07/27/2024"
    sunday_date_str  =  "07/28/2024"  #M/D/Y
    # print(monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    data = generate_weekly_report(file_path, monday_date_str, friday_date_str, saturday_date_str, sunday_date_str)
    
    print(data)
    
    with open("washify_data.json","w") as f:
        json.dump(data,f,indent=4)
    
    clinet = washifyClient()
    monday_str  = "06/03/2024"
    saturday= "06/09/2024"
    
    arm_plans =clinet.GetRevenuReportFinancialUnlimitedSales([90],monday_str,saturday)
    print(arm_plans)
    
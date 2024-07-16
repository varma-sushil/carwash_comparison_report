import json
import os 
import sys
import datetime
import requests
import locale
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta
from dotenv import load_dotenv


sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
# Set the locale to US English
locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

file_path="washift.xlsx"

load_dotenv()

class Config:
    # get environment variables as a dictionary
    env_vars           = os.environ
    PROXY_USER_NAME    = env_vars.get('PROXY_USER_NAME')
    PROXY_PASSWORD     = env_vars.get('PROXY_PASSWORD')
    PROXY_HOST         = env_vars.get('PROXY_HOST')
    PROXY_PORT         = env_vars.get('PROXY_PORT')
    PROXY_ZONE         = env_vars.get("PROXY_ZONE")
    IS_PROXY           =env_vars.get("IS_PROXY")



# Bright Data proxy credentials
username = Config.PROXY_USER_NAME
password = Config.PROXY_PASSWORD
zone = Config.PROXY_ZONE

# # Proxy configuration
# # proxy_host = 'zproxy.lum-superproxy.io'
proxy_host = Config.PROXY_HOST
proxy_port = Config.PROXY_PORT

# # Proxy URL format for datacenter proxy
proxy_url = f'http://{username}-zone-{zone}:{password}@{proxy_host}:{proxy_port}'

username = Config.PROXY_USER_NAME
password = Config.PROXY_PASSWORD
zone = Config.PROXY_ZONE

# # Proxy configuration
# # proxy_host = 'zproxy.lum-superproxy.io'
proxy_host = Config.PROXY_HOST
proxy_port = Config.PROXY_PORT
IS_PROXY = Config.IS_PROXY
# # Proxy URL format for datacenter proxy
proxies =None
# print(IS_PROXY)
if IS_PROXY:
    proxy_url = f'http://{username}-zone-{zone}:{password}@{proxy_host}:{proxy_port}'
    proxies={"http":proxy_url,"http":proxy_url}

def get_week_dates():
    # Get the current date
    today = datetime.today()
    
    # Find the current week's Monday date
    current_week_monday = today - timedelta(days=today.weekday())
    
    # Find the current week's Sunday date
    current_week_sunday = current_week_monday + timedelta(days=6)
    
    # Find the current week's Friday date
    current_week_friday = current_week_monday + timedelta(days=4)
    
    # Find the current week's Saturday date
    current_week_saturday = current_week_monday + timedelta(days=5)
    
    # Format the dates in mm/dd/yyyy format
    monday_date_str = current_week_monday.strftime("%m/%d/%Y")
    friday_date_str = current_week_friday.strftime("%m/%d/%Y")
    saturday_date_str = current_week_saturday.strftime("%m/%d/%Y")
    sunday_date_str = current_week_sunday.strftime("%m/%d/%Y")
    
    return monday_date_str, friday_date_str, saturday_date_str, sunday_date_str

def generate_past_4_weeks_days(date_str):
    # Convert the string date to a datetime object
    date_format = "%m/%d/%Y"
    monday = datetime.strptime(date_str, date_format)
    
    # Subtract one day
    one_day_before = monday - timedelta(days=1)
    four_weeks_before = monday - timedelta(days=(7*4) + 1)

    # Format the dates in "dd/mm/yyyy" format
    formatted_date = one_day_before.strftime("%m/%d/%Y")
    four_weeks_before_fmt = four_weeks_before.strftime("%m/%d/%Y")

    print("One day before the current date:", formatted_date)
    print("4 weeks before day :", four_weeks_before_fmt)

    return four_weeks_before_fmt, formatted_date
def append_dict_to_excel(file_path, data, num_lines,add_headers=True):
    try:
        # Try to load an existing workbook
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
    else:
        # If the file exists, get the active sheet
        sheet = workbook.active

    # Append blank lines
    for _ in range(num_lines):
        sheet.append([])

    if add_headers:
        # Append the header row with bold keys
        bold_font = Font(bold=True)
        header_row = list(data.keys())
        sheet.append(header_row)
        for cell in sheet[sheet.max_row]:
            cell.font = bold_font

    # Append the data row
    data_row = list(data.values())
    sheet.append(data_row)

    # Save the workbook
    workbook.save(file_path)


current_file_path = os.path.dirname(os.path.abspath(__file__))
# print(current_file_path)

cookies_path = os.path.join(current_file_path,"cookies")

cookie_file_path = os.path.join(cookies_path,"cookie.json")

data_path = os.path.join(current_file_path,"data")
print(cookies_path)

proxy_url=None

proxy = {
    "http":proxy_url,
    "https":proxy_url
}




class washifyClient():
    def __init__(self,) -> None:
        self.proxies = proxies

    def login(self,username,password,companyName,userType,proxy)->bool:
        """login fucntion

        Args:
            username (str):username
            password (str): password
            companyName (str): companyname
            userType (str): usertype
        """
        headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'en-US,en;q=0.9',
                'content-type': 'application/json',
                'origin': 'https://washifyapi.com:1000',
                'priority': 'u=1, i',
                'referer': 'https://washifyapi.com:1000/',
                'sec-ch-ua': '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Linux"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-site',
                'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
            }
        
        # json_data = {
        #         'username': 'Cameron',
        #         'password': 'Password1',
        #         'companyName': 'cleangetawayexpress',
        #         'userType': 'CWA',
        #     }
        
        json_data = {
                'username': username,
                'password': password,
                'companyName': companyName,
                'userType': userType,
            }
        
        try:
            response = requests.post('https://washifyapi.com:8298/api/AccountLogin/ValidateUserCredentials', headers=headers, json=json_data,proxies=self.proxies)
            if response.status_code==200:
                data = response.json()
                print(len(data['data']))
                mainCompanyID = data['data'].get("mainCompanyID")
                companyID = data['data'].get("companyID")
                serverID = data['data'].get("serverID")
                userRoleID = data['data'].get("userRoleID")
                userID= data['data'].get("userID")
                authToken= data['data'].get("authToken")
                userLocations= data['data'].get("userLocations")
                timeOffset= data['data'].get("timeOffset")
                # Store data in a dictionary
                extracted_data = {
                    "mainCompanyID": mainCompanyID,
                    "companyID": companyID,
                    "serverID": serverID,
                    "userRoleID": userRoleID,
                    "userID":userID,
                    "authToken":authToken,
                    "userLocations":userLocations,
                    "timeOffset":timeOffset

                }

                try:
                    with open(cookie_file_path, 'w') as json_file:
                        json.dump(extracted_data, json_file, indent=4)
                    print("Data saved to", cookie_file_path)
                    return True
                except Exception as e:
                    print("Failed to save data to JSON file. Error:", e)
        except Exception as e:
            print(f"Exception in login() {e}")
        return False

    def get_common_data(self):
        common_data={}

        try:
            with open(cookie_file_path,'r') as f:
                cookie_data = json.load(f)
            companyID = cookie_data.get("companyID")
            serverID = cookie_data.get("serverID")
            userRoleID = cookie_data.get("userRoleID")
            userID     = cookie_data.get("userID")
            authToken   = cookie_data.get("authToken")
            timeOffset  = cookie_data.get("timeOffset")

            if all([companyID,serverID,userRoleID,userID,authToken,timeOffset]):
                common_data['commonCompanyID'] = companyID
                common_data['commonServerID'] = serverID
                common_data['commonUserRoleID'] = userRoleID
                common_data['commonUserID'] = userID
                common_data['authKey'] = authToken
                common_data['commonTimeoffset'] =timeOffset
        except (FileExistsError,FileNotFoundError):
            print("Cookeis file not found !")
        
        except Exception as e:
            print(f"Exception : {e}")
        
        return common_data

    def check_login(self,proxy)->bool:
        "cheks is we can use previous login data or not"

        login_passed = False

        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        # with open(cookie_file_path,'r') as f:
        #     cookie_data = json.load(f)
        # companyID = cookie_data.get("companyID")
        # serverID = cookie_data.get("serverID")
        # userRoleID = cookie_data.get("userRoleID")
        # userID     = cookie_data.get("userID")
        # authToken   = cookie_data.get("authToken")
        # timeOffset  = cookie_data.get("timeOffset")

        json_data = {
            'CompanyID': 0,
            'UserLocations': '',
            'ServerID': 0,
            'UserRoleID': 0,
            'ID': 0,
            'CommonCompanySettings': self.get_common_data(),
        }
        try:
            response = requests.post('https://washifyapi.com:8298/api/UserRoles/GetRoleId', headers=headers, json=json_data,proxies=self.proxies)
            if response.status_code==200:
                msg = response.json().get("message")
                login_passed =True if msg=="Success" else False
        except Exception as e:
            print(f"Exception in check_login() : {e}")
        
        return login_passed

    def get_user_locations(self):
        data ={}

        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        json_data = {
            'companyID': 0,
            'ServerID': 0,
            'ID': 0,
            'CommonCompanySettings': self.get_common_data(),
        }
        try:
            response = requests.post('https://washifyapi.com:8298/api/CommonMethod/getUserLocations', headers=headers, json=json_data,proxies=self.proxies)
            if response.status_code==200:
                data = response.json().get("data")
                data = { location.get("locationName").split("-")[-1].strip() :location.get("locationID") for location in data if location.get("locationID")!=0}
        except Exception as e:
            print(f"Error in get_user_locations() : {e}")
        
        return data

    def get_car_count_report(self,location:list,StartDate,EndDate):
        result={}

        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }
        json_data = {
            'Locations': location,
            'StartDate': StartDate , #'06/24/2024'
            'EndDate': EndDate,  # '06/28/2024'
            'ReportBy': 'Day',
            'GroupAll': True,
            'CommonCompanySettings': self.get_common_data(),
        }
        try:
            response = requests.post('https://washifyapi.com:8298/api/Reports/GetCarCountReport', headers=headers, json=json_data)
            if response.status_code==200:
                data = response.json().get("data")
                #print(f"car count : {response} {response.json()}")
                for single_data in data:
                    car_count=single_data.get("carwashed",0)
                    unlimited_cars_washed = single_data.get("unilitedCarwashed",0)
                    staff_hours = single_data.get("totalhrs",0.0)
                    result["car_count"] = result.get("car_count",0)+car_count
                    
                    result["retail_car_count"]=result.get("retail_car_count",0)+(car_count-unlimited_cars_washed)
                    result['totalhrs'] = result.get("totalhrs",0.0)+staff_hours
                
                
        except Exception as e :
            print(f"Excpetion in get_car_count_report() {e}")
            
        return result
    
    def get_financal_revenue_summary(self):
        data =None
        headers = {
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'en-US,en;q=0.9',
        'content-type': 'application/json',
        'dnt': '1',
        'origin': 'https://washifyapi.com:1000',
        'priority': 'u=1, i',
        'referer': 'https://washifyapi.com:1000/',
        'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-site',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        json_data = {
            'Locations': [
                88,
                89,
                87,
                90,
            ],
            'StartDate': '06/10/2024 12:00 AM',
            'EndDate': '06/21/2024 11:59 PM',
            'LogOutDate': '06/21/2024 11:59 PM',
            'locationName': '',
            'ReportBy': '',
            'CommonCompanySettings': self.get_common_data(),
        }

        try:
            response = requests.post(
            'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialRevenueSummary',
            headers=headers,
            json=json_data,
            proxies=self.proxies
        )
            if response.status_code==200:
                data = response.json()
        
        except Exception as e:
            print(f"Error in get_financal_revenue_summary() {e}")

        return data
 
 
    
    def GetRevenuReportFinancialWashPackage(self,client_locations:list, monday,sunday):
        "WASH PACKAGES"
        data = None

        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        json_data = {
            'Locations': client_locations,
            'StartDate': f'{monday} 12:00 AM', #'06/10/2024 12:00 AM'
            'EndDate': f'{sunday} 11:59 PM', #'06/21/2024 11:59 PM'
            'LogOutDate': f'{sunday} 11:59 PM', #'06/21/2024 11:59 PM'
            'locationName': '',
            'ReportBy': '',
            'CommonCompanySettings': self.get_common_data(),
        }
        try:
            response = requests.post(
                'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialWashPackage',
                headers=headers,
                json=json_data,
                proxies=self.proxies
            )
            if response.status_code==200:
                data = response.json()
        except Exception as e:
            print(f" Exception in GetRevenuReportFinancialWashPackage()  {e}")
            
        return data

    def GetRevenuReportFinancialWashPackage_formatter(self,data):
        "WASH PACKAGES formatter"
        wash_packages_all = []
        Wash_Packages_Unlimited_total = 0
        Wash_Packages_Virtual_Wash_total = 0
        Wash_Packages_Non_Unlimited_total = 0
        Wash_Packages_Total_total  = 0
        Wash_Packages_Amount_total = 0
        Wash_Packages_Total_Amount_total = 0
        try:
            data = data.get("data")
            financialWashPackage = data.get("financialWashPackage")
            for wash_package in financialWashPackage:
                wash_package_structure={}
                
                cUnlimited  = wash_package.get("cUnlimited")
                Wash_Packages_Unlimited_total+=cUnlimited
                
                virtualWashNumber  = wash_package.get("virtualWashNumber")
                Wash_Packages_Virtual_Wash_total+=virtualWashNumber
                
                nonUnlimited    =   wash_package.get("nonUnlimited")
                Wash_Packages_Non_Unlimited_total+=nonUnlimited
                
                total    =   wash_package.get("total")
                Wash_Packages_Total_total += total
                
                price  =  wash_package.get("price")
                Wash_Packages_Amount_total +=price
                
                amount   =  wash_package.get("amount")
                Wash_Packages_Total_Amount_total +=amount
                
                wash_package_structure['Wash_Packages_ServiceName'] = wash_package.get("serviceName")
                wash_package_structure["Wash_Packages_Unlimited"]   = cUnlimited
                wash_package_structure["Wash_Packages_Virtual_Wash"] = virtualWashNumber
                wash_package_structure["Wash_Packages_Non_Unlimited"]  = nonUnlimited
                wash_package_structure["Wash_Packages_Total"]   =   total
                wash_package_structure["Wash_Packages_Amount"]   =    locale.currency(float(price), grouping=True)
                wash_package_structure["Wash_Packages_Total_Amount"]  =locale.currency(float(amount), grouping=True)
                # print(type(wash_package.get("amount")))
                wash_packages_all.append(wash_package_structure)
            
            wash_package_total_structure={
                "Wash_Packages_ServiceName":"Total:",
                "Wash_Packages_Unlimited":Wash_Packages_Unlimited_total,
                "Wash_Packages_Virtual_Wash":Wash_Packages_Virtual_Wash_total,
                "Wash_Packages_Non_Unlimited":Wash_Packages_Non_Unlimited_total,
                "Wash_Packages_Total":Wash_Packages_Total_total,
                "Wash_Packages_Amount":locale.currency(float(Wash_Packages_Amount_total), grouping=True),
                "Wash_Packages_Total_Amount":locale.currency(float(Wash_Packages_Total_Amount_total), grouping=True)   
            }
            
            wash_packages_all.append(wash_package_total_structure) #last ALl total row
                
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialWashPackage_formatter() {e}")
        return wash_packages_all



    def GetRevenuReportFinancialWashDiscounts(self,client_locations,monday,sunday):
        "DISCOUNTS"
        data=None
        


        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        json_data = {
            'Locations': client_locations,
            'StartDate': f'{monday} 12:00 AM',
            'EndDate': f'{sunday} 11:59 PM',
            'LogOutDate': f'{sunday} 11:59 PM',
            'locationName': '',
            'ReportBy': '',
            'CommonCompanySettings': self.get_common_data(),
        }

        try:
            response = requests.post(
                'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialWashDiscounts',
                headers=headers,
                json=json_data,
                proxies=self.proxies
            )

            if response.status_code==200:
                data = response.json()
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialWashDiscounts()  {e}")

        return data
    
    def  GetRevenuReportFinancialWashDiscounts_formatter(self,data):
        "DISCOUNTS formatter"
        discount_all =[]
        Wash_Packages_Discount_Number_total  = 0
        Wash_Packages_Discount_Service_Price_total = 0
        Wash_Packages_Discount_Total_Discount = 0
        
        try:
            data = data.get("data")
            financialWashDiscounts = data.get("financialWashDiscounts")
            
            for wash_discount in financialWashDiscounts:
                wash_discount_structure = {}
                
                number   =  wash_discount.get("number")
                Wash_Packages_Discount_Number_total+=number
                
                discountPrice  = wash_discount.get("discountPrice")
                Wash_Packages_Discount_Service_Price_total+=discountPrice
                
                totalAmt   =  wash_discount.get("totalAmt")
                Wash_Packages_Discount_Total_Discount+=totalAmt
                
                
                wash_discount_structure["Wash_Packages_Discount_ServiceName"] = wash_discount.get("discountName")
                wash_discount_structure["Wash_Packages_Discount_Number"]      = number
                wash_discount_structure["Wash_Packages_Discount_Service Price ($)"] = locale.currency(float(discountPrice), grouping=True)
                wash_discount_structure["Wash_Packages_Discount_Total Discount ($)"] = locale.currency(float(totalAmt), grouping=True)
                
                discount_all.append(wash_discount_structure)
                
            #appending Grand total
            discounts_all_total_structure = {
                "Wash_Packages_Discount_ServiceName":"Total:",
                "Wash_Packages_Discount_Number":Wash_Packages_Discount_Number_total,
                "Wash_Packages_Discount_Service Price ($)":locale.currency(float(Wash_Packages_Discount_Service_Price_total), grouping=True),
                "Wash_Packages_Discount_Total Discount ($)":locale.currency(float(Wash_Packages_Discount_Total_Discount), grouping=True)
            }
            discount_all.append(discounts_all_total_structure)
            
            
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialWashDiscounts_formatter() {e}")
          
        return discount_all

    def GetRevenuReportFinancialWashDiscounts_formatter2(self,data):
        "DISCOUNTS DISCOUNTS formatter2"
        discount_discount_all =[]
        DISCOUNTS_Number_total = 0
        DISCOUNTS_Price_total = 0
        DISCOUNTS_Revenue_total = 0
        
        try:
            
            data = data.get("data")

            financialWashDiscounts = data.get("financialWashDiscounts")     
            
            for wash_discount in financialWashDiscounts:
                
                number  = int(wash_discount.get("number",0))
                DISCOUNTS_Number_total+=number
                
                discountPrice  = float(wash_discount.get("discountPrice",0.0))
                DISCOUNTS_Price_total+=discountPrice
                
                totalAmt  =  float(wash_discount.get("totalAmt",0.0))
                DISCOUNTS_Revenue_total+=totalAmt
                
                discount_discount_structure = {}                                                             # Discount Discount
                discount_discount_structure["DISCOUNTS_Discount"] = wash_discount.get("discountName")
                discount_discount_structure["DISCOUNTS_Number"]      = number
                discount_discount_structure["DISCOUNTS_Price ($)"] = locale.currency(float(discountPrice), grouping=True)
                discount_discount_structure["DISCOUNTS_Revenue"] = locale.currency(float(totalAmt), grouping=True)
                
                discount_discount_all.append(discount_discount_structure)
                
            #Total table Discount discount
            discount_discount_total={
                "DISCOUNTS_Discount":"Total:",
                "DISCOUNTS_Number":DISCOUNTS_Number_total,
                "DISCOUNTS_Price ($)":locale.currency(float(DISCOUNTS_Price_total), grouping=True),
                "DISCOUNTS_Revenue":locale.currency(float(DISCOUNTS_Revenue_total), grouping=True)
            }
            
            discount_discount_all.append(discount_discount_total)
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialWashDiscounts_fromatter2 {e}")
            
        return discount_discount_all

    # def GetRevenuReportFinancialWashDiscounts(self):
        
    #     data = None
        
    #     headers = {
    #         'accept': 'application/json, text/plain, */*',
    #         'accept-language': 'en-US,en;q=0.9',
    #         'content-type': 'application/json',
    #         'dnt': '1',
    #         'origin': 'https://washifyapi.com:1000',
    #         'priority': 'u=1, i',
    #         'referer': 'https://washifyapi.com:1000/',
    #         'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
    #         'sec-ch-ua-mobile': '?0',
    #         'sec-ch-ua-platform': '"Windows"',
    #         'sec-fetch-dest': 'empty',
    #         'sec-fetch-mode': 'cors',
    #         'sec-fetch-site': 'same-site',
    #         'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
    #     }

    #     json_data = {
    #         'Locations': [
    #             88,
    #             89,
    #             87,
    #             90,
    #         ],
    #         'StartDate': '06/10/2024 12:00 AM',
    #         'EndDate': '06/21/2024 11:59 PM',
    #         'LogOutDate': '06/21/2024 11:59 PM',
    #         'locationName': '',
    #         'ReportBy': '',
    #         'CommonCompanySettings': self.get_common_data(),
    #     }
    #     try:
    #         response = requests.post(
    #             'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialWashDiscounts',
    #             headers=headers,
    #             json=json_data,
    #         )
            
    #         if response.status_code==200:
    #             data = response.json()
    #     except Exception as e:
    #         print(f"Exception in GetRevenuReportFinancialWashDiscounts() {e}")

    #     return data
   
 
 
   
    
    def GetRevenuReportFinancialPackagesDiscount(self,client_locations,monday,sunday):
        "WASH EXTRAS"
        data = None

        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        json_data = {
            'Locations': client_locations,
            'StartDate': f'{monday} 12:00 AM',
            'EndDate': f'{sunday} 11:59 PM',
            'LogOutDate': f'{sunday} 11:59 PM',
            'locationName': '',
            'ReportBy': '',
            'CommonCompanySettings': self.get_common_data(),
        }

        try:
            response = requests.post(
                'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialPackagesDiscount',
                headers=headers,
                json=json_data,
                proxies=self.proxies
            )
            if response.status_code==200:
                data = response.json()
        except Exception as e:
            print(f"Exception on GetRevenuReportFinancialPackagesDiscount() {e}")
            
        return data

    def GetRevenuReportFinancialPackagesDiscount_formatter(self,data):
        "WASH EXTRAS formatter"
        wash_extras_all = []
        Wash_Extras_Number_total=0
        Wash_Extras_Amount_total = 0
        Wash_Extras_Total_Amount_total = 0
        
        try:
            data = data.get("data")
            financialPackagesDiscount = data.get("financialPackagesDiscount") 
            
            for wash_extra in financialPackagesDiscount:
                wash_extra_structure={}
                
                number = int(wash_extra.get("number",0))
                # print("number:",type(number))
                Wash_Extras_Number_total+=number
                
                servicePrice = float(wash_extra.get("servicePrice",0.0))
                Wash_Extras_Amount_total+=servicePrice
                
                totalAmount = float(wash_extra.get("totalAmount",0.0))
                Wash_Extras_Total_Amount_total+=totalAmount
                
                wash_extra_structure["Wash_Extras_ServiceName"] = wash_extra.get("serviceName")
                wash_extra_structure["Wash_Extras_Number"]      = number
                wash_extra_structure["Wash_Extras_Amount ($)"]  = locale.currency(float(servicePrice), grouping=True)
                wash_extra_structure["Wash_Extras_Total_Amount ($)"] = locale.currency(float(totalAmount), grouping=True)
                
                wash_extras_all.append(wash_extra_structure)
            #adding final total value 
            wash_extras_total_structure = {
                "Wash_Extras_ServiceName":"Total:",
                "Wash_Extras_Number":Wash_Extras_Number_total,
                "Wash_Extras_Amount ($)":locale.currency(float(Wash_Extras_Amount_total), grouping=True),
                "Wash_Extras_Total_Amount ($)":locale.currency(float(Wash_Extras_Total_Amount_total), grouping=True)
            }   
            
            wash_extras_all.append(wash_extras_total_structure)
                
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialPackagesDiscount_formatter()  {e}")

        return wash_extras_all





    def GetRevenuReportFinancialUnlimitedSales(self,client_locations,monday,sunday):
        "UNLIMITED SALES clubplans sold"
        data = None
        sale_count_total=0
        


        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        json_data = {
            'Locations': client_locations,
            'StartDate': f'{monday} 12:00 AM',
            'EndDate': f'{sunday} 11:59 PM',
            'LogOutDate': f'{sunday} 11:59 PM',
            'locationName': '',
            'ReportBy': '',
            'CommonCompanySettings':self.get_common_data(),
        }

        try:
            response = requests.post(
                'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialUnlimitedSales',
                headers=headers,
                json=json_data,
                proxies=self.proxies
            )
            if response.status_code==200:
                data = response.json().get("data")
                unlimited_sales = data.get("financialUnlimitedSales")
                sale_count_total = 0
                for unlimited_sale in unlimited_sales:
                    unlimite_sale_type = unlimited_sale.get("unlimited_Sales")
                    
                    if unlimite_sale_type in ["New Sales","Re Signups"]:
                        sale_cnt = unlimited_sale.get("number",0)
                        sale_count_total += sale_cnt
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialUnlimitedSales() {e}")

        return sale_count_total

    def GetRevenuReportFinancialUnlimitedSales_formatter(self,data):
        "UNLIMITED SALES formatter"
        
        unlimited_sales_all =[]
        
        Unlimited_Sales_Number_total = 0
        Unlimited_Sales_Revenue_total =0
        try:
            data = data.get("data")
            financialUnlimitedSales  = data.get("financialUnlimitedSales")
            
            for sales_data in financialUnlimitedSales:
                sales_data_structure ={}
                # print(sales_data)
                number  = int(sales_data.get("number",0))
                Unlimited_Sales_Number_total+=number
                
                price  =  float(sales_data.get("price",0.0))
                Unlimited_Sales_Revenue_total+=price
                
                sales_data_structure["Unlimited_Sales"] = sales_data.get("unlimited_Sales")
                sales_data_structure["Unlimited_Sales_Service"] = sales_data.get("serviceName")
                sales_data_structure["Unlimited_Sales_Number"]  = number
                sales_data_structure["Unlimited_Sales_Revenue ($)"] = locale.currency(float(price), grouping=True) 
                
                unlimited_sales_all.append(sales_data_structure)
              
            #final table total
            unlimited_sales__total = {
                "Unlimited_Sales":"Total:",
                "Unlimited_Sales_Service":"",
                "Unlimited_Sales_Number":Unlimited_Sales_Number_total,
                "Unlimited_Sales_Revenue ($)":locale.currency(float(Unlimited_Sales_Revenue_total), grouping=True)
            }    
            
            unlimited_sales_all.append(unlimited_sales__total)
            
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialUnlimitedSales_formatter() {e}")
            
        return unlimited_sales_all

 
 
 

    def GetRevenuReportFinancialGiftcardsale(self,client_locations,monday,sunday):
        "GIFT CARD SALES"
        data = None
        


        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        json_data = {
            'Locations': client_locations,
            'StartDate': f'{monday} 12:00 AM',
            'EndDate': f'{sunday} 11:59 PM',
            'LogOutDate': f'{sunday} 11:59 PM',
            'locationName': '',
            'ReportBy': '',
            'CommonCompanySettings': self.get_common_data(),
        }

        try:
            response = requests.post(
                'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialGiftcardsale',
                headers=headers,
                json=json_data,
                proxies=self.proxies
            )
            
            if response.status_code==200:
                data = response.json()
        except Exception as e:
            print("Exception in GetRevenuReportFinancialGiftcardsale() {e}")
            
        return data

    def GetRevenuReportFinancialGiftcardsale_formatter(self,data):
        "GIFT CARD SALES formatter "
        
        gift_card_sale_all =[]
        GIFT_CARD_SALESr_Amount_total = 0
        try:
            data = data.get("data")

            financialGiftcardsale = data.get("financialGiftcardsale")
            
            for gift_card in financialGiftcardsale:
                gift_card_sale_structure = {}
                
                price   =  float(gift_card.get("price",0.0))
                GIFT_CARD_SALESr_Amount_total+=price
                
                gift_card_sale_structure["GIFT_CARD_SALES_DATE"] = gift_card.get("date")  #giftcarsd sales
                gift_card_sale_structure["GIFT_CARD_SALES_TIME"] = gift_card.get("time")
                gift_card_sale_structure["GIFT_CARD_SALES_Card_Number"] = gift_card.get("coupanNumber")
                gift_card_sale_structure["GIFT_CARD_SALESr_Amount ($)"] = locale.currency(float(price), grouping=True)
                gift_card_sale_structure["GIFT_CARD_SALES_Source"]      = gift_card.get("transactionFrom")
                
                gift_card_sale_all.append(gift_card_sale_structure)
                
            #final total table 
            giftcard_sale_total = {
                "GIFT_CARD_SALES_DATE":"Total:",
                "GIFT_CARD_SALES_TIME":"",
                "GIFT_CARD_SALES_Card_Number":"",
                "GIFT_CARD_SALESr_Amount ($)":locale.currency(float(GIFT_CARD_SALESr_Amount_total), grouping=True),
                "GIFT_CARD_SALES_Source":""
            }
            gift_card_sale_all.append(giftcard_sale_total)
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialGiftcardsale_formatter() {e}")
            
        return gift_card_sale_all



    # def GetRevenuReportFinancialWashDiscounts(self):
    #     "DISCOUNTS"
    #     data = None
        

    #     headers = {
    #         'accept': 'application/json, text/plain, */*',
    #         'accept-language': 'en-US,en;q=0.9',
    #         'content-type': 'application/json',
    #         'dnt': '1',
    #         'origin': 'https://washifyapi.com:1000',
    #         'priority': 'u=1, i',
    #         'referer': 'https://washifyapi.com:1000/',
    #         'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
    #         'sec-ch-ua-mobile': '?0',
    #         'sec-ch-ua-platform': '"Windows"',
    #         'sec-fetch-dest': 'empty',
    #         'sec-fetch-mode': 'cors',
    #         'sec-fetch-site': 'same-site',
    #         'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
    #     }

    #     json_data = {
    #         'Locations': [
    #             88,
    #             89,
    #             87,
    #             90,
    #         ],
    #         'StartDate': '06/10/2024 12:00 AM',
    #         'EndDate': '06/21/2024 11:59 PM',
    #         'LogOutDate': '06/21/2024 11:59 PM',
    #         'locationName': '',
    #         'ReportBy': '',
    #         'CommonCompanySettings': {
    #             'commonCompanyID': 'WWhtk4RrRQ8KymvwT0BMaw==',
    #             'commonServerID': '+iMYUwYx079az+3TrcOsag==',
    #             'commonUserRoleID': 'Hud67uIVSx9QDTsSrUCfjg==',
    #             'commonUserID': 'bKN5Lonb9871/yWYGAEuAQ==',
    #             'authKey': 'jF9inrLiZCOgWlXjXi0Z13m4qUvjpsZcQujn6kXp6iE=',
    #             'commonTimeoffset': '120',
    #         },
    #     }

    #     try:
    #         response = requests.post(
    #             'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialWashDiscounts',
    #             headers=headers,
    #             json=json_data,
    #         )
    #         if response.status_code==200:
    #             data = response.json()
    #     except Exception as e:
    #         print(f"Exception in GetRevenuReportFinancialWashDiscounts() {e}")
            
    #     return data

    def GetRevenuReportFinancialRevenueSummary(self,client_locations,monday,sunday):
        "GIFT CARD REDEEMED netPrice total(totalrevenue)"
        
        result = {}
        


        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        json_data = {
            'Locations': client_locations,
            'StartDate': f'{monday} 12:00 AM',
            'EndDate': f'{sunday} 11:59 PM',
            'LogOutDate': f'{sunday} 11:59 PM',
            'locationName': '',
            'ReportBy': '',
            'CommonCompanySettings': self.get_common_data(),
        }
        try:
            response = requests.post(
                'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialRevenueSummary',
                headers=headers,
                json=json_data,
                proxies=self.proxies
            )
            
            if response.status_code==200:
                data = response.json().get("data")
                finanacial_sumamry= data.get("financialRevenueSummary")[0]
                financialReportOther = data.get("financialReportOther")[0]
                result['netPrice'] = finanacial_sumamry.get("netPrice")
                result["total"]    =financialReportOther.get("total")
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialRevenueSummary() {e}")

        return result 
    
    def GetRevenuReportFinancialRevenueSummary_formatted(self,data):
        "GIFT CARD REDEEMED formatter"
        
        reedemed_giftcard_all =[]
        GIFT_CARD_REDEEMED_Amount_total = 0
        try:
            data = data.get("data")

            financialGiftcardRedeemed = data.get("financialGiftcardRedeemed")
            
            
            for reedemed_giftcard in financialGiftcardRedeemed:
                reedemed_giftcard_structure = {}
                # print(reedemed_giftcard)
                
                price = float(reedemed_giftcard.get("price",0.0))
                GIFT_CARD_REDEEMED_Amount_total+=price
                
                reedemed_giftcard_structure["GIFT_CARD_REDEEMED_DATE"] = reedemed_giftcard.get("date")
                reedemed_giftcard_structure["GIFT_CARD_REDEEMED_TIME"] = reedemed_giftcard.get("time")
                reedemed_giftcard_structure["GIFT_CARD_REDEEMED_CARD_NUMBER"] = reedemed_giftcard.get("coupanNumber")
                reedemed_giftcard_structure["GIFT_CARD_REDEEMED_Amount ($)"]  = locale.currency(float(price), grouping=True)
                
                reedemed_giftcard_all.append(reedemed_giftcard_structure)
            #reedem giftcard total
            
            reedem_total_structure ={
                "GIFT_CARD_REDEEMED_DATE":"Total:",
                "GIFT_CARD_REDEEMED_TIME":"",
                "GIFT_CARD_REDEEMED_CARD_NUMBER":"",
                "GIFT_CARD_REDEEMED_Amount ($)":locale.currency(float(GIFT_CARD_REDEEMED_Amount_total), grouping=True)
            }    
            reedemed_giftcard_all.append(reedem_total_structure)
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialRevenueSummary_formatter()  {e}")
            
        return reedemed_giftcard_all





    def GetRevenuReportFinancialPaymentNew(self,client_locations,monday,sunday):
        "Payment"
        data = None

        headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/json',
            'dnt': '1',
            'origin': 'https://washifyapi.com:1000',
            'priority': 'u=1, i',
            'referer': 'https://washifyapi.com:1000/',
            'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        }

        json_data = {
            'Locations': client_locations,
            'StartDate': f'{monday} 12:00 AM',
            'EndDate': f'{sunday} 11:59 PM',
            'LogOutDate': f'{sunday} 11:59 PM',
            'locationName': '',
            'ReportBy': '',
            'CommonCompanySettings': self.get_common_data(),
        }

        try:
            response = requests.post(
                'https://washifyapi.com:8298/api/Reports/GetRevenuReportFinancialPaymentNew',
                headers=headers,
                json=json_data,
                proxies=self.proxies
            )
            if response.status_code==200:
                data = response.json()
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialPaymentNew() {e}")
            
        return data

    def GetRevenuReportFinancialPaymentNew_formatter(self,data):
        "Payment formatter"
        
        payment_data_all =[]
        Payment_Cash_total = 0
        Payment_Credit_Card_total = 0
        Payment_Check_total = 0
        Payment_Invoice_total = 0
        Payment_ACH_total = 0
        Payment_Total_total = 0
        try:
            data = data.get('data')
            financialPaymentNew = data.get("financialPaymentNew")
            
            for payment in financialPaymentNew:
                payment_structure = {}
                # print(payment)

                cash = float(payment.get("cash",0.0))
                Payment_Cash_total+=cash
                
                creditCard = float(payment.get("creditCard",0.0))
                Payment_Credit_Card_total +=creditCard
                
                checkpayment  = float(payment.get("checkpayment",0.0))
                Payment_Check_total+=checkpayment
                
                invoiceCustomer = float(payment.get("invoiceCustomer",0.0))
                Payment_Invoice_total+=invoiceCustomer
                
                
                ach = float(payment.get("ach",0.0))
                Payment_ACH_total +=ach
                
                total_payment = sum([cash,creditCard,checkpayment,invoiceCustomer,ach])
                Payment_Total_total+=total_payment
                
                payment_structure["Payment_Location"] = payment.get("locationName")
                payment_structure["Payment_Cash"]     = locale.currency(float(cash), grouping=True)
                payment_structure["Payment_Credit_Card"]  = locale.currency(float(creditCard), grouping=True)
                payment_structure["Payment_Check"]     = locale.currency(float(checkpayment), grouping=True)
                payment_structure["Payment_Invoice"]   = locale.currency(float(invoiceCustomer), grouping=True)
                payment_structure["Payment_ACH"]       = locale.currency(float(ach), grouping=True)
                payment_structure["Payment_Total ($)"] = locale.currency(float(total_payment), grouping=True)  ##payment
                
                payment_data_all.append(payment_structure)
                  
            #total payments row 
            payment_total_structure = {
                "Payment_Location":"Total Payments:",
                "Payment_Cash":locale.currency(float(Payment_Cash_total), grouping=True),
                "Payment_Credit_Card":locale.currency(float(Payment_Credit_Card_total), grouping=True),
                "Payment_Check":locale.currency(float(Payment_Check_total), grouping=True),
                "Payment_Invoice":locale.currency(float(Payment_Invoice_total), grouping=True),
                "Payment_ACH":locale.currency(float(Payment_ACH_total), grouping=True),
                "Payment_Total ($)":locale.currency(float(Payment_Total_total), grouping=True)
            }
            
            payment_data_all.append(payment_total_structure)
                  
        except Exception as e:
            print(f"Exception in GetRevenuReportFinancialPaymentNew_formatter()  {e}")


        return payment_data_all

   
    def get_club_plan_members(self,locationcode:int):
        "This function will give toital club plan memebers based on user lcoation"
        total_plan_members = 0
       
        try:

            headers = {
                'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
                'DNT': '1',
                'sec-ch-ua-mobile': '?0',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
                'Content-Type': 'application/json',
                'Accept': 'application/json, text/plain, */*',
                'Referer': 'https://washifyapi.com:1000/',
                'sec-ch-ua-platform': '"Windows"',
            }

            json_data = {
                'CompanyID': 0,
                'UserLocations': locationcode,
                'ServerID': 0,
                'UserRoleID': 0,
                'ID': 0,
                'CommonCompanySettings': self.get_common_data(),
            }

            response = requests.post(
                'https://washifyapi.com:8298/api/Dashboard/DashBoardDailyStatisticList',
                headers=headers,
                json=json_data,
                proxies=self.proxies
            )
            if response.status_code==200:
                data = response.json().get("data")
                dailyStatisticList = data.get("dailyStatisticList")[0]
                total_plan_members = dailyStatisticList.get("vehicles")
        except Exception as e:
            print(f"Exception in get_club_plan_mberbers() {e}")

        return total_plan_members
    
    
if __name__=="__main__":
    
# json_data = {
        #         'username': 'Cameron',
        #         'password': 'Password1',
        #         'companyName': 'cleangetawayexpress',
        #         'userType': 'CWA',
        #     }
    username = 'Cameron'
    password = 'Password1'
    companyName = 'cleangetawayexpress'
    userType = 'CWA'
    client  = washifyClient()
    # login = client.login(username=username,password=password,
    #                     companyName=companyName,userType=userType,proxy=proxy)
    # print(f"login check : {login}")
    is_logged_in = client.check_login(proxy=proxy)
    client_locations = client.get_user_locations()
    print(f"is logged in :{is_logged_in}")
    print(f"user lcoations : {client_locations}")
    if client_locations:
        # for location_name,location_id in client_locations.items():
        #     print(location_name,":",client.get_car_count_report([location_id,]))
        print(client_locations)
        
        # response_reneue = client.GetRevenuReportFinancialWashPackage()
        # with open(f"{data_path}\GetRevenuReportFinancialWashPackage.json","w") as f:  #for wah package
        #     json.dump(response_reneue,f,indent=4)
        
        # response_reneue = client.GetRevenuReportFinancialWashDiscounts()
        # with open(f"{data_path}\GetRevenuReportFinancialWashDiscounts.json","w") as f:  #for Discount
        #     json.dump(response_reneue,f,indent=4)
            
        # response_reneue = client.GetRevenuReportFinancialPackagesDiscount()
        # with open(f"{data_path}\GetRevenuReportFinancialPackagesDiscount.json","w") as f:  #for Discount
        #     json.dump(response_reneue,f,indent=4)
        # response_reneue = client.GetRevenuReportFinancialUnlimitedSales()
        # with open(f"{data_path}\GetRevenuReportFinancialUnlimitedSales.json","w") as f:  #for Unlimited sales
        #     json.dump(response_reneue,f,indent=4)
            
        # response_reneue = client.GetRevenuReportFinancialGiftcardsale()
        # with open(f"{data_path}\GetRevenuReportFinancialGiftcardsale.json","w") as f:  #for Gift card sale
        #     json.dump(response_reneue,f,indent=4)
        # response_reneue = client.GetRevenuReportFinancialRevenueSummary()
        # with open(f"{data_path}\GetRevenuReportFinancialRevenueSummary.json","w") as f:  #for Unlimited sales
        #     json.dump(response_reneue,f,indent=4)
        
        # response_reneue = client.GetRevenuReportFinancialPaymentNew()
        # with open(f"{data_path}\GetRevenuReportFinancialPaymentNew.json","w") as f:  #for Payment
        #     json.dump(response_reneue,f,indent=4)
            
            
        ## Formatter logic
        # wash_packages_response = client.GetRevenuReportFinancialWashPackage()
        # wash_packages_data = client.GetRevenuReportFinancialWashPackage_formatter(wash_packages_response)  #first table 
        
        # for index,data in enumerate(wash_packages_data):
        #     if index == 0:
        #         append_dict_to_excel(file_path,data,0)
        #     else:
        #         append_dict_to_excel(file_path,data,0,False)
        #     # print(data)
        # # print(json.dumps(formatted_response,indent=4))

        # wash_package_discount_response = client.GetRevenuReportFinancialWashDiscounts()
        # washpack_discount_data = client.GetRevenuReportFinancialWashDiscounts_formatter(wash_package_discount_response)  #secound table 
        # # print(json.dumps(formatted_response,indent=4))
        
        # for index,data in enumerate(washpack_discount_data):
        #     if index == 0:
        #         append_dict_to_excel(file_path,data,2)
        #     else:
        #         append_dict_to_excel(file_path,data,0,False)

        # wash_extra_response = client.GetRevenuReportFinancialPackagesDiscount()
        # wash_extra_data = client.GetRevenuReportFinancialPackagesDiscount_formatter(wash_extra_response)  #3rd table 
        # # print(json.dumps(formatted_response,indent=4))
        # for index,data in enumerate(wash_extra_data):
        #     if index == 0:
        #         append_dict_to_excel(file_path,data,2)
        #     else:
        #         append_dict_to_excel(file_path,data,0,False)

        # unlimited_sales_response = client.GetRevenuReportFinancialUnlimitedSales()
        # unlimited_sales_data  = client.GetRevenuReportFinancialUnlimitedSales_formatter(unlimited_sales_response) #unlimited sales
        # # print(unlimited_sales_data)
        # # print(json.dumps(formatted_response,indent=4))
        # for index,data in enumerate( unlimited_sales_data):
        #     if index == 0:
        #         append_dict_to_excel(file_path,data,2)
        #     else:
        #         append_dict_to_excel(file_path,data,0,False)
            
        # giftcard_sales_response = client.GetRevenuReportFinancialGiftcardsale()
        # giftcards_sales_data = client.GetRevenuReportFinancialGiftcardsale_formatter(giftcard_sales_response)  #4rd table  gift card sale
        # # print(formatted_response) 
        # # print(json.dumps(formatted_response,indent=4))
        
        # for index,data in enumerate(giftcards_sales_data):
        #     if index == 0:
        #         append_dict_to_excel(file_path,data,2)
        #     else:
        #         append_dict_to_excel(file_path,data,0,False)
        
        
        # discount_discount_response = client.GetRevenuReportFinancialWashDiscounts()
        # discount_discount_data = client.GetRevenuReportFinancialWashDiscounts_formatter2(discount_discount_response)  #5rd table    Discount discount
        # # print(json.dumps(formatted_response,indent=4))
        
        # for index,data in enumerate(discount_discount_data):
        #     if index == 0:
        #         append_dict_to_excel(file_path,data,2)
        #     else:
        #         append_dict_to_excel(file_path,data,0,False)
        
        
        # giftcard_reedemed_response = client.GetRevenuReportFinancialRevenueSummary()
        # giftcard_reedemed_data = client.GetRevenuReportFinancialRevenueSummary_formatted(giftcard_reedemed_response)  #6rd table   Discount discount
        # # print(json.dumps(formatted_response,indent=4))
        # for index,data in enumerate(giftcard_reedemed_data):
        #     if index == 0:
        #         append_dict_to_excel(file_path,data,2)
        #     else:
        #         append_dict_to_excel(file_path,data,0,False)
        
        
        
        # # response = client.GetRevenuReportFinancialRevenueSummary()   #duplicate
        # # formatted_response = client.GetRevenuReportFinancialRevenueSummary_formatted(response)  #8rd table  gift card reedem
        # # print(json.dumps(formatted_response,indent=4))
        
        
        # payment_response  = client.GetRevenuReportFinancialPaymentNew()
        # payment_data = client.GetRevenuReportFinancialPaymentNew_formatter(payment_response)  #8rd table payment location
        # # print(json.dumps(payment_data,indent=4))
        # for index,data in enumerate(payment_data):
        #     if index == 0:
        #         append_dict_to_excel(file_path,data,2)
        #     else:
        #         append_dict_to_excel(file_path,data,0,False)
        
        
        # monday_date_str, friday_date_str, saturday_date_str, sunday_date_str =  get_week_dates()
        # data = client.get_car_count_report([88],monday_date_str, friday_date_str)
        # print(data)
        
        print(client.get_club_plan_members(90))
        
## need to chanege site locations dynamic and and time stamp also dynamic and need to use type casting and need to write xl conversion code